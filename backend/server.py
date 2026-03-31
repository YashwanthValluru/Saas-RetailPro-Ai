from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, Request, HTTPException, Depends, Query
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from starlette.responses import JSONResponse, StreamingResponse
import os
import logging
import uuid
import secrets
import hashlib
import base64
import io
import json
import csv
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
from bson import ObjectId

import bcrypt
import jwt
import pyotp
import qrcode
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from urllib.parse import quote as url_quote

# PDF & Excel
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook, load_workbook
import barcode as python_barcode
from barcode.writer import ImageWriter

# ─── Config ───────────────────────────────────────────────────────
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGORITHM = "HS256"
MFA_KEY = os.environ.get('MFA_ENCRYPTION_KEY', 'default_key_32chars_here_ok1234')

app = FastAPI()
api_router = APIRouter(prefix="/api")

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter

@app.exception_handler(RateLimitExceeded)
async def rate_limit_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(status_code=429, content={"detail": "Too many requests. Please try again later."})

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ─── In-Memory Tenant-Scoped Cache ────────────────────────────────
import time as _time
from collections import OrderedDict

class TenantCache:
    """Thread-safe, tenant-scoped in-memory cache with TTL and max size eviction."""
    def __init__(self, max_size=2000, default_ttl=60):
        self._cache = OrderedDict()
        self._max_size = max_size
        self._default_ttl = default_ttl
        self._hits = 0
        self._misses = 0

    def get(self, tenant_id: str, key: str):
        cache_key = f"{tenant_id}:{key}"
        if cache_key in self._cache:
            value, expiry = self._cache[cache_key]
            if _time.time() < expiry:
                self._cache.move_to_end(cache_key)
                self._hits += 1
                return value
            else:
                del self._cache[cache_key]
        self._misses += 1
        return None

    def set(self, tenant_id: str, key: str, value, ttl: int = None):
        cache_key = f"{tenant_id}:{key}"
        ttl = ttl or self._default_ttl
        self._cache[cache_key] = (value, _time.time() + ttl)
        self._cache.move_to_end(cache_key)
        # Evict oldest if over max_size
        while len(self._cache) > self._max_size:
            self._cache.popitem(last=False)

    def invalidate(self, tenant_id: str, key: str = None):
        if key:
            cache_key = f"{tenant_id}:{key}"
            self._cache.pop(cache_key, None)
        else:
            keys_to_delete = [k for k in self._cache if k.startswith(f"{tenant_id}:")]
            for k in keys_to_delete:
                del self._cache[k]

    def invalidate_prefix(self, tenant_id: str, prefix: str):
        full_prefix = f"{tenant_id}:{prefix}"
        keys_to_delete = [k for k in self._cache if k.startswith(full_prefix)]
        for k in keys_to_delete:
            del self._cache[k]

    def clear(self):
        self._cache.clear()

    @property
    def stats(self):
        return {"size": len(self._cache), "hits": self._hits, "misses": self._misses,
                "hit_rate": round(self._hits / max(self._hits + self._misses, 1) * 100, 1)}

# Global cache instances — different TTLs per data type
product_cache = TenantCache(max_size=5000, default_ttl=45)       # Products: 45s
category_cache = TenantCache(max_size=500, default_ttl=300)      # Categories: 5min
barcode_cache = TenantCache(max_size=10000, default_ttl=86400)   # Barcode lookups: 24hr
dashboard_cache = TenantCache(max_size=200, default_ttl=120)     # Dashboard stats: 2min
customer_cache = TenantCache(max_size=2000, default_ttl=60)      # Customers: 1min

# ─── Pydantic Models ─────────────────────────────────────────────

class RegisterRequest(BaseModel):
    email: str
    password: str
    name: str
    shop_name: str
    business_type: str = "general"

class LoginRequest(BaseModel):
    email: str
    password: str

class MFAVerifyRequest(BaseModel):
    temp_token: str
    otp_code: str

class MFABackupLoginRequest(BaseModel):
    temp_token: str
    backup_code: str

class MFAEnableRequest(BaseModel):
    otp_code: str

class ProductCreate(BaseModel):
    name: str
    sku: str = ""
    barcode: str = ""
    category: str = ""
    price: float = 0.0
    cost_price: float = 0.0
    stock: int = 0
    low_stock_threshold: int = 10
    unit: str = "pcs"
    batch_number: str = ""
    expiry_date: Optional[str] = None
    description: str = ""
    hsn_code: str = ""
    gst_rate: float = 0.0

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    sku: Optional[str] = None
    barcode: Optional[str] = None
    category: Optional[str] = None
    price: Optional[float] = None
    cost_price: Optional[float] = None
    stock: Optional[int] = None
    low_stock_threshold: Optional[int] = None
    unit: Optional[str] = None
    batch_number: Optional[str] = None
    expiry_date: Optional[str] = None
    description: Optional[str] = None
    hsn_code: Optional[str] = None
    gst_rate: Optional[float] = None

class InvoiceItem(BaseModel):
    product_id: str
    name: str
    quantity: int
    price: float
    gst_rate: float = 0.0

class InvoiceCreate(BaseModel):
    customer_name: str = "Walk-in Customer"
    customer_phone: str = ""
    customer_id: Optional[str] = None
    items: List[InvoiceItem]
    discount: float = 0.0
    payment_method: str = "cash"
    notes: str = ""
    device_source: str = "desktop"

class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str = "STAFF"

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None

class CheckoutRequest(BaseModel):
    plan_id: str
    origin_url: str

class StockAdjust(BaseModel):
    product_id: str
    adjustment: int
    reason: str = ""

# ─── Purchase Models ──────────────────────────────────────────────

class SupplierCreate(BaseModel):
    name: str
    contact_person: str = ""
    phone: str = ""
    email: str = ""
    address: str = ""
    gst_number: str = ""

class PurchaseItemModel(BaseModel):
    product_id: str = ""
    product_name: str
    quantity: int
    unit_cost: float
    gst_rate: float = 0.0

class PurchaseCreate(BaseModel):
    supplier_id: str
    items: List[PurchaseItemModel]
    notes: str = ""
    expected_date: str = ""

class PurchaseUpdate(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = None
    expected_date: Optional[str] = None

# ─── Scan Session Models ──────────────────────────────────────────

class ScanSessionCreate(BaseModel):
    type: str  # "inventory" or "pos"

# ─── Customer Models ─────────────────────────────────────────────

class CustomerCreate(BaseModel):
    name: str
    phone: str = ""
    email: str = ""
    address: str = ""
    gst_number: str = ""
    credit_limit: float = 0.0
    notes: str = ""

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    gst_number: Optional[str] = None
    credit_limit: Optional[float] = None
    notes: Optional[str] = None

class CreditAdjust(BaseModel):
    amount: float
    type: str  # "credit" or "payment"
    reference: str = ""
    notes: str = ""

# ─── API Key Models ──────────────────────────────────────────────

class APIKeyCreate(BaseModel):
    name: str
    permissions: List[str] = ["read_inventory", "read_invoices"]

# ─── IP Whitelist Model ─────────────────────────────────────────

class IPWhitelistUpdate(BaseModel):
    allowed_ips: List[str] = []

# ─── Support Ticket Models ───────────────────────────────────────

class SupportTicketCreate(BaseModel):
    subject: str
    description: str
    channel: str = "email"  # email, phone, whatsapp
    priority: str = "normal"  # low, normal, high, urgent

class TicketNoteCreate(BaseModel):
    message: str

class TicketStatusUpdate(BaseModel):
    status: str  # open, in_progress, resolved, closed

# ─── Financial Access Request Models ─────────────────────────────

class AccessRequestCreate(BaseModel):
    owner_id: str
    tenant_id: str
    request_type: str  # revenue, transactions, full_financial
    reason: str
    duration_hours: int = 24

class AccessRequestRespond(BaseModel):
    action: str  # approve, reject
    response_note: str = ""

# ─── Temp Access Models ──────────────────────────────────────────

class TempAccessCreate(BaseModel):
    user_id: str
    reason: str
    allowed_ip: str
    duration_hours: int = 24

# ─── User Permission Models ──────────────────────────────────────

class UserPermissionsUpdate(BaseModel):
    can_view_revenue: Optional[bool] = None
    can_manage_inventory: Optional[bool] = None

# ─── Promo Code Models ────────────────────────────────────────────

class PromoCodeCreate(BaseModel):
    code: str
    discount_type: str = "percentage"  # percentage or fixed
    value: float
    min_order_amount: float = 0
    max_discount: float = 0  # 0 = no cap
    valid_from: str = ""
    valid_to: str = ""
    max_uses: int = 0  # 0 = unlimited
    description: str = ""

class PromoCodeUpdate(BaseModel):
    discount_type: Optional[str] = None
    value: Optional[float] = None
    min_order_amount: Optional[float] = None
    max_discount: Optional[float] = None
    valid_from: Optional[str] = None
    valid_to: Optional[str] = None
    max_uses: Optional[int] = None
    is_active: Optional[bool] = None
    description: Optional[str] = None

class PromoCodeValidate(BaseModel):
    code: str
    order_amount: float

# ─── Reorder Rule Models ─────────────────────────────────────────

class ReorderRuleCreate(BaseModel):
    product_id: str
    threshold: int
    reorder_quantity: int
    notify_whatsapp: bool = False
    notify_email: bool = False
    notify_voice: bool = False
    supplier_phone: str = ""
    supplier_email: str = ""
    is_active: bool = True

class ReorderRuleUpdate(BaseModel):
    threshold: Optional[int] = None
    reorder_quantity: Optional[int] = None
    notify_whatsapp: Optional[bool] = None
    notify_email: Optional[bool] = None
    notify_voice: Optional[bool] = None
    supplier_phone: Optional[str] = None
    supplier_email: Optional[str] = None
    is_active: Optional[bool] = None

# ─── Notification Template Models ─────────────────────────────────

class NotificationTemplateCreate(BaseModel):
    channel: str  # whatsapp, email, voice
    name: str
    subject: str = ""  # for email
    template_text: str

class NotificationTemplateUpdate(BaseModel):
    name: Optional[str] = None
    subject: Optional[str] = None
    template_text: Optional[str] = None
    is_active: Optional[bool] = None

# ─── Advance Order Models ────────────────────────────────────────

class AdvanceOrderCreate(BaseModel):
    customer_name: str
    customer_phone: str = ""
    products: List[Dict[str, Any]]  # [{name, quantity, price, notes}]
    advance_amount: float
    total_estimated: float
    notes: str = ""

class AdvanceOrderFulfill(BaseModel):
    final_items: List[Dict[str, Any]] = []  # optional override

# ─── SMTP Settings Model ─────────────────────────────────────────

class SMTPSettingsUpdate(BaseModel):
    smtp_host: str = "smtp.gmail.com"
    smtp_port: int = 587
    smtp_email: str = ""
    smtp_password: str = ""
    sender_name: str = ""

# ─── Owner Account Management Models ─────────────────────────────

class CreateOwnerAccount(BaseModel):
    email: str
    password: str
    name: str
    shop_name: str
    business_type: str = "general"
    plan: str = "basic"
    valid_days: int = 365  # validity in days

class CreateAdminAccount(BaseModel):
    email: str
    password: str
    name: str

# ─── Branch Models ───────────────────────────────────────────────

class BranchCreate(BaseModel):
    name: str
    code: str = ""
    address: str = ""
    phone: str = ""
    manager_name: str = ""
    is_main: bool = False

class BranchUpdate(BaseModel):
    name: Optional[str] = None
    code: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    manager_name: Optional[str] = None
    is_active: Optional[bool] = None

# ─── Category Hierarchy Models ───────────────────────────────────

class CategoryCreate(BaseModel):
    name: str
    parent_id: Optional[str] = None
    description: str = ""
    icon: str = ""
    sort_order: int = 0

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    parent_id: Optional[str] = None
    description: Optional[str] = None
    icon: Optional[str] = None
    sort_order: Optional[int] = None
    is_active: Optional[bool] = None

# ─── Utility Functions ────────────────────────────────────────────

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str, tenant_id: str, role: str) -> str:
    payload = {
        "sub": user_id, "email": email, "tenant_id": tenant_id, "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=30), "type": "access"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_temp_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(minutes=5), "type": "temp_mfa"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def encrypt_mfa_secret(secret: str) -> str:
    key_bytes = MFA_KEY.encode('utf-8')[:32].ljust(32, b'\0')
    xored = bytes(a ^ b for a, b in zip(secret.encode('utf-8'), key_bytes * (len(secret) // 32 + 1)))
    return base64.b64encode(xored).decode('utf-8')

def decrypt_mfa_secret(encrypted: str) -> str:
    key_bytes = MFA_KEY.encode('utf-8')[:32].ljust(32, b'\0')
    decoded = base64.b64decode(encrypted.encode('utf-8'))
    xored = bytes(a ^ b for a, b in zip(decoded, key_bytes * (len(decoded) // 32 + 1)))
    return xored.decode('utf-8')

def generate_backup_codes(count=8):
    codes = [secrets.token_hex(4).upper() for _ in range(count)]
    hashed = [hashlib.sha256(c.encode()).hexdigest() for c in codes]
    return codes, hashed

def serialize_doc(doc):
    if doc is None:
        return None
    doc = dict(doc)
    if '_id' in doc:
        doc['id'] = str(doc.pop('_id'))
    for k, v in doc.items():
        if isinstance(v, ObjectId):
            doc[k] = str(v)
        elif isinstance(v, datetime):
            doc[k] = v.isoformat()
    return doc

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        if not user.get("is_active", True):
            raise HTTPException(status_code=403, detail="Account disabled")
        result = serialize_doc(user)
        result.pop("password_hash", None)
        result.pop("mfa_secret", None)
        return result
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(*roles):
    async def checker(request: Request):
        user = await get_current_user(request)
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return checker

async def log_audit(tenant_id: str, user_id: str, action: str, details: str = "", ip: str = "", event_category: str = "general"):
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()), "tenant_id": tenant_id, "user_id": user_id,
        "action": action, "details": details, "ip_address": ip,
        "event_category": event_category,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })

async def get_platform_admin(request: Request) -> dict:
    """Authenticate platform admin (separate from tenant users)"""
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        if not user.get("is_platform_admin", False):
            raise HTTPException(status_code=403, detail="Platform admin access required")
        result = serialize_doc(user)
        result.pop("password_hash", None)
        result.pop("mfa_secret", None)
        return result
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_platform_or_admin(request: Request) -> dict:
    """Authenticate platform admin OR admin user — both product-side roles."""
    user = await get_current_user(request)
    if not user.get("is_platform_admin", False) and not user.get("is_admin", False):
        raise HTTPException(status_code=403, detail="Requires admin or platform admin access")
    return user

async def create_security_alert(tenant_id: str, user_id: str, alert_type: str, severity: str, details: dict):
    """Create a security alert for fraud detection"""
    await db.security_alerts.insert_one({
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "user_id": user_id,
        "alert_type": alert_type,
        "severity": severity,
        "details": details,
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })

async def check_temp_access(user_id: str, client_ip: str) -> bool:
    """Check if a temporary IP access grant exists for this user/IP"""
    now = datetime.now(timezone.utc).isoformat()
    temp = await db.temp_access.find_one({
        "user_id": user_id,
        "allowed_ip": client_ip,
        "is_active": True,
        "expires_at": {"$gt": now}
    })
    return temp is not None

def get_user_permission(user: dict, perm: str) -> bool:
    """Check a granular permission on a user. OWNER always has all permissions."""
    if user.get("role") == "OWNER":
        return True
    perms = user.get("permissions", {})
    defaults = {"can_view_revenue": False, "can_manage_inventory": True}
    return perms.get(perm, defaults.get(perm, False))

async def alert_unauthorized_access(user: dict, resource: str, request: Request):
    """Create security alert + audit log when unauthorized access is attempted"""
    client_ip = request.client.host if request.client else "unknown"
    await create_security_alert(
        user["tenant_id"], user["id"], "unauthorized_access", "high",
        {"email": user.get("email", ""), "name": user.get("name", ""),
         "role": user.get("role", ""), "resource": resource, "ip": client_ip,
         "message": f"{user.get('name', 'Unknown')} ({user.get('role', '')}) attempted to access {resource} without permission"}
    )
    await log_audit(user["tenant_id"], user["id"], "unauthorized_access_attempt",
                    f"User {user.get('email')} tried to access {resource} without permission from IP {client_ip}",
                    client_ip, "security")

async def require_premium(user: dict):
    """Check if the user's tenant has premium plan"""
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"plan": 1})
    if not tenant or tenant.get("plan") != "premium":
        raise HTTPException(403, "This feature requires a Premium plan. Upgrade to access.")

async def check_tenant_validity(user: dict):
    """Check if tenant account is still valid (not expired/revoked)"""
    if user.get("is_platform_admin") or user.get("is_admin"):
        return
    tenant = await db.tenants.find_one({"id": user["tenant_id"]})
    if not tenant:
        raise HTTPException(403, "Tenant not found")
    if tenant.get("is_revoked"):
        raise HTTPException(403, "Your account has been suspended. Contact platform administrator.")
    valid_until = tenant.get("valid_until")
    if valid_until and valid_until < datetime.now(timezone.utc).isoformat():
        raise HTTPException(403, "Your account has expired. Contact platform administrator to renew.")

def render_template(template_text: str, variables: dict) -> str:
    """Render a notification template with dynamic variables"""
    result = template_text
    for key, value in variables.items():
        result = result.replace(f"{{{{{key}}}}}", str(value))
    return result

async def send_email_notification(tenant_id: str, recipient_email: str, subject: str, body: str) -> bool:
    """Send email via SMTP settings stored for the tenant"""
    tenant = await db.tenants.find_one({"id": tenant_id})
    smtp = tenant.get("smtp_settings", {})
    if not smtp.get("smtp_email") or not smtp.get("smtp_password"):
        logger.warning(f"SMTP not configured for tenant {tenant_id}")
        return False
    try:
        msg = MIMEMultipart()
        msg['From'] = f"{smtp.get('sender_name', 'RetailPro')} <{smtp['smtp_email']}>"
        msg['To'] = recipient_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))
        server = smtplib.SMTP(smtp.get('smtp_host', 'smtp.gmail.com'), smtp.get('smtp_port', 587))
        server.starttls()
        server.login(smtp['smtp_email'], smtp['smtp_password'])
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        logger.error(f"Email send failed: {e}")
        return False

def generate_whatsapp_url(phone: str, message: str) -> str:
    """Generate WhatsApp click-to-send URL"""
    clean_phone = ''.join(filter(str.isdigit, phone))
    if not clean_phone.startswith('91') and len(clean_phone) == 10:
        clean_phone = '91' + clean_phone
    return f"https://wa.me/{clean_phone}?text={url_quote(message)}"

async def trigger_reorder_notifications(tenant_id: str, product: dict, rule: dict, user: dict):
    """Trigger notifications for a low-stock product"""
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    variables = {
        "shop_name": tenant.get("shop_name", ""),
        "owner_name": user.get("name", ""),
        "product_name": product.get("name", ""),
        "current_stock": str(product.get("stock", 0)),
        "threshold": str(rule.get("threshold", 0)),
        "reorder_quantity": str(rule.get("reorder_quantity", 0)),
        "sku": product.get("sku", ""),
        "branch_name": tenant.get("branch_name", tenant.get("shop_name", "")),
    }

    # WhatsApp notification
    if rule.get("notify_whatsapp") and rule.get("supplier_phone"):
        template = await db.notification_templates.find_one(
            {"tenant_id": tenant_id, "channel": "whatsapp", "is_active": True}
        )
        default_msg = f"Reorder Alert from {{{{shop_name}}}}: {{{{product_name}}}} stock is at {{{{current_stock}}}} (threshold: {{{{threshold}}}}). Please supply {{{{reorder_quantity}}}} units. - {{{{owner_name}}}}"
        msg = render_template(template["template_text"] if template else default_msg, variables)
        wa_url = generate_whatsapp_url(rule["supplier_phone"], msg)
        await db.notification_logs.insert_one({
            "id": str(uuid.uuid4()), "tenant_id": tenant_id, "channel": "whatsapp",
            "recipient": rule["supplier_phone"], "subject": "Reorder Alert",
            "message": msg, "whatsapp_url": wa_url,
            "status": "ready", "related_product_id": product.get("id", ""),
            "product_name": product.get("name", ""),
            "created_at": datetime.now(timezone.utc).isoformat()
        })

    # Email notification
    if rule.get("notify_email") and rule.get("supplier_email"):
        template = await db.notification_templates.find_one(
            {"tenant_id": tenant_id, "channel": "email", "is_active": True}
        )
        default_subject = "Reorder Alert: {{product_name}} - Low Stock"
        default_body = "<h2>Reorder Alert</h2><p>Dear Supplier,</p><p>Product <strong>{{product_name}}</strong> (SKU: {{sku}}) at <strong>{{shop_name}}</strong> has reached low stock.</p><p>Current Stock: <strong>{{current_stock}}</strong><br>Threshold: {{threshold}}<br>Required Quantity: <strong>{{reorder_quantity}}</strong></p><p>Please arrange delivery at the earliest.</p><p>Regards,<br>{{owner_name}}<br>{{shop_name}}</p>"
        subject = render_template(template.get("subject", default_subject) if template else default_subject, variables)
        body = render_template(template["template_text"] if template else default_body, variables)
        sent = await send_email_notification(tenant_id, rule["supplier_email"], subject, body)
        await db.notification_logs.insert_one({
            "id": str(uuid.uuid4()), "tenant_id": tenant_id, "channel": "email",
            "recipient": rule["supplier_email"], "subject": subject,
            "message": body, "status": "sent" if sent else "failed",
            "related_product_id": product.get("id", ""),
            "product_name": product.get("name", ""),
            "created_at": datetime.now(timezone.utc).isoformat()
        })

    # Voice notification (MOCKED)
    if rule.get("notify_voice") and rule.get("supplier_phone"):
        template = await db.notification_templates.find_one(
            {"tenant_id": tenant_id, "channel": "voice", "is_active": True}
        )
        default_voice = "Hello, this is an automated reorder alert from {{shop_name}}. Product {{product_name}} stock is at {{current_stock}} units, which is below the threshold of {{threshold}}. Please supply {{reorder_quantity}} units. This message was sent on behalf of {{owner_name}}. Thank you."
        msg = render_template(template["template_text"] if template else default_voice, variables)
        await db.notification_logs.insert_one({
            "id": str(uuid.uuid4()), "tenant_id": tenant_id, "channel": "voice",
            "recipient": rule["supplier_phone"], "subject": "Reorder Voice Alert",
            "message": msg, "status": "queued_mock",
            "related_product_id": product.get("id", ""),
            "product_name": product.get("name", ""),
            "created_at": datetime.now(timezone.utc).isoformat()
        })

async def check_reorder_for_product(tenant_id: str, product_id: str, user: dict):
    """Check if a product needs reordering and trigger notifications"""
    rule = await db.reorder_rules.find_one({"tenant_id": tenant_id, "product_id": product_id, "is_active": True})
    if not rule:
        return
    product = await db.products.find_one({"tenant_id": tenant_id, "id": product_id}, {"_id": 0})
    if not product:
        return
    if product.get("stock", 0) <= rule.get("threshold", 0):
        # Check if we already sent notification recently (within 24h)
        recent = await db.notification_logs.find_one({
            "tenant_id": tenant_id, "related_product_id": product_id,
            "created_at": {"$gt": (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()}
        })
        if not recent:
            await trigger_reorder_notifications(tenant_id, product, rule, user)

# ─── SUBSCRIPTION PLANS ──────────────────────────────────────────

PLANS = {
    "basic": {"name": "Basic", "price": 999.00, "max_users": 2, "features": ["single_store", "basic_billing", "inventory"]},
    "standard": {"name": "Standard", "price": 2999.00, "max_users": 10, "features": ["single_store", "gst_billing", "inventory", "batch_tracking", "reports", "rbac"]},
    "premium": {"name": "Premium", "price": 7999.00, "max_users": 999, "features": ["multi_branch", "gst_billing", "inventory", "batch_tracking", "reports", "rbac", "ai_forecasting", "api_access"]},
}

# ═══════════════════════════════════════════════════════════════════
#  AUTH ROUTES
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/auth/register")
async def register(req: RegisterRequest, request: Request):
    email = req.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(400, "Email already registered")

    # Check if this is the first user (allow) or if registration is restricted
    total_tenants = await db.tenants.count_documents({})
    if total_tenants > 0:
        # Only allow registration if no tenants exist (first setup) or via admin
        # After initial setup, users must be created by OWNER
        raise HTTPException(403, "Registration is restricted. Please contact your administrator to create an account.")

    tenant_id = str(uuid.uuid4())
    tenant = {
        "id": tenant_id, "shop_name": req.shop_name, "business_type": req.business_type,
        "plan": "basic", "max_users": 2, "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.tenants.insert_one(tenant)

    user_doc = {
        "email": email, "password_hash": hash_password(req.password), "name": req.name,
        "role": "OWNER", "tenant_id": tenant_id, "is_active": True,
        "mfa_enabled": False, "mfa_secret": None, "allowed_ips": [],
        "last_activity": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)

    access_token = create_access_token(user_id, email, tenant_id, "OWNER")
    refresh_token = create_refresh_token(user_id)

    await log_audit(tenant_id, user_id, "register", f"New tenant created: {req.shop_name}", request.client.host if request.client else "")

    response = JSONResponse(content={
        "id": user_id, "email": email, "name": req.name, "role": "OWNER",
        "tenant_id": tenant_id, "shop_name": req.shop_name, "mfa_enabled": False,
        "plan": "basic"
    })
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=3600, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")
    return response

@api_router.post("/auth/login")
@limiter.limit("10/minute")
async def login(req: LoginRequest, request: Request):
    email = req.email.lower().strip()
    client_ip = request.client.host if request.client else "unknown"
    identifier = f"{client_ip}:{email}"

    attempt = await db.login_attempts.find_one({"identifier": identifier}, {"_id": 0})
    if attempt and attempt.get("count", 0) >= 5:
        locked_until = datetime.fromisoformat(attempt["locked_until"]) if isinstance(attempt.get("locked_until"), str) else attempt.get("locked_until")
        if locked_until and datetime.now(timezone.utc) < locked_until:
            raise HTTPException(429, "Account temporarily locked. Try again in 15 minutes.")
        else:
            await db.login_attempts.delete_one({"identifier": identifier})

    user = await db.users.find_one({"email": email})
    if not user or not verify_password(req.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$inc": {"count": 1}, "$set": {"locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()}},
            upsert=True
        )
        # Fraud Detection: alert on multiple failed attempts
        attempt_doc = await db.login_attempts.find_one({"identifier": identifier})
        fail_count = attempt_doc.get("count", 1) if attempt_doc else 1
        if fail_count >= 3 and user:
            await create_security_alert(
                user["tenant_id"], str(user["_id"]), "failed_logins",
                "high" if fail_count >= 5 else "medium",
                {"ip": client_ip, "email": email, "attempt_count": fail_count,
                 "message": f"{fail_count} failed login attempts from {client_ip}"}
            )
        raise HTTPException(401, "Invalid email or password")

    if not user.get("is_active", True):
        raise HTTPException(403, "Account disabled")

    # Check tenant validity (expiry/revocation) — skip for platform admin and admin
    if not user.get("is_platform_admin") and not user.get("is_admin"):
        tenant_check = await db.tenants.find_one({"id": user["tenant_id"]})
        if tenant_check and tenant_check.get("is_revoked"):
            raise HTTPException(403, "Your account has been suspended. Contact platform administrator.")
        if tenant_check and tenant_check.get("valid_until"):
            if tenant_check["valid_until"] < datetime.now(timezone.utc).isoformat():
                raise HTTPException(403, "Your account has expired. Contact platform administrator to renew.")

    # IP whitelist check - also check temp_access
    allowed_ips = user.get("allowed_ips", [])
    if allowed_ips and client_ip not in allowed_ips and client_ip != "127.0.0.1":
        # Check for temporary IP access
        has_temp = await check_temp_access(str(user["_id"]), client_ip)
        if not has_temp:
            await log_audit(user["tenant_id"], str(user["_id"]), "login_ip_blocked", f"Blocked IP: {client_ip}", client_ip, "security")
            await create_security_alert(
                user["tenant_id"], str(user["_id"]), "ip_blocked",
                "medium", {"ip": client_ip, "email": email, "message": f"Login attempt from non-whitelisted IP: {client_ip}"}
            )
            raise HTTPException(403, "Access denied from this IP address. Contact your administrator.")

    await db.login_attempts.delete_one({"identifier": identifier})
    user_id = str(user["_id"])
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})

    # ── Fraud Detection: New IP detection ──
    known_ips = user.get("known_login_ips", [])
    if client_ip not in known_ips and client_ip != "127.0.0.1" and client_ip != "unknown":
        if known_ips:  # Only alert if user has logged in before
            await create_security_alert(
                user["tenant_id"], user_id, "new_ip_login",
                "medium", {"ip": client_ip, "email": email, "known_ips": known_ips[:5],
                           "message": f"Login from new IP address: {client_ip}"}
            )
        # Add to known IPs
        await db.users.update_one(
            {"_id": user["_id"]},
            {"$addToSet": {"known_login_ips": client_ip}}
        )

    # Determine device source from User-Agent
    ua = request.headers.get("User-Agent", "").lower()
    device_source = "mobile" if any(k in ua for k in ["mobile", "android", "iphone", "ipad"]) else "desktop"

    # MFA required for all non-OWNER users, or if MFA is enabled
    if user.get("mfa_enabled") and user.get("mfa_secret"):
        temp_token = create_temp_token(user_id)
        await log_audit(user["tenant_id"], user_id, "login_mfa_required", f"Device: {device_source}", client_ip)
        return {"mfa_required": True, "temp_token": temp_token}

    # If user is non-OWNER and MFA not set up, require setup
    mfa_setup_required = not user.get("mfa_enabled", False) and user["role"] != "OWNER" and not user.get("is_platform_admin") and not user.get("is_admin")

    access_token = create_access_token(user_id, email, user["tenant_id"], user["role"])
    refresh_token = create_refresh_token(user_id)

    # Update last activity
    await db.users.update_one({"_id": user["_id"]}, {"$set": {"last_activity": datetime.now(timezone.utc).isoformat(), "last_login_ip": client_ip, "last_login_device": device_source}})

    await log_audit(user["tenant_id"], user_id, "login", f"Successful login from {device_source} ({client_ip})", client_ip)

    response_content = {
        "id": user_id, "email": email, "name": user["name"], "role": user["role"],
        "tenant_id": user["tenant_id"], "shop_name": tenant["shop_name"] if tenant else "",
        "mfa_enabled": user.get("mfa_enabled", False),
        "mfa_setup_required": mfa_setup_required,
        "plan": tenant.get("plan", "basic") if tenant else "basic",
        "is_platform_admin": user.get("is_platform_admin", False),
        "is_admin": user.get("is_admin", False),
    }
    response = JSONResponse(content=response_content)
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=3600, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")
    return response

@api_router.post("/auth/mfa/verify")
@limiter.limit("5/minute")
async def mfa_verify(req: MFAVerifyRequest, request: Request):
    try:
        payload = jwt.decode(req.temp_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "temp_mfa":
            raise HTTPException(401, "Invalid token")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid or expired token")

    user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
    if not user or not user.get("mfa_secret"):
        raise HTTPException(400, "MFA not configured")

    secret = decrypt_mfa_secret(user["mfa_secret"])
    totp = pyotp.TOTP(secret)
    if not totp.verify(req.otp_code, valid_window=1):
        raise HTTPException(401, "Invalid OTP code")

    user_id = str(user["_id"])
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    access_token = create_access_token(user_id, user["email"], user["tenant_id"], user["role"])
    refresh_token = create_refresh_token(user_id)

    await log_audit(user["tenant_id"], user_id, "mfa_verified", "MFA verification successful", request.client.host if request.client else "")

    response = JSONResponse(content={
        "id": user_id, "email": user["email"], "name": user["name"], "role": user["role"],
        "tenant_id": user["tenant_id"], "shop_name": tenant["shop_name"] if tenant else "",
        "mfa_enabled": True, "plan": tenant.get("plan", "basic") if tenant else "basic"
    })
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=1800, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")
    return response

@api_router.post("/auth/mfa/backup-login")
@limiter.limit("5/minute")
async def mfa_backup_login(req: MFABackupLoginRequest, request: Request):
    try:
        payload = jwt.decode(req.temp_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "temp_mfa":
            raise HTTPException(401, "Invalid token")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid or expired token")

    user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
    if not user:
        raise HTTPException(400, "User not found")

    code_hash = hashlib.sha256(req.backup_code.upper().strip().encode()).hexdigest()
    backup = await db.mfa_backup_codes.find_one({"user_id": str(user["_id"]), "code_hash": code_hash, "used": False})
    if not backup:
        raise HTTPException(401, "Invalid or already used backup code")

    await db.mfa_backup_codes.update_one({"_id": backup["_id"]}, {"$set": {"used": True, "used_at": datetime.now(timezone.utc).isoformat()}})

    user_id = str(user["_id"])
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    access_token = create_access_token(user_id, user["email"], user["tenant_id"], user["role"])
    refresh_token = create_refresh_token(user_id)

    await log_audit(user["tenant_id"], user_id, "mfa_backup_login", "Login via backup code", request.client.host if request.client else "")

    response = JSONResponse(content={
        "id": user_id, "email": user["email"], "name": user["name"], "role": user["role"],
        "tenant_id": user["tenant_id"], "shop_name": tenant["shop_name"] if tenant else "",
        "mfa_enabled": True, "plan": tenant.get("plan", "basic") if tenant else "basic"
    })
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=1800, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")
    return response

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    user["shop_name"] = tenant["shop_name"] if tenant else ""
    user["plan"] = tenant.get("plan", "basic") if tenant else "basic"
    user["mfa_setup_required"] = not user.get("mfa_enabled", False) and user.get("role") != "OWNER" and not user.get("is_platform_admin") and not user.get("is_admin")
    # Include permissions
    user["permissions"] = user.get("permissions", {"can_view_revenue": user.get("role") == "OWNER", "can_manage_inventory": True})
    # Ensure is_admin flag is included
    user["is_admin"] = user.get("is_admin", False)
    user["is_platform_admin"] = user.get("is_platform_admin", False)
    # Update last activity
    await db.users.update_one({"_id": ObjectId(user["id"])}, {"$set": {"last_activity": datetime.now(timezone.utc).isoformat()}})
    return user

@api_router.post("/auth/heartbeat")
async def heartbeat(request: Request):
    """Track user activity for idle timeout"""
    user = await get_current_user(request)
    await db.users.update_one({"_id": ObjectId(user["id"])}, {"$set": {"last_activity": datetime.now(timezone.utc).isoformat()}})
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}

@api_router.post("/auth/logout")
async def logout(request: Request):
    user = await get_current_user(request)
    await log_audit(user["tenant_id"], user["id"], "logout", "", request.client.host if request.client else "")
    response = JSONResponse(content={"message": "Logged out"})
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return response

@api_router.post("/auth/refresh")
async def refresh_token(request: Request):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(401, "No refresh token")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(401, "Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(401, "User not found")
        access_token = create_access_token(str(user["_id"]), user["email"], user["tenant_id"], user["role"])
        response = JSONResponse(content={"message": "Token refreshed"})
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=1800, path="/")
        return response
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid refresh token")

# ─── MFA Setup ────────────────────────────────────────────────────

@api_router.post("/auth/mfa/setup")
async def mfa_setup(request: Request):
    user = await get_current_user(request)
    secret = pyotp.random_base32()
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    shop = tenant["shop_name"] if tenant else "RetailSaaS"
    totp = pyotp.TOTP(secret)
    uri = totp.provisioning_uri(name=user["email"], issuer_name=f"RetailPro - {shop}")

    img = qrcode.make(uri)
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

    encrypted = encrypt_mfa_secret(secret)
    await db.users.update_one({"_id": ObjectId(user["id"])}, {"$set": {"mfa_secret": encrypted}})

    await log_audit(user["tenant_id"], user["id"], "mfa_setup_initiated", "", request.client.host if request.client else "")
    return {"qr_code": f"data:image/png;base64,{qr_base64}", "secret_key": secret, "uri": uri}

@api_router.post("/auth/mfa/enable")
async def mfa_enable(req: MFAEnableRequest, request: Request):
    user = await get_current_user(request)
    user_doc = await db.users.find_one({"_id": ObjectId(user["id"])})
    if not user_doc or not user_doc.get("mfa_secret"):
        raise HTTPException(400, "Run MFA setup first")

    secret = decrypt_mfa_secret(user_doc["mfa_secret"])
    totp = pyotp.TOTP(secret)
    if not totp.verify(req.otp_code, valid_window=1):
        raise HTTPException(400, "Invalid OTP. Try again.")

    codes, hashed = generate_backup_codes()
    for h in hashed:
        await db.mfa_backup_codes.insert_one({"user_id": user["id"], "code_hash": h, "used": False, "created_at": datetime.now(timezone.utc).isoformat()})

    await db.users.update_one({"_id": ObjectId(user["id"])}, {"$set": {"mfa_enabled": True}})
    await log_audit(user["tenant_id"], user["id"], "mfa_enabled", "MFA enabled with backup codes", request.client.host if request.client else "")
    return {"message": "MFA enabled", "backup_codes": codes}

@api_router.post("/auth/mfa/backup-codes")
async def regenerate_backup_codes(request: Request):
    user = await get_current_user(request)
    if not user.get("mfa_enabled"):
        raise HTTPException(400, "MFA not enabled")

    await db.mfa_backup_codes.delete_many({"user_id": user["id"]})
    codes, hashed = generate_backup_codes()
    for h in hashed:
        await db.mfa_backup_codes.insert_one({"user_id": user["id"], "code_hash": h, "used": False, "created_at": datetime.now(timezone.utc).isoformat()})

    await log_audit(user["tenant_id"], user["id"], "backup_codes_regenerated", "", request.client.host if request.client else "")
    return {"backup_codes": codes}

@api_router.post("/admin/users/{user_id}/mfa/reset")
async def admin_reset_mfa(user_id: str, request: Request):
    admin = await get_current_user(request)
    if admin["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can reset MFA")

    target = await db.users.find_one({"_id": ObjectId(user_id), "tenant_id": admin["tenant_id"]})
    if not target:
        raise HTTPException(404, "User not found")

    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"mfa_enabled": False, "mfa_secret": None}})
    await db.mfa_backup_codes.delete_many({"user_id": user_id})
    await log_audit(admin["tenant_id"], admin["id"], "admin_mfa_reset", f"Reset MFA for user {user_id}", request.client.host if request.client else "")
    return {"message": "MFA reset successfully"}

# ═══════════════════════════════════════════════════════════════════
#  INVENTORY ROUTES
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/inventory/products")
async def list_products(request: Request, search: str = "", category: str = "", page: int = 1, limit: int = 50, branch_id: str = ""):
    user = await get_current_user(request)
    if not get_user_permission(user, "can_manage_inventory"):
        await alert_unauthorized_access(user, "inventory_products", request)
        raise HTTPException(403, "You don't have permission to access inventory. Contact your store owner.")

    # Staff can only see their assigned branch
    user_branch = user.get("branch_id", "")
    effective_branch = branch_id
    if user["role"] == "STAFF" and user_branch:
        effective_branch = user_branch

    # Check cache for non-search requests
    cache_key = f"products:{search}:{category}:{page}:{limit}:{effective_branch}"
    if not search:  # Only cache non-search requests (full lists)
        cached = product_cache.get(user["tenant_id"], cache_key)
        if cached is not None:
            return cached

    query = {"tenant_id": user["tenant_id"]}
    if effective_branch:
        query["branch_id"] = effective_branch
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"sku": {"$regex": search, "$options": "i"}},
            {"barcode": {"$regex": search, "$options": "i"}}
        ]
    if category:
        query["category"] = category

    total = await db.products.count_documents(query)
    skip = (page - 1) * limit
    products = await db.products.find(query, {"_id": 0}).sort("name", 1).skip(skip).limit(limit).to_list(limit)
    result = {"products": products, "total": total, "page": page, "pages": (total + limit - 1) // limit}

    # Cache the result
    product_cache.set(user["tenant_id"], cache_key, result, ttl=45)
    return result

@api_router.post("/inventory/products")
async def create_product(product: ProductCreate, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot add products")

    existing = await db.products.find_one({"tenant_id": user["tenant_id"], "sku": product.sku}) if product.sku else None
    if existing:
        raise HTTPException(400, "SKU already exists")

    doc = product.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["tenant_id"] = user["tenant_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    doc["created_by"] = user["id"]

    await db.products.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "product_created", f"Product: {product.name}", request.client.host if request.client else "")
    # Invalidate product & category caches
    product_cache.invalidate(user["tenant_id"])
    category_cache.invalidate(user["tenant_id"])
    doc.pop("_id", None)
    return doc

@api_router.put("/inventory/products/{product_id}")
async def update_product(product_id: str, update: ProductUpdate, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot edit products")

    existing = await db.products.find_one({"id": product_id, "tenant_id": user["tenant_id"]})
    if not existing:
        raise HTTPException(404, "Product not found")

    updates = {k: v for k, v in update.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.products.update_one({"id": product_id, "tenant_id": user["tenant_id"]}, {"$set": updates})

    await log_audit(user["tenant_id"], user["id"], "product_updated", f"Product: {product_id}", request.client.host if request.client else "")
    # Invalidate caches
    product_cache.invalidate(user["tenant_id"])
    category_cache.invalidate(user["tenant_id"])
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    return updated

@api_router.delete("/inventory/products/{product_id}")
async def delete_product(product_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")

    result = await db.products.delete_one({"id": product_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Product not found")

    await log_audit(user["tenant_id"], user["id"], "product_deleted", f"Product: {product_id}", request.client.host if request.client else "")
    # Invalidate caches
    product_cache.invalidate(user["tenant_id"])
    category_cache.invalidate(user["tenant_id"])
    return {"message": "Product deleted"}

@api_router.get("/inventory/categories")
async def list_categories(request: Request):
    user = await get_current_user(request)
    # Check cache
    cached = category_cache.get(user["tenant_id"], "categories")
    if cached is not None:
        return cached
    cats = await db.products.distinct("category", {"tenant_id": user["tenant_id"]})
    result = {"categories": [c for c in cats if c]}
    category_cache.set(user["tenant_id"], "categories", result, ttl=300)
    return result

@api_router.post("/inventory/stock-adjust")
async def adjust_stock(adj: StockAdjust, request: Request):
    user = await get_current_user(request)
    product = await db.products.find_one({"id": adj.product_id, "tenant_id": user["tenant_id"]})
    if not product:
        raise HTTPException(404, "Product not found")

    new_stock = product.get("stock", 0) + adj.adjustment
    if new_stock < 0:
        raise HTTPException(400, "Stock cannot be negative")

    await db.products.update_one(
        {"id": adj.product_id, "tenant_id": user["tenant_id"]},
        {"$set": {"stock": new_stock, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await log_audit(user["tenant_id"], user["id"], "stock_adjusted", f"Product: {adj.product_id}, adjustment: {adj.adjustment}, reason: {adj.reason}", request.client.host if request.client else "")
    # Check reorder rules
    await check_reorder_for_product(user["tenant_id"], adj.product_id, user)
    return {"message": "Stock adjusted", "new_stock": new_stock}

@api_router.get("/inventory/low-stock")
async def low_stock_alerts(request: Request):
    user = await get_current_user(request)
    pipeline = [
        {"$match": {"tenant_id": user["tenant_id"]}},
        {"$match": {"$expr": {"$lte": ["$stock", "$low_stock_threshold"]}}},
        {"$project": {"_id": 0}},
        {"$sort": {"stock": 1}},
        {"$limit": 50}
    ]
    products = await db.products.aggregate(pipeline).to_list(50)
    return {"alerts": products}

# ═══════════════════════════════════════════════════════════════════
#  POS / BILLING ROUTES
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/pos/invoice")
async def create_invoice(inv: InvoiceCreate, request: Request):
    user = await get_current_user(request)

    subtotal = 0.0
    tax_total = 0.0
    validated_items = []

    for item in inv.items:
        product = await db.products.find_one({"id": item.product_id, "tenant_id": user["tenant_id"]})
        if not product:
            raise HTTPException(400, f"Product {item.product_id} not found")
        if product.get("stock", 0) < item.quantity:
            raise HTTPException(400, f"Insufficient stock for {product['name']}")

        item_total = item.price * item.quantity
        item_tax = item_total * (item.gst_rate / 100)
        subtotal += item_total
        tax_total += item_tax
        validated_items.append({
            "product_id": item.product_id, "name": item.name,
            "quantity": item.quantity, "price": item.price,
            "gst_rate": item.gst_rate, "total": item_total, "tax": item_tax
        })

        await db.products.update_one(
            {"id": item.product_id, "tenant_id": user["tenant_id"]},
            {"$inc": {"stock": -item.quantity}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
        )

    grand_total = subtotal + tax_total - inv.discount
    count = await db.invoices.count_documents({"tenant_id": user["tenant_id"]})
    invoice_number = f"INV-{count + 1:06d}"

    invoice_doc = {
        "id": str(uuid.uuid4()), "invoice_number": invoice_number,
        "tenant_id": user["tenant_id"], "created_by": user["id"],
        "customer_name": inv.customer_name, "customer_phone": inv.customer_phone,
        "customer_id": inv.customer_id or "",
        "items": validated_items, "subtotal": round(subtotal, 2),
        "tax_total": round(tax_total, 2), "discount": round(inv.discount, 2),
        "grand_total": round(grand_total, 2), "payment_method": inv.payment_method,
        "notes": inv.notes, "status": "completed",
        "device_source": inv.device_source,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.invoices.insert_one(invoice_doc)

    # Update customer credit if applicable
    if inv.customer_id:
        customer = await db.customers.find_one({"id": inv.customer_id, "tenant_id": user["tenant_id"]})
        if customer:
            await db.customers.update_one(
                {"id": inv.customer_id},
                {"$inc": {"total_purchases": grand_total}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
            )

    await log_audit(user["tenant_id"], user["id"], "invoice_created", f"Invoice: {invoice_number}, Total: {grand_total}, Device: {inv.device_source}", request.client.host if request.client else "")
    # Invalidate caches (stock changed, new invoice)
    product_cache.invalidate(user["tenant_id"])
    dashboard_cache.invalidate(user["tenant_id"])
    customer_cache.invalidate(user["tenant_id"])
    invoice_doc.pop("_id", None)
    return invoice_doc

@api_router.get("/pos/invoices")
async def list_invoices(request: Request, page: int = 1, limit: int = 20, search: str = ""):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    if search:
        query["$or"] = [
            {"invoice_number": {"$regex": search, "$options": "i"}},
            {"customer_name": {"$regex": search, "$options": "i"}}
        ]
    total = await db.invoices.count_documents(query)
    skip = (page - 1) * limit
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"invoices": invoices, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.get("/pos/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, request: Request):
    user = await get_current_user(request)
    invoice = await db.invoices.find_one({"id": invoice_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not invoice:
        raise HTTPException(404, "Invoice not found")
    return invoice

# ═══════════════════════════════════════════════════════════════════
#  REPORTS ROUTES
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/reports/dashboard")
async def dashboard_stats(request: Request):
    user = await get_current_user(request)
    tid = user["tenant_id"]

    # Check cache (cache per user role since revenue visibility differs)
    cache_key = f"dashboard:{user['role']}:{user['id']}"
    cached = dashboard_cache.get(tid, cache_key)
    if cached is not None:
        return cached

    total_products = await db.products.count_documents({"tenant_id": tid})
    total_invoices = await db.invoices.count_documents({"tenant_id": tid})

    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    today_invoices = await db.invoices.find({"tenant_id": tid, "created_at": {"$gte": today}}, {"_id": 0}).to_list(1000)
    today_revenue = sum(inv.get("grand_total", 0) for inv in today_invoices)
    today_count = len(today_invoices)

    low_stock_pipeline = [
        {"$match": {"tenant_id": tid}},
        {"$match": {"$expr": {"$lte": ["$stock", "$low_stock_threshold"]}}},
        {"$count": "count"}
    ]
    low_stock_result = await db.products.aggregate(low_stock_pipeline).to_list(1)
    low_stock_count = low_stock_result[0]["count"] if low_stock_result else 0

    total_stock_value = 0
    async for p in db.products.find({"tenant_id": tid}, {"stock": 1, "cost_price": 1, "_id": 0}):
        total_stock_value += p.get("stock", 0) * p.get("cost_price", 0)

    # Last 7 days revenue
    revenue_by_day = []
    for i in range(6, -1, -1):
        day = (datetime.now(timezone.utc) - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        next_day = day + timedelta(days=1)
        day_invoices = await db.invoices.find(
            {"tenant_id": tid, "created_at": {"$gte": day.isoformat(), "$lt": next_day.isoformat()}},
            {"_id": 0, "grand_total": 1}
        ).to_list(1000)
        day_total = sum(inv.get("grand_total", 0) for inv in day_invoices)
        revenue_by_day.append({"date": day.strftime("%b %d"), "revenue": round(day_total, 2)})

    recent_invoices = await db.invoices.find({"tenant_id": tid}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)

    # Revenue Visibility: check granular permission
    can_view = get_user_permission(user, "can_view_revenue")
    if not can_view and user["role"] != "OWNER":
        pass  # Dashboard is always shown, just hide data

    result = {
        "total_products": total_products, "total_invoices": total_invoices,
        "today_revenue": round(today_revenue, 2) if can_view else None,
        "today_invoices": today_count,
        "low_stock_count": low_stock_count,
        "stock_value": round(total_stock_value, 2) if can_view else None,
        "revenue_by_day": revenue_by_day if can_view else [],
        "recent_invoices": recent_invoices if can_view else [],
        "revenue_hidden": not can_view
    }

    # Cache for 2 minutes
    dashboard_cache.set(tid, cache_key, result, ttl=120)
    return result

@api_router.get("/reports/sales")
async def sales_report(request: Request, start_date: str = "", end_date: str = ""):
    user = await get_current_user(request)
    if not get_user_permission(user, "can_view_revenue"):
        await alert_unauthorized_access(user, "sales_report", request)
        raise HTTPException(403, "You don't have permission to view revenue data. Contact your store owner.")
    query = {"tenant_id": user["tenant_id"]}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date

    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    total_revenue = sum(inv.get("grand_total", 0) for inv in invoices)
    total_tax = sum(inv.get("tax_total", 0) for inv in invoices)

    product_sales = {}
    for inv in invoices:
        for item in inv.get("items", []):
            pid = item["product_id"]
            if pid not in product_sales:
                product_sales[pid] = {"name": item["name"], "quantity": 0, "revenue": 0}
            product_sales[pid]["quantity"] += item["quantity"]
            product_sales[pid]["revenue"] += item["total"]

    top_products = sorted(product_sales.values(), key=lambda x: x["revenue"], reverse=True)[:10]
    return {
        "total_revenue": round(total_revenue, 2), "total_tax": round(total_tax, 2),
        "invoice_count": len(invoices), "top_products": top_products
    }

# ═══════════════════════════════════════════════════════════════════
#  USER MANAGEMENT ROUTES
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/users")
async def list_users(request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Insufficient permissions")
    users = await db.users.find({"tenant_id": user["tenant_id"]}, {"password_hash": 0, "mfa_secret": 0}).to_list(500)
    # Get branches for enrichment
    branches = await db.branches.find({"tenant_id": user["tenant_id"]}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    bmap = {b["id"]: b["name"] for b in branches}
    for u in users:
        if "_id" in u:
            u["id"] = str(u.pop("_id"))
        # Ensure permissions are included
        if "permissions" not in u:
            u["permissions"] = {"can_view_revenue": u.get("role") == "OWNER", "can_manage_inventory": True}
        u["branch_name"] = bmap.get(u.get("branch_id", ""), "All Branches")
    return {"users": users, "branches": branches}

@api_router.post("/users")
async def create_user(req: UserCreate, request: Request):
    admin = await get_current_user(request)
    if admin["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")

    tenant = await db.tenants.find_one({"id": admin["tenant_id"]}, {"_id": 0})
    plan = tenant.get("plan", "basic") if tenant else "basic"
    max_users = PLANS.get(plan, {}).get("max_users", 2)
    current_count = await db.users.count_documents({"tenant_id": admin["tenant_id"]})
    if current_count >= max_users:
        raise HTTPException(400, f"User limit ({max_users}) reached for {plan} plan")

    if req.role not in ["OWNER", "MANAGER", "STAFF"]:
        raise HTTPException(400, "Invalid role")
    if admin["role"] == "MANAGER" and req.role in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Managers can only create STAFF users")

    email = req.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(400, "Email already registered")

    user_doc = {
        "email": email, "password_hash": hash_password(req.password), "name": req.name,
        "role": req.role, "tenant_id": admin["tenant_id"], "is_active": True,
        "mfa_enabled": False, "mfa_secret": None, "branch_id": "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.users.insert_one(user_doc)
    await log_audit(admin["tenant_id"], admin["id"], "user_created", f"Created user {email} as {req.role}", request.client.host if request.client else "")
    return {"id": str(result.inserted_id), "email": email, "name": req.name, "role": req.role}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, req: UserUpdate, request: Request):
    admin = await get_current_user(request)
    if admin["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")

    target = await db.users.find_one({"_id": ObjectId(user_id), "tenant_id": admin["tenant_id"]})
    if not target:
        raise HTTPException(404, "User not found")

    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    if "role" in updates and admin["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can change roles")

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": updates})
    await log_audit(admin["tenant_id"], admin["id"], "user_updated", f"Updated user {user_id}", request.client.host if request.client else "")
    return {"message": "User updated"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request):
    admin = await get_current_user(request)
    if admin["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can delete users")
    if admin["id"] == user_id:
        raise HTTPException(400, "Cannot delete yourself")

    result = await db.users.delete_one({"_id": ObjectId(user_id), "tenant_id": admin["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "User not found")
    await log_audit(admin["tenant_id"], admin["id"], "user_deleted", f"Deleted user {user_id}", request.client.host if request.client else "")
    return {"message": "User deleted"}

@api_router.get("/users/{user_id}/permissions")
async def get_user_permissions(user_id: str, request: Request):
    """Get a user's granular permissions"""
    admin = await get_current_user(request)
    if admin["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can view user permissions")

    target = await db.users.find_one({"_id": ObjectId(user_id), "tenant_id": admin["tenant_id"]})
    if not target:
        raise HTTPException(404, "User not found")

    perms = target.get("permissions", {"can_view_revenue": False, "can_manage_inventory": True})
    return {"user_id": user_id, "name": target.get("name", ""), "role": target.get("role", ""), "permissions": perms}

@api_router.put("/users/{user_id}/permissions")
async def update_user_permissions(user_id: str, req: UserPermissionsUpdate, request: Request):
    """Owner toggles revenue visibility and inventory access for a user"""
    admin = await get_current_user(request)
    if admin["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can manage user permissions")

    target = await db.users.find_one({"_id": ObjectId(user_id), "tenant_id": admin["tenant_id"]})
    if not target:
        raise HTTPException(404, "User not found")

    if target.get("role") == "OWNER":
        raise HTTPException(400, "Cannot change OWNER permissions")

    current_perms = target.get("permissions", {"can_view_revenue": False, "can_manage_inventory": True})
    updates = {}
    if req.can_view_revenue is not None:
        current_perms["can_view_revenue"] = req.can_view_revenue
        updates["revenue_access"] = "granted" if req.can_view_revenue else "revoked"
    if req.can_manage_inventory is not None:
        current_perms["can_manage_inventory"] = req.can_manage_inventory
        updates["inventory_access"] = "granted" if req.can_manage_inventory else "revoked"

    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"permissions": current_perms, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    details = ", ".join([f"{k}: {v}" for k, v in updates.items()])
    await log_audit(admin["tenant_id"], admin["id"], "user_permissions_updated",
                    f"User {target.get('email')}: {details}",
                    request.client.host if request.client else "", "security")
    return {"message": "Permissions updated", "permissions": current_perms}

# ═══════════════════════════════════════════════════════════════════
#  SUBSCRIPTION / STRIPE ROUTES
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/subscription/plans")
async def get_plans():
    return {"plans": PLANS}

@api_router.get("/subscription/current")
async def get_current_subscription(request: Request):
    user = await get_current_user(request)
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    plan = tenant.get("plan", "basic") if tenant else "basic"
    return {"plan": plan, "details": PLANS.get(plan, PLANS["basic"]), "tenant": tenant}

@api_router.post("/subscription/checkout")
async def create_checkout(req: CheckoutRequest, request: Request):
    user = await get_current_user(request)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can manage subscription")

    if req.plan_id not in PLANS:
        raise HTTPException(400, "Invalid plan")

    plan = PLANS[req.plan_id]
    api_key = os.environ.get("STRIPE_API_KEY")
    if not api_key:
        raise HTTPException(500, "Payment not configured")

    from emergentintegrations.payments.stripe.checkout import StripeCheckout, CheckoutSessionRequest

    webhook_url = f"{str(request.base_url).rstrip('/')}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=api_key, webhook_url=webhook_url)

    success_url = f"{req.origin_url}/settings?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{req.origin_url}/settings"

    checkout_req = CheckoutSessionRequest(
        amount=plan["price"],
        currency="inr",
        success_url=success_url,
        cancel_url=cancel_url,
        metadata={"tenant_id": user["tenant_id"], "plan_id": req.plan_id, "user_id": user["id"]}
    )
    session = await stripe_checkout.create_checkout_session(checkout_req)

    await db.payment_transactions.insert_one({
        "id": str(uuid.uuid4()), "session_id": session.session_id,
        "tenant_id": user["tenant_id"], "user_id": user["id"],
        "plan_id": req.plan_id, "amount": plan["price"], "currency": "inr",
        "payment_status": "initiated", "created_at": datetime.now(timezone.utc).isoformat()
    })

    await log_audit(user["tenant_id"], user["id"], "checkout_created", f"Plan: {req.plan_id}", request.client.host if request.client else "")
    return {"url": session.url, "session_id": session.session_id}

@api_router.get("/subscription/checkout/status/{session_id}")
async def checkout_status(session_id: str, request: Request):
    user = await get_current_user(request)
    txn = await db.payment_transactions.find_one({"session_id": session_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not txn:
        raise HTTPException(404, "Transaction not found")

    if txn.get("payment_status") == "paid":
        return txn

    api_key = os.environ.get("STRIPE_API_KEY")
    from emergentintegrations.payments.stripe.checkout import StripeCheckout
    webhook_url = f"{str(request.base_url).rstrip('/')}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=api_key, webhook_url=webhook_url)

    status = await stripe_checkout.get_checkout_status(session_id)

    update_data = {"payment_status": status.payment_status, "status": status.status, "updated_at": datetime.now(timezone.utc).isoformat()}
    await db.payment_transactions.update_one({"session_id": session_id}, {"$set": update_data})

    if status.payment_status == "paid" and txn.get("payment_status") != "paid":
        plan_id = txn.get("plan_id", "basic")
        plan_details = PLANS.get(plan_id, PLANS["basic"])
        await db.tenants.update_one(
            {"id": user["tenant_id"]},
            {"$set": {"plan": plan_id, "max_users": plan_details["max_users"], "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        await log_audit(user["tenant_id"], user["id"], "subscription_upgraded", f"Upgraded to {plan_id}", request.client.host if request.client else "")

    return {**txn, **update_data}

@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    body = await request.body()
    api_key = os.environ.get("STRIPE_API_KEY")
    from emergentintegrations.payments.stripe.checkout import StripeCheckout
    webhook_url = f"{str(request.base_url).rstrip('/')}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=api_key, webhook_url=webhook_url)

    try:
        event = await stripe_checkout.handle_webhook(body, request.headers.get("Stripe-Signature"))
        if event.payment_status == "paid":
            txn = await db.payment_transactions.find_one({"session_id": event.session_id}, {"_id": 0})
            if txn and txn.get("payment_status") != "paid":
                await db.payment_transactions.update_one(
                    {"session_id": event.session_id},
                    {"$set": {"payment_status": "paid", "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                plan_id = txn.get("plan_id", "basic")
                plan_details = PLANS.get(plan_id, PLANS["basic"])
                await db.tenants.update_one(
                    {"id": txn["tenant_id"]},
                    {"$set": {"plan": plan_id, "max_users": plan_details["max_users"]}}
                )
        return {"status": "ok"}
    except Exception as e:
        logger.error(f"Webhook error: {e}")
        return {"status": "error"}

# ═══════════════════════════════════════════════════════════════════
#  AUDIT LOGS
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/audit-logs")
async def get_audit_logs(request: Request, page: int = 1, limit: int = 50):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Insufficient permissions")
    total = await db.audit_logs.count_documents({"tenant_id": user["tenant_id"]})
    skip = (page - 1) * limit
    logs = await db.audit_logs.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "page": page, "pages": (total + limit - 1) // limit}

# ═══════════════════════════════════════════════════════════════════
#  AI DEMAND FORECASTING (GPT-5.2)
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/reports/ai-forecast")
async def ai_demand_forecast(request: Request):
    user = await get_current_user(request)
    tid = user["tenant_id"]

    # Check premium plan
    tenant = await db.tenants.find_one({"id": tid}, {"_id": 0})
    if tenant and tenant.get("plan") not in ["premium", "standard"]:
        raise HTTPException(403, "AI Forecasting is available on Standard and Premium plans")

    # Gather last 30 days of sales data
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    invoices = await db.invoices.find(
        {"tenant_id": tid, "created_at": {"$gte": thirty_days_ago}},
        {"_id": 0}
    ).to_list(5000)

    # Aggregate product sales
    product_sales = {}
    daily_revenue = {}
    for inv in invoices:
        day = inv.get("created_at", "")[:10]
        daily_revenue[day] = daily_revenue.get(day, 0) + inv.get("grand_total", 0)
        for item in inv.get("items", []):
            pid = item.get("product_id", item.get("name", "unknown"))
            if pid not in product_sales:
                product_sales[pid] = {"name": item["name"], "total_qty": 0, "total_revenue": 0, "days_sold": set()}
            product_sales[pid]["total_qty"] += item["quantity"]
            product_sales[pid]["total_revenue"] += item.get("total", item["price"] * item["quantity"])
            product_sales[pid]["days_sold"].add(day)

    # Get low stock products
    low_stock = await db.products.find(
        {"tenant_id": tid, "$expr": {"$lte": ["$stock", "$low_stock_threshold"]}},
        {"_id": 0, "name": 1, "stock": 1, "low_stock_threshold": 1, "category": 1}
    ).to_list(50)

    # Prepare data summary for AI
    sales_summary = []
    for pid, data in sorted(product_sales.items(), key=lambda x: x[1]["total_revenue"], reverse=True)[:20]:
        data["days_sold"] = len(data["days_sold"])
        sales_summary.append({
            "product": data["name"],
            "quantity_sold": data["total_qty"],
            "revenue": round(data["total_revenue"], 2),
            "days_with_sales": data["days_sold"],
            "avg_daily": round(data["total_qty"] / max(data["days_sold"], 1), 1)
        })

    revenue_trend = [{"date": k, "revenue": round(v, 2)} for k, v in sorted(daily_revenue.items())]

    # Call GPT-5.2 for analysis
    llm_key = os.environ.get("EMERGENT_LLM_KEY")
    ai_analysis = None
    if llm_key and sales_summary:
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
            chat = LlmChat(
                api_key=llm_key,
                session_id=f"forecast-{tid}-{datetime.now(timezone.utc).strftime('%Y%m%d')}",
                system_message="You are a retail business analyst AI. Analyze sales data and provide demand forecasts, restock recommendations, and business insights. Be specific with numbers. Always respond in valid JSON format."
            ).with_model("openai", "gpt-5.2")

            prompt = f"""Analyze this retail sales data from the last 30 days and provide demand forecasting:

## Top Products by Revenue:
{json.dumps(sales_summary, indent=2)}

## Daily Revenue Trend:
{json.dumps(revenue_trend[-14:], indent=2)}

## Low Stock Items:
{json.dumps(low_stock[:10], indent=2)}

Business type: {tenant.get('business_type', 'general')}
Shop: {tenant.get('shop_name', 'Retail Store')}

Respond with ONLY valid JSON (no markdown, no code blocks) in this exact structure:
{{
  "forecast_summary": "Brief 2-3 sentence overview of business performance",
  "demand_predictions": [
    {{"product": "name", "predicted_weekly_demand": 0, "confidence": "high/medium/low", "trend": "up/stable/down", "recommendation": "specific action"}}
  ],
  "restock_alerts": [
    {{"product": "name", "current_stock": 0, "estimated_days_left": 0, "suggested_order_qty": 0, "priority": "urgent/normal/low"}}
  ],
  "business_insights": [
    "Specific actionable insight 1",
    "Specific actionable insight 2",
    "Specific actionable insight 3"
  ],
  "revenue_forecast": {{
    "next_week_estimate": 0,
    "next_month_estimate": 0,
    "growth_trend": "growing/stable/declining"
  }}
}}"""

            user_msg = UserMessage(text=prompt)
            response = await chat.send_message(user_msg)

            # Parse the response - handle potential markdown wrapping
            response_text = response.strip()
            if response_text.startswith("```"):
                lines = response_text.split("\n")
                response_text = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
            ai_analysis = json.loads(response_text)
        except json.JSONDecodeError:
            ai_analysis = {"forecast_summary": response_text if response_text else "Unable to parse AI response", "demand_predictions": [], "restock_alerts": [], "business_insights": [], "revenue_forecast": {}}
        except Exception as e:
            logger.error(f"AI Forecast error: {e}")
            ai_analysis = {"forecast_summary": f"AI analysis temporarily unavailable: {str(e)}", "demand_predictions": [], "restock_alerts": [], "business_insights": [], "revenue_forecast": {}}

    await log_audit(tid, user["id"], "ai_forecast_generated", "Demand forecast requested", request.client.host if request.client else "")

    return {
        "sales_summary": sales_summary,
        "revenue_trend": revenue_trend,
        "low_stock": low_stock,
        "ai_analysis": ai_analysis,
        "generated_at": datetime.now(timezone.utc).isoformat()
    }

# ═══════════════════════════════════════════════════════════════════
#  PURCHASE MANAGEMENT ROUTES
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/purchases/suppliers")
async def list_suppliers(request: Request):
    user = await get_current_user(request)
    suppliers = await db.suppliers.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("name", 1).to_list(500)
    return {"suppliers": suppliers}

@api_router.post("/purchases/suppliers")
async def create_supplier(req: SupplierCreate, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot manage suppliers")

    doc = req.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["tenant_id"] = user["tenant_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.suppliers.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "supplier_created", f"Supplier: {req.name}", request.client.host if request.client else "")
    doc.pop("_id", None)
    return doc

@api_router.delete("/purchases/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")
    result = await db.suppliers.delete_one({"id": supplier_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Supplier not found")
    return {"message": "Supplier deleted"}

@api_router.get("/purchases")
async def list_purchases(request: Request, page: int = 1, limit: int = 20, status: str = ""):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    if status:
        query["status"] = status
    total = await db.purchases.count_documents(query)
    skip = (page - 1) * limit
    purchases = await db.purchases.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"purchases": purchases, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.post("/purchases")
async def create_purchase(req: PurchaseCreate, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot create purchase orders")

    supplier = await db.suppliers.find_one({"id": req.supplier_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not supplier:
        raise HTTPException(404, "Supplier not found")

    subtotal = 0.0
    tax_total = 0.0
    items = []
    for item in req.items:
        item_total = item.unit_cost * item.quantity
        item_tax = item_total * (item.gst_rate / 100)
        subtotal += item_total
        tax_total += item_tax
        items.append({
            "product_id": item.product_id, "product_name": item.product_name,
            "quantity": item.quantity, "unit_cost": item.unit_cost,
            "gst_rate": item.gst_rate, "total": round(item_total, 2), "tax": round(item_tax, 2),
            "received_qty": 0
        })

    count = await db.purchases.count_documents({"tenant_id": user["tenant_id"]})
    po_number = f"PO-{count + 1:06d}"

    doc = {
        "id": str(uuid.uuid4()), "po_number": po_number,
        "tenant_id": user["tenant_id"], "created_by": user["id"],
        "supplier_id": req.supplier_id, "supplier_name": supplier["name"],
        "items": items, "subtotal": round(subtotal, 2),
        "tax_total": round(tax_total, 2), "grand_total": round(subtotal + tax_total, 2),
        "status": "pending", "notes": req.notes,
        "expected_date": req.expected_date,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.purchases.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "purchase_created", f"PO: {po_number}, Supplier: {supplier['name']}", request.client.host if request.client else "")
    doc.pop("_id", None)
    return doc

@api_router.put("/purchases/{purchase_id}")
async def update_purchase(purchase_id: str, req: PurchaseUpdate, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot update purchase orders")

    existing = await db.purchases.find_one({"id": purchase_id, "tenant_id": user["tenant_id"]})
    if not existing:
        raise HTTPException(404, "Purchase order not found")

    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.purchases.update_one({"id": purchase_id, "tenant_id": user["tenant_id"]}, {"$set": updates})
    await log_audit(user["tenant_id"], user["id"], "purchase_updated", f"PO: {existing.get('po_number', purchase_id)}", request.client.host if request.client else "")
    updated = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    return updated

@api_router.post("/purchases/{purchase_id}/receive")
async def receive_purchase(purchase_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot receive purchases")

    purchase = await db.purchases.find_one({"id": purchase_id, "tenant_id": user["tenant_id"]})
    if not purchase:
        raise HTTPException(404, "Purchase order not found")
    if purchase.get("status") == "received":
        raise HTTPException(400, "Purchase already received")

    body = await request.json()
    received_items = body.get("items", [])

    for ri in received_items:
        pid = ri.get("product_id")
        recv_qty = ri.get("received_qty", 0)
        if not pid or recv_qty <= 0:
            continue

        # Update stock
        product = await db.products.find_one({"id": pid, "tenant_id": user["tenant_id"]})
        if product:
            await db.products.update_one(
                {"id": pid, "tenant_id": user["tenant_id"]},
                {"$inc": {"stock": recv_qty}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
            )

        # Update received_qty in purchase items
        for i, item in enumerate(purchase.get("items", [])):
            if item.get("product_id") == pid:
                await db.purchases.update_one(
                    {"id": purchase_id, f"items.{i}.product_id": pid},
                    {"$set": {f"items.{i}.received_qty": recv_qty}}
                )

    # Check if all items received
    all_received = all(
        ri.get("received_qty", 0) >= next((it["quantity"] for it in purchase["items"] if it.get("product_id") == ri.get("product_id")), 0)
        for ri in received_items
    ) if received_items else False

    new_status = "received" if all_received else "partial"
    await db.purchases.update_one(
        {"id": purchase_id, "tenant_id": user["tenant_id"]},
        {"$set": {"status": new_status, "received_at": datetime.now(timezone.utc).isoformat()}}
    )

    await log_audit(user["tenant_id"], user["id"], "purchase_received", f"PO: {purchase.get('po_number')}, Status: {new_status}", request.client.host if request.client else "")
    return {"message": f"Purchase {new_status}", "status": new_status}

@api_router.delete("/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")

    purchase = await db.purchases.find_one({"id": purchase_id, "tenant_id": user["tenant_id"]})
    if not purchase:
        raise HTTPException(404, "Purchase order not found")
    if purchase.get("status") == "received":
        raise HTTPException(400, "Cannot delete a received purchase order")

    await db.purchases.delete_one({"id": purchase_id, "tenant_id": user["tenant_id"]})
    await log_audit(user["tenant_id"], user["id"], "purchase_deleted", f"PO: {purchase.get('po_number')}", request.client.host if request.client else "")
    return {"message": "Purchase order deleted"}

# ═══════════════════════════════════════════════════════════════════
#  BARCODE & SCAN SESSION ROUTES
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/inventory/barcode/{barcode}")
async def lookup_barcode(barcode: str, request: Request):
    user = await get_current_user(request)
    product = await db.products.find_one({"tenant_id": user["tenant_id"], "barcode": barcode}, {"_id": 0})
    if not product:
        product = await db.products.find_one({"tenant_id": user["tenant_id"], "sku": barcode}, {"_id": 0})
    if not product:
        product = await db.products.find_one(
            {"tenant_id": user["tenant_id"], "$or": [
                {"barcode": {"$regex": barcode, "$options": "i"}},
                {"sku": {"$regex": barcode, "$options": "i"}}
            ]}, {"_id": 0}
        )
    if not product:
        raise HTTPException(404, f"Product not found for barcode/SKU: {barcode}")
    return product


# ─── External Barcode Lookup (UPCitemdb + Open Food Facts) ──────────
@api_router.get("/inventory/barcode-lookup/{barcode}")
async def external_barcode_lookup(barcode: str, request: Request):
    """Look up product info from external barcode databases.
    First checks local cache, then UPCitemdb, then Open Food Facts."""
    import httpx

    user = await get_current_user(request)
    barcode = barcode.strip()
    if not barcode:
        raise HTTPException(400, "Barcode is required")

    # 1. Check local inventory first
    product = await db.products.find_one({"tenant_id": user["tenant_id"], "barcode": barcode}, {"_id": 0})
    if not product:
        product = await db.products.find_one({"tenant_id": user["tenant_id"], "sku": barcode}, {"_id": 0})
    if product:
        return {"source": "local_inventory", "found": True, "product": product, "message": "Product already exists in your inventory"}

    # 2. Check cache (barcode_lookups collection) to avoid repeat API calls
    cached = await db.barcode_lookups.find_one({"barcode": barcode}, {"_id": 0})
    if cached and cached.get("found"):
        return {"source": "cache", "found": True, "product_info": cached.get("product_info", {}), "message": "Product info retrieved from cache"}

    product_info = {}
    source = "not_found"

    # 3. Try UPCitemdb (covers all product types)
    try:
        async with httpx.AsyncClient(timeout=8.0) as client:
            resp = await client.get(f"https://api.upcitemdb.com/prod/trial/lookup?upc={barcode}")
            if resp.status_code == 200:
                data = resp.json()
                if data.get("code") == "OK" and data.get("items") and len(data["items"]) > 0:
                    item = data["items"][0]
                    product_info = {
                        "name": item.get("title", ""),
                        "brand": item.get("brand", ""),
                        "description": item.get("description", ""),
                        "category": item.get("category", ""),
                        "weight": item.get("weight", ""),
                        "dimension": item.get("dimension", ""),
                        "model": item.get("model", ""),
                        "images": item.get("images", []),
                        "ean": item.get("ean", ""),
                        "upc": item.get("upc", ""),
                    }
                    # Extract a useful price hint from offers
                    offers = item.get("offers", [])
                    if offers:
                        prices = [o.get("price", 0) for o in offers if o.get("price", 0) > 0]
                        if prices:
                            product_info["price_hint"] = round(min(prices), 2)
                            product_info["price_currency"] = offers[0].get("currency", "")
                    source = "upcitemdb"
    except Exception as e:
        logging.warning(f"UPCitemdb lookup failed for {barcode}: {e}")

    # 4. If UPCitemdb didn't find it, try Open Food Facts
    if not product_info.get("name"):
        try:
            async with httpx.AsyncClient(timeout=8.0) as client:
                resp = await client.get(
                    f"https://world.openfoodfacts.org/api/v2/product/{barcode}.json",
                    params={"fields": "code,product_name,brands,categories,generic_name,quantity,image_url,ingredients_text"}
                )
                if resp.status_code == 200:
                    data = resp.json()
                    if data.get("status") == 1 and data.get("product"):
                        p = data["product"]
                        product_info = {
                            "name": p.get("product_name", "") or p.get("generic_name", ""),
                            "brand": p.get("brands", ""),
                            "description": p.get("generic_name", ""),
                            "category": p.get("categories", ""),
                            "weight": p.get("quantity", ""),
                            "images": [p["image_url"]] if p.get("image_url") else [],
                            "ingredients": p.get("ingredients_text", ""),
                        }
                        source = "openfoodfacts"
        except Exception as e:
            logging.warning(f"Open Food Facts lookup failed for {barcode}: {e}")

    # 5. Cache the result (whether found or not, to avoid repeated calls)
    found = bool(product_info.get("name"))
    await db.barcode_lookups.update_one(
        {"barcode": barcode},
        {"$set": {
            "barcode": barcode,
            "found": found,
            "source": source,
            "product_info": product_info,
            "looked_up_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )

    if found:
        return {"source": source, "found": True, "product_info": product_info, "message": f"Product info found via {source}"}
    else:
        return {"source": "not_found", "found": False, "product_info": {}, "message": "Product not found in any barcode database. Please enter details manually."}

@api_router.post("/scan/session")
async def create_scan_session(req: ScanSessionCreate, request: Request):
    user = await get_current_user(request)
    if req.type not in ["inventory", "pos"]:
        raise HTTPException(400, "Type must be 'inventory' or 'pos'")

    # Expire at 10 PM today (or 10 PM tomorrow if already past 10 PM)
    now = datetime.now(timezone.utc)
    # Use a fixed offset for IST (UTC+5:30) to calculate 10 PM IST
    ist_offset = timedelta(hours=5, minutes=30)
    ist_now = now + ist_offset
    expire_today = ist_now.replace(hour=22, minute=0, second=0, microsecond=0)
    if ist_now >= expire_today:
        expire_today += timedelta(days=1)
    expires_at = (expire_today - ist_offset).isoformat()

    # Deactivate any existing session of same type
    await db.scan_sessions.update_many(
        {"tenant_id": user["tenant_id"], "user_id": user["id"], "type": req.type, "is_active": True},
        {"$set": {"is_active": False}}
    )

    session_id = str(uuid.uuid4())[:12]
    doc = {
        "id": session_id, "tenant_id": user["tenant_id"], "user_id": user["id"],
        "type": req.type, "is_active": True, "expires_at": expires_at,
        "created_at": now.isoformat()
    }
    await db.scan_sessions.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "scan_session_created", f"Type: {req.type}, Session: {session_id}", request.client.host if request.client else "")
    doc.pop("_id", None)
    return doc

@api_router.get("/scan/session/{session_type}")
async def get_active_session(session_type: str, request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    session = await db.scan_sessions.find_one(
        {"tenant_id": user["tenant_id"], "user_id": user["id"], "type": session_type, "is_active": True, "expires_at": {"$gt": now}},
        {"_id": 0}
    )
    if not session:
        return {"session": None}
    return {"session": session}

@api_router.delete("/scan/session/{session_id}")
async def revoke_scan_session(session_id: str, request: Request):
    user = await get_current_user(request)
    await db.scan_sessions.update_one(
        {"id": session_id, "tenant_id": user["tenant_id"]},
        {"$set": {"is_active": False}}
    )
    return {"message": "Session revoked"}

# Mobile scanner endpoints (no auth needed - session validates)
@api_router.get("/scan/mobile/{session_id}")
async def get_mobile_session(session_id: str):
    now = datetime.now(timezone.utc).isoformat()
    session = await db.scan_sessions.find_one(
        {"id": session_id, "is_active": True, "expires_at": {"$gt": now}},
        {"_id": 0}
    )
    if not session:
        raise HTTPException(404, "Scan session expired or invalid")
    tenant = await db.tenants.find_one({"id": session["tenant_id"]}, {"_id": 0, "shop_name": 1})
    return {"type": session["type"], "shop_name": tenant.get("shop_name", ""), "expires_at": session["expires_at"]}

@api_router.post("/scan/mobile/{session_id}/barcode")
async def submit_mobile_barcode(session_id: str, request: Request):
    body = await request.json()
    barcode = body.get("barcode", "").strip()
    if not barcode:
        raise HTTPException(400, "Barcode is required")

    now = datetime.now(timezone.utc)
    session = await db.scan_sessions.find_one(
        {"id": session_id, "is_active": True, "expires_at": {"$gt": now.isoformat()}},
        {"_id": 0}
    )
    if not session:
        raise HTTPException(404, "Scan session expired or invalid")

    # Store the scanned barcode
    await db.scanned_barcodes.insert_one({
        "session_id": session_id, "tenant_id": session["tenant_id"],
        "barcode": barcode, "type": session["type"],
        "processed": False, "scanned_at": now.isoformat()
    })

    # Try to find the product by barcode, then by SKU
    product = await db.products.find_one({"tenant_id": session["tenant_id"], "barcode": barcode}, {"_id": 0})
    if not product:
        product = await db.products.find_one({"tenant_id": session["tenant_id"], "sku": barcode}, {"_id": 0})
    if not product:
        # Try partial match on barcode or sku
        product = await db.products.find_one(
            {"tenant_id": session["tenant_id"], "$or": [
                {"barcode": {"$regex": barcode, "$options": "i"}},
                {"sku": {"$regex": barcode, "$options": "i"}}
            ]}, {"_id": 0}
        )
    return {"message": "Barcode submitted", "product_found": product is not None, "product_name": product.get("name") if product else None, "barcode_value": barcode}

@api_router.get("/scan/poll/{session_id}")
async def poll_scanned_barcodes(session_id: str, request: Request):
    user = await get_current_user(request)
    # Get unprocessed barcodes for this session
    barcodes = await db.scanned_barcodes.find(
        {"session_id": session_id, "tenant_id": user["tenant_id"], "processed": False},
        {"_id": 0}
    ).sort("scanned_at", 1).to_list(10)

    # Mark as processed
    if barcodes:
        await db.scanned_barcodes.update_many(
            {"session_id": session_id, "tenant_id": user["tenant_id"], "processed": False},
            {"$set": {"processed": True}}
        )

    # Resolve products for each barcode (try barcode, then SKU, then partial)
    results = []
    for b in barcodes:
        product = await db.products.find_one({"tenant_id": user["tenant_id"], "barcode": b["barcode"]}, {"_id": 0})
        if not product:
            product = await db.products.find_one({"tenant_id": user["tenant_id"], "sku": b["barcode"]}, {"_id": 0})
        if not product:
            product = await db.products.find_one(
                {"tenant_id": user["tenant_id"], "$or": [
                    {"barcode": {"$regex": b["barcode"], "$options": "i"}},
                    {"sku": {"$regex": b["barcode"], "$options": "i"}}
                ]}, {"_id": 0}
            )
        results.append({"barcode": b["barcode"], "product": product, "scanned_at": b["scanned_at"]})

    return {"scans": results}

# ═══════════════════════════════════════════════════════════════════
#  INVOICE PDF GENERATION
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/pos/invoices/{invoice_id}/pdf")
async def generate_invoice_pdf(invoice_id: str, request: Request):
    user = await get_current_user(request)
    invoice = await db.invoices.find_one({"id": invoice_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    shop_name = tenant.get("shop_name", "RetailPro") if tenant else "RetailPro"
    shop_address = tenant.get("address", "") if tenant else ""
    shop_phone = tenant.get("phone", "") if tenant else ""
    shop_gst = tenant.get("gst_number", "") if tenant else ""

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, spaceAfter=2*mm, alignment=TA_CENTER, textColor=colors.HexColor('#0F172A'))
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor('#475569'))
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=11, spaceAfter=2*mm, textColor=colors.HexColor('#0F172A'), leading=14)
    small_style = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#64748B'))

    elements = []
    elements.append(Paragraph(shop_name, title_style))
    if shop_address:
        elements.append(Paragraph(shop_address, subtitle_style))
    shop_info = []
    if shop_phone:
        shop_info.append(f"Phone: {shop_phone}")
    if shop_gst:
        shop_info.append(f"GSTIN: {shop_gst}")
    if shop_info:
        elements.append(Paragraph(" | ".join(shop_info), subtitle_style))
    elements.append(Spacer(1, 4*mm))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#E5E7EB')))
    elements.append(Spacer(1, 4*mm))

    # Invoice details
    inv_date = invoice.get("created_at", "")[:19].replace("T", " ")
    elements.append(Paragraph(f"<b>Invoice #:</b> {invoice.get('invoice_number', '')}", header_style))
    elements.append(Paragraph(f"<b>Date:</b> {inv_date}", header_style))
    elements.append(Paragraph(f"<b>Customer:</b> {invoice.get('customer_name', 'Walk-in')}", header_style))
    if invoice.get("customer_phone"):
        elements.append(Paragraph(f"<b>Phone:</b> {invoice.get('customer_phone')}", header_style))
    elements.append(Paragraph(f"<b>Payment:</b> {invoice.get('payment_method', 'cash').upper()}", header_style))
    elements.append(Spacer(1, 4*mm))

    # Items table
    table_data = [['#', 'Item', 'Qty', 'Price', 'GST%', 'Total']]
    for i, item in enumerate(invoice.get("items", []), 1):
        table_data.append([
            str(i), item.get("name", ""), str(item.get("quantity", 0)),
            f"₹{item.get('price', 0):.2f}", f"{item.get('gst_rate', 0)}%",
            f"₹{item.get('total', 0):.2f}"
        ])

    t = Table(table_data, colWidths=[8*mm, 70*mm, 15*mm, 25*mm, 15*mm, 30*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F5F9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 4*mm))

    # Totals
    totals_data = [
        ['Subtotal:', f"₹{invoice.get('subtotal', 0):.2f}"],
        ['Tax (GST):', f"₹{invoice.get('tax_total', 0):.2f}"],
    ]
    if invoice.get("discount", 0) > 0:
        totals_data.append(['Discount:', f"-₹{invoice.get('discount', 0):.2f}"])
    totals_data.append(['GRAND TOTAL:', f"₹{invoice.get('grand_total', 0):.2f}"])

    tt = Table(totals_data, colWidths=[130*mm, 33*mm])
    tt.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#0F172A')),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(tt)

    if invoice.get("notes"):
        elements.append(Spacer(1, 6*mm))
        elements.append(Paragraph(f"<b>Notes:</b> {invoice['notes']}", small_style))

    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph("Thank you for your business!", ParagraphStyle('Thanks', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=colors.HexColor('#475569'))))

    doc.build(elements)
    buffer.seek(0)

    await log_audit(user["tenant_id"], user["id"], "invoice_pdf_generated", f"Invoice: {invoice.get('invoice_number')}", request.client.host if request.client else "")
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"inline; filename=invoice_{invoice.get('invoice_number', invoice_id)}.pdf"})

# ═══════════════════════════════════════════════════════════════════
#  CUSTOMER MANAGEMENT
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/customers")
async def list_customers(request: Request, search: str = "", page: int = 1, limit: int = 50):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]
    total = await db.customers.count_documents(query)
    skip = (page - 1) * limit
    customers = await db.customers.find(query, {"_id": 0}).sort("name", 1).skip(skip).limit(limit).to_list(limit)
    return {"customers": customers, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.post("/customers")
async def create_customer(req: CustomerCreate, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot manage customers")

    doc = req.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["tenant_id"] = user["tenant_id"]
    doc["credit_balance"] = 0.0
    doc["total_purchases"] = 0.0
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.customers.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "customer_created", f"Customer: {req.name}", request.client.host if request.client else "")
    doc.pop("_id", None)
    return doc

@api_router.get("/customers/{customer_id}")
async def get_customer(customer_id: str, request: Request):
    user = await get_current_user(request)
    customer = await db.customers.find_one({"id": customer_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not customer:
        raise HTTPException(404, "Customer not found")
    return customer

@api_router.put("/customers/{customer_id}")
async def update_customer(customer_id: str, req: CustomerUpdate, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot manage customers")

    existing = await db.customers.find_one({"id": customer_id, "tenant_id": user["tenant_id"]})
    if not existing:
        raise HTTPException(404, "Customer not found")

    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.customers.update_one({"id": customer_id, "tenant_id": user["tenant_id"]}, {"$set": updates})
    await log_audit(user["tenant_id"], user["id"], "customer_updated", f"Customer: {customer_id}", request.client.host if request.client else "")
    updated = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    return updated

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")

    result = await db.customers.delete_one({"id": customer_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Customer not found")
    await log_audit(user["tenant_id"], user["id"], "customer_deleted", f"Customer: {customer_id}", request.client.host if request.client else "")
    return {"message": "Customer deleted"}

@api_router.post("/customers/{customer_id}/credit")
async def adjust_credit(customer_id: str, req: CreditAdjust, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot manage credit")

    customer = await db.customers.find_one({"id": customer_id, "tenant_id": user["tenant_id"]})
    if not customer:
        raise HTTPException(404, "Customer not found")

    if req.type == "credit":
        new_balance = customer.get("credit_balance", 0) + req.amount
        if customer.get("credit_limit", 0) > 0 and new_balance > customer["credit_limit"]:
            raise HTTPException(400, f"Credit limit exceeded. Limit: {customer['credit_limit']}")
    elif req.type == "payment":
        new_balance = customer.get("credit_balance", 0) - req.amount
    else:
        raise HTTPException(400, "Type must be 'credit' or 'payment'")

    await db.customers.update_one(
        {"id": customer_id, "tenant_id": user["tenant_id"]},
        {"$set": {"credit_balance": round(new_balance, 2), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    txn = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        "customer_id": customer_id, "type": req.type,
        "amount": req.amount, "balance_after": round(new_balance, 2),
        "reference": req.reference, "notes": req.notes,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.customer_transactions.insert_one(txn)
    await log_audit(user["tenant_id"], user["id"], "credit_adjusted", f"Customer: {customer_id}, {req.type}: {req.amount}", request.client.host if request.client else "")
    txn.pop("_id", None)
    return txn

@api_router.get("/customers/{customer_id}/transactions")
async def customer_transactions(customer_id: str, request: Request, page: int = 1, limit: int = 50):
    user = await get_current_user(request)
    customer = await db.customers.find_one({"id": customer_id, "tenant_id": user["tenant_id"]})
    if not customer:
        raise HTTPException(404, "Customer not found")

    query = {"customer_id": customer_id, "tenant_id": user["tenant_id"]}
    total = await db.customer_transactions.count_documents(query)
    skip = (page - 1) * limit
    txns = await db.customer_transactions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"transactions": txns, "total": total, "page": page, "customer": {"name": customer.get("name"), "credit_balance": customer.get("credit_balance", 0)}}

@api_router.get("/customers/{customer_id}/invoices")
async def customer_invoices(customer_id: str, request: Request, page: int = 1, limit: int = 20):
    user = await get_current_user(request)
    query = {"customer_id": customer_id, "tenant_id": user["tenant_id"]}
    total = await db.invoices.count_documents(query)
    skip = (page - 1) * limit
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"invoices": invoices, "total": total}

# ═══════════════════════════════════════════════════════════════════
#  BATCH & EXPIRY TRACKING ALERTS
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/inventory/expiry-alerts")
async def expiry_alerts(request: Request, days: int = 90):
    user = await get_current_user(request)
    cutoff = (datetime.now(timezone.utc) + timedelta(days=days)).isoformat()[:10]
    today = datetime.now(timezone.utc).isoformat()[:10]

    products = await db.products.find(
        {"tenant_id": user["tenant_id"], "expiry_date": {"$nin": [None, ""], "$lte": cutoff}},
        {"_id": 0}
    ).sort("expiry_date", 1).to_list(200)

    expired = []
    expiring_soon = []
    for p in products:
        exp = p.get("expiry_date", "")
        if not exp:
            continue
        if exp <= today:
            p["status"] = "expired"
            expired.append(p)
        else:
            p["status"] = "expiring_soon"
            days_left = (datetime.fromisoformat(exp) - datetime.fromisoformat(today)).days
            p["days_until_expiry"] = days_left
            expiring_soon.append(p)

    return {"expired": expired, "expiring_soon": expiring_soon, "total_expired": len(expired), "total_expiring": len(expiring_soon)}

# ═══════════════════════════════════════════════════════════════════
#  ADVANCED REPORTS
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/reports/profit-margins")
async def profit_margins_report(request: Request, start_date: str = "", end_date: str = ""):
    user = await get_current_user(request)
    if not get_user_permission(user, "can_view_revenue"):
        await alert_unauthorized_access(user, "profit_margins_report", request)
        raise HTTPException(403, "You don't have permission to view revenue data. Contact your store owner.")

    query = {"tenant_id": user["tenant_id"]}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date

    invoices = await db.invoices.find(query, {"_id": 0}).to_list(5000)

    product_data = {}
    for inv in invoices:
        for item in inv.get("items", []):
            pid = item.get("product_id", "")
            if pid not in product_data:
                product_data[pid] = {"name": item["name"], "total_revenue": 0, "total_cost": 0, "total_qty": 0}
            product_data[pid]["total_revenue"] += item.get("total", item["price"] * item["quantity"])
            product_data[pid]["total_qty"] += item["quantity"]

    # Get cost prices
    for pid in product_data:
        product = await db.products.find_one({"id": pid, "tenant_id": user["tenant_id"]}, {"_id": 0, "cost_price": 1})
        if product:
            product_data[pid]["total_cost"] = product.get("cost_price", 0) * product_data[pid]["total_qty"]

    margins = []
    for pid, data in product_data.items():
        profit = data["total_revenue"] - data["total_cost"]
        margin_pct = (profit / data["total_revenue"] * 100) if data["total_revenue"] > 0 else 0
        margins.append({
            "product_id": pid, "name": data["name"],
            "revenue": round(data["total_revenue"], 2),
            "cost": round(data["total_cost"], 2),
            "profit": round(profit, 2),
            "margin_pct": round(margin_pct, 1),
            "quantity_sold": data["total_qty"]
        })

    margins.sort(key=lambda x: x["profit"], reverse=True)
    total_revenue = sum(m["revenue"] for m in margins)
    total_cost = sum(m["cost"] for m in margins)
    total_profit = sum(m["profit"] for m in margins)
    overall_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0

    return {
        "products": margins,
        "summary": {
            "total_revenue": round(total_revenue, 2),
            "total_cost": round(total_cost, 2),
            "total_profit": round(total_profit, 2),
            "overall_margin_pct": round(overall_margin, 1)
        }
    }

@api_router.get("/reports/category-analysis")
async def category_analysis_report(request: Request, start_date: str = "", end_date: str = ""):
    user = await get_current_user(request)
    if not get_user_permission(user, "can_view_revenue"):
        await alert_unauthorized_access(user, "category_analysis_report", request)
        raise HTTPException(403, "You don't have permission to view revenue data. Contact your store owner.")

    query = {"tenant_id": user["tenant_id"]}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date

    invoices = await db.invoices.find(query, {"_id": 0}).to_list(5000)

    # Build product->category map
    products_list = await db.products.find({"tenant_id": user["tenant_id"]}, {"_id": 0, "id": 1, "category": 1, "cost_price": 1}).to_list(5000)
    prod_map = {p["id"]: p for p in products_list}

    category_data = {}
    for inv in invoices:
        for item in inv.get("items", []):
            pid = item.get("product_id", "")
            cat = prod_map.get(pid, {}).get("category", "Uncategorized") or "Uncategorized"
            if cat not in category_data:
                category_data[cat] = {"revenue": 0, "cost": 0, "qty": 0, "invoices": set(), "products": set()}
            category_data[cat]["revenue"] += item.get("total", item["price"] * item["quantity"])
            cost_price = prod_map.get(pid, {}).get("cost_price", 0)
            category_data[cat]["cost"] += cost_price * item["quantity"]
            category_data[cat]["qty"] += item["quantity"]
            category_data[cat]["invoices"].add(inv.get("id", ""))
            category_data[cat]["products"].add(pid)

    categories = []
    total_revenue = sum(d["revenue"] for d in category_data.values())
    for cat, data in category_data.items():
        profit = data["revenue"] - data["cost"]
        share = (data["revenue"] / total_revenue * 100) if total_revenue > 0 else 0
        categories.append({
            "category": cat,
            "revenue": round(data["revenue"], 2),
            "cost": round(data["cost"], 2),
            "profit": round(profit, 2),
            "quantity_sold": data["qty"],
            "invoice_count": len(data["invoices"]),
            "product_count": len(data["products"]),
            "revenue_share_pct": round(share, 1)
        })

    categories.sort(key=lambda x: x["revenue"], reverse=True)
    return {"categories": categories, "total_revenue": round(total_revenue, 2)}

@api_router.get("/reports/purchase-analytics")
async def purchase_analytics_report(request: Request, start_date: str = "", end_date: str = ""):
    user = await get_current_user(request)
    if not get_user_permission(user, "can_view_revenue"):
        await alert_unauthorized_access(user, "purchase_analytics_report", request)
        raise HTTPException(403, "You don't have permission to view revenue data. Contact your store owner.")

    query = {"tenant_id": user["tenant_id"]}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date

    purchases = await db.purchases.find(query, {"_id": 0}).to_list(5000)

    supplier_data = {}
    monthly_spending = {}
    for po in purchases:
        sid = po.get("supplier_id", "")
        sname = po.get("supplier_name", "Unknown")
        if sid not in supplier_data:
            supplier_data[sid] = {"name": sname, "total_spent": 0, "total_tax": 0, "order_count": 0, "items_ordered": 0, "statuses": {}}
        supplier_data[sid]["total_spent"] += po.get("grand_total", 0)
        supplier_data[sid]["total_tax"] += po.get("tax_total", 0)
        supplier_data[sid]["order_count"] += 1
        supplier_data[sid]["items_ordered"] += sum(i.get("quantity", 0) for i in po.get("items", []))
        status = po.get("status", "pending")
        supplier_data[sid]["statuses"][status] = supplier_data[sid]["statuses"].get(status, 0) + 1

        # Monthly spending
        month = po.get("created_at", "")[:7]
        if month:
            monthly_spending[month] = monthly_spending.get(month, 0) + po.get("grand_total", 0)

    suppliers = []
    for sid, data in supplier_data.items():
        avg_order = data["total_spent"] / data["order_count"] if data["order_count"] > 0 else 0
        suppliers.append({
            "supplier_id": sid, "name": data["name"],
            "total_spent": round(data["total_spent"], 2),
            "total_tax": round(data["total_tax"], 2),
            "order_count": data["order_count"],
            "avg_order_value": round(avg_order, 2),
            "items_ordered": data["items_ordered"],
            "statuses": data["statuses"]
        })
    suppliers.sort(key=lambda x: x["total_spent"], reverse=True)

    monthly = [{"month": k, "spending": round(v, 2)} for k, v in sorted(monthly_spending.items())]

    total_spent = sum(s["total_spent"] for s in suppliers)
    return {"suppliers": suppliers, "monthly_spending": monthly, "total_spent": round(total_spent, 2), "total_orders": len(purchases)}

# ═══════════════════════════════════════════════════════════════════
#  EXPORT DATA (CSV / EXCEL)
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/export/inventory")
async def export_inventory(request: Request, format: str = "csv"):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Insufficient permissions")

    products = await db.products.find({"tenant_id": user["tenant_id"]}, {"_id": 0, "tenant_id": 0}).sort("name", 1).to_list(10000)
    headers = ["Name", "SKU", "Barcode", "Category", "Price", "Cost Price", "Stock", "Low Stock Threshold", "Unit", "Batch", "Expiry Date", "HSN Code", "GST Rate"]
    keys = ["name", "sku", "barcode", "category", "price", "cost_price", "stock", "low_stock_threshold", "unit", "batch_number", "expiry_date", "hsn_code", "gst_rate"]

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        ws.append(headers)
        for p in products:
            ws.append([p.get(k, "") for k in keys])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               headers={"Content-Disposition": "attachment; filename=inventory.xlsx"})
    else:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        for p in products:
            writer.writerow([p.get(k, "") for k in keys])
        buf.seek(0)
        return StreamingResponse(io.BytesIO(buf.getvalue().encode()), media_type="text/csv",
                               headers={"Content-Disposition": "attachment; filename=inventory.csv"})

@api_router.get("/export/invoices")
async def export_invoices(request: Request, format: str = "csv", start_date: str = "", end_date: str = ""):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Insufficient permissions")

    query = {"tenant_id": user["tenant_id"]}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date

    invoices = await db.invoices.find(query, {"_id": 0, "tenant_id": 0}).sort("created_at", -1).to_list(10000)
    headers = ["Invoice #", "Date", "Customer", "Phone", "Subtotal", "Tax", "Discount", "Grand Total", "Payment Method", "Device", "Status"]
    keys = ["invoice_number", "created_at", "customer_name", "customer_phone", "subtotal", "tax_total", "discount", "grand_total", "payment_method", "device_source", "status"]

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append(headers)
        for inv in invoices:
            ws.append([inv.get(k, "") for k in keys])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               headers={"Content-Disposition": "attachment; filename=invoices.xlsx"})
    else:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        for inv in invoices:
            writer.writerow([inv.get(k, "") for k in keys])
        buf.seek(0)
        return StreamingResponse(io.BytesIO(buf.getvalue().encode()), media_type="text/csv",
                               headers={"Content-Disposition": "attachment; filename=invoices.csv"})

@api_router.get("/export/customers")
async def export_customers(request: Request, format: str = "csv"):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Insufficient permissions")

    customers = await db.customers.find({"tenant_id": user["tenant_id"]}, {"_id": 0, "tenant_id": 0}).sort("name", 1).to_list(10000)
    headers = ["Name", "Phone", "Email", "Address", "GST Number", "Credit Balance", "Credit Limit", "Total Purchases"]
    keys = ["name", "phone", "email", "address", "gst_number", "credit_balance", "credit_limit", "total_purchases"]

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.append(headers)
        for c in customers:
            ws.append([c.get(k, "") for k in keys])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               headers={"Content-Disposition": "attachment; filename=customers.xlsx"})
    else:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        for c in customers:
            writer.writerow([c.get(k, "") for k in keys])
        buf.seek(0)
        return StreamingResponse(io.BytesIO(buf.getvalue().encode()), media_type="text/csv",
                               headers={"Content-Disposition": "attachment; filename=customers.csv"})

# ═══════════════════════════════════════════════════════════════════
#  API KEY MANAGEMENT (External API Access)
# ═══════════════════════════════════════════════════════════════════

VALID_API_PERMISSIONS = ["read_inventory", "read_invoices", "read_customers", "read_reports", "write_inventory", "create_invoice"]

async def get_api_key_user(request: Request) -> dict:
    """Authenticate via API key for external access"""
    api_key = request.headers.get("X-API-Key", "")
    if not api_key:
        raise HTTPException(401, "API key required")

    key_hash = hashlib.sha256(api_key.encode()).hexdigest()
    key_doc = await db.api_keys.find_one({"key_hash": key_hash, "is_active": True}, {"_id": 0})
    if not key_doc:
        raise HTTPException(401, "Invalid API key")

    await db.api_keys.update_one({"key_hash": key_hash}, {"$set": {"last_used": datetime.now(timezone.utc).isoformat()}})
    return key_doc

@api_router.get("/admin/api-keys")
async def list_api_keys(request: Request):
    user = await get_current_user(request)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can manage API keys")

    keys = await db.api_keys.find({"tenant_id": user["tenant_id"]}, {"_id": 0, "key_hash": 0}).sort("created_at", -1).to_list(100)
    return {"api_keys": keys}

@api_router.post("/admin/api-keys")
async def create_api_key(req: APIKeyCreate, request: Request):
    user = await get_current_user(request)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can create API keys")

    # Validate permissions
    for p in req.permissions:
        if p not in VALID_API_PERMISSIONS:
            raise HTTPException(400, f"Invalid permission: {p}")

    raw_key = f"rpk_{secrets.token_hex(24)}"
    key_hash = hashlib.sha256(raw_key.encode()).hexdigest()
    key_prefix = raw_key[:12] + "..."

    doc = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        "name": req.name, "key_hash": key_hash, "key_prefix": key_prefix,
        "permissions": req.permissions, "created_by": user["id"],
        "is_active": True, "last_used": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.api_keys.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "api_key_created", f"Key: {req.name}", request.client.host if request.client else "")
    doc.pop("_id", None)
    doc.pop("key_hash", None)
    return {"api_key": raw_key, "details": doc}

@api_router.delete("/admin/api-keys/{key_id}")
async def delete_api_key(key_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can delete API keys")

    result = await db.api_keys.update_one({"id": key_id, "tenant_id": user["tenant_id"]}, {"$set": {"is_active": False}})
    if result.modified_count == 0:
        raise HTTPException(404, "API key not found")
    await log_audit(user["tenant_id"], user["id"], "api_key_deleted", f"Key ID: {key_id}", request.client.host if request.client else "")
    return {"message": "API key revoked"}

# External API endpoints (authenticated via API key)
@api_router.get("/external/inventory")
async def external_inventory(request: Request, search: str = "", page: int = 1, limit: int = 50):
    key = await get_api_key_user(request)
    if "read_inventory" not in key.get("permissions", []):
        raise HTTPException(403, "Insufficient permissions")

    query = {"tenant_id": key["tenant_id"]}
    if search:
        query["$or"] = [{"name": {"$regex": search, "$options": "i"}}, {"sku": {"$regex": search, "$options": "i"}}]
    total = await db.products.count_documents(query)
    skip = (page - 1) * limit
    products = await db.products.find(query, {"_id": 0, "tenant_id": 0}).sort("name", 1).skip(skip).limit(limit).to_list(limit)
    return {"products": products, "total": total, "page": page}

@api_router.get("/external/invoices")
async def external_invoices(request: Request, page: int = 1, limit: int = 20):
    key = await get_api_key_user(request)
    if "read_invoices" not in key.get("permissions", []):
        raise HTTPException(403, "Insufficient permissions")

    query = {"tenant_id": key["tenant_id"]}
    total = await db.invoices.count_documents(query)
    skip = (page - 1) * limit
    invoices = await db.invoices.find(query, {"_id": 0, "tenant_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"invoices": invoices, "total": total, "page": page}

@api_router.get("/external/customers")
async def external_customers(request: Request, page: int = 1, limit: int = 50):
    key = await get_api_key_user(request)
    if "read_customers" not in key.get("permissions", []):
        raise HTTPException(403, "Insufficient permissions")

    query = {"tenant_id": key["tenant_id"]}
    total = await db.customers.count_documents(query)
    skip = (page - 1) * limit
    customers = await db.customers.find(query, {"_id": 0, "tenant_id": 0}).sort("name", 1).skip(skip).limit(limit).to_list(limit)
    return {"customers": customers, "total": total, "page": page}

# ═══════════════════════════════════════════════════════════════════
#  IP WHITELIST & USER ACTIVITY
# ═══════════════════════════════════════════════════════════════════

@api_router.put("/admin/users/{user_id}/ip-whitelist")
async def update_ip_whitelist(user_id: str, req: IPWhitelistUpdate, request: Request):
    admin = await get_current_user(request)
    if admin["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can manage IP restrictions")

    target = await db.users.find_one({"_id": ObjectId(user_id), "tenant_id": admin["tenant_id"]})
    if not target:
        raise HTTPException(404, "User not found")

    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"allowed_ips": req.allowed_ips}})
    await log_audit(admin["tenant_id"], admin["id"], "ip_whitelist_updated", f"User: {user_id}, IPs: {req.allowed_ips}", request.client.host if request.client else "")
    return {"message": "IP whitelist updated", "allowed_ips": req.allowed_ips}

@api_router.get("/admin/users/{user_id}/ip-whitelist")
async def get_ip_whitelist(user_id: str, request: Request):
    admin = await get_current_user(request)
    if admin["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can view IP restrictions")

    target = await db.users.find_one({"_id": ObjectId(user_id), "tenant_id": admin["tenant_id"]})
    if not target:
        raise HTTPException(404, "User not found")
    return {"allowed_ips": target.get("allowed_ips", [])}

@api_router.get("/admin/user-activity")
async def get_user_activity(request: Request, user_id: str = "", page: int = 1, limit: int = 100):
    admin = await get_current_user(request)
    if admin["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")

    query = {"tenant_id": admin["tenant_id"]}
    if user_id:
        query["user_id"] = user_id

    total = await db.audit_logs.count_documents(query)
    skip = (page - 1) * limit
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)

    # Get user info for each log
    user_ids = list(set(entry.get("user_id") for entry in logs if entry.get("user_id")))
    users_map = {}
    for uid in user_ids:
        try:
            u = await db.users.find_one({"_id": ObjectId(uid)}, {"_id": 0, "name": 1, "email": 1, "role": 1})
            if u:
                users_map[uid] = u
        except Exception:
            pass

    for log in logs:
        uid = log.get("user_id", "")
        log["user_name"] = users_map.get(uid, {}).get("name", "Unknown")
        log["user_email"] = users_map.get(uid, {}).get("email", "")
        log["user_role"] = users_map.get(uid, {}).get("role", "")

    return {"activities": logs, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.get("/admin/users-status")
async def get_users_status(request: Request):
    """Get all users with their last activity and login info for owner dashboard"""
    admin = await get_current_user(request)
    if admin["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can view user status")

    users = await db.users.find(
        {"tenant_id": admin["tenant_id"]},
        {"password_hash": 0, "mfa_secret": 0}
    ).to_list(100)

    result = []
    now = datetime.now(timezone.utc)
    for u in users:
        u_data = serialize_doc(u)
        last_activity = u.get("last_activity", "")
        if last_activity:
            try:
                last_dt = datetime.fromisoformat(last_activity.replace("Z", "+00:00")) if isinstance(last_activity, str) else last_activity
                idle_minutes = (now - last_dt).total_seconds() / 60
                u_data["idle_minutes"] = round(idle_minutes, 1)
                u_data["is_idle"] = idle_minutes > 60
            except Exception:
                u_data["idle_minutes"] = None
                u_data["is_idle"] = None
        result.append(u_data)

    return {"users": result}

# ═══════════════════════════════════════════════════════════════════
#  SUPPORT TICKET SYSTEM
# ═══════════════════════════════════════════════════════════════════

SUPPORT_CONTACT_INFO = {
    "email": "support@retailpro.com",
    "phone": "+91-1800-RETAIL",
    "whatsapp": "+919876543210",
    "whatsapp_url": "https://wa.me/919876543210",
    "hours": "Mon-Sat 9AM-6PM IST"
}

@api_router.get("/support/contact-info")
async def get_contact_info(request: Request):
    await get_current_user(request)
    return SUPPORT_CONTACT_INFO

@api_router.post("/support/tickets")
async def create_support_ticket(req: SupportTicketCreate, request: Request):
    user = await get_current_user(request)
    if req.channel not in ["email", "phone", "whatsapp"]:
        raise HTTPException(400, "Channel must be email, phone, or whatsapp")
    if req.priority not in ["low", "normal", "high", "urgent"]:
        raise HTTPException(400, "Priority must be low, normal, high, or urgent")

    doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": user["tenant_id"],
        "owner_id": user["id"],
        "owner_name": user.get("name", ""),
        "owner_email": user.get("email", ""),
        "subject": req.subject,
        "description": req.description,
        "channel": req.channel,
        "priority": req.priority,
        "status": "open",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.support_tickets.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "support_ticket_created", f"Ticket: {doc['id']}, Subject: {req.subject}", request.client.host if request.client else "", "support")
    doc.pop("_id", None)
    return doc

@api_router.get("/support/tickets")
async def list_support_tickets(request: Request, status: str = "", page: int = 1, limit: int = 20):
    user = await get_current_user(request)

    # Platform admin and admin see all tickets, owners see their own tenant's tickets
    if user.get("is_platform_admin") or user.get("is_admin"):
        query = {}
    else:
        query = {"tenant_id": user["tenant_id"]}

    if status:
        query["status"] = status

    total = await db.support_tickets.count_documents(query)
    skip = (page - 1) * limit
    tickets = await db.support_tickets.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"tickets": tickets, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.get("/support/tickets/{ticket_id}")
async def get_support_ticket(ticket_id: str, request: Request):
    user = await get_current_user(request)

    if user.get("is_platform_admin") or user.get("is_admin"):
        ticket = await db.support_tickets.find_one({"id": ticket_id}, {"_id": 0})
    else:
        ticket = await db.support_tickets.find_one({"id": ticket_id, "tenant_id": user["tenant_id"]}, {"_id": 0})

    if not ticket:
        raise HTTPException(404, "Ticket not found")

    # Get notes for this ticket
    notes = await db.ticket_notes.find({"ticket_id": ticket_id}, {"_id": 0}).sort("created_at", 1).to_list(500)
    ticket["notes"] = notes
    return ticket

@api_router.put("/support/tickets/{ticket_id}/status")
async def update_ticket_status(ticket_id: str, req: TicketStatusUpdate, request: Request):
    user = await get_current_user(request)
    if req.status not in ["open", "in_progress", "resolved", "closed"]:
        raise HTTPException(400, "Status must be open, in_progress, resolved, or closed")

    if user.get("is_platform_admin") or user.get("is_admin"):
        ticket = await db.support_tickets.find_one({"id": ticket_id})
    else:
        ticket = await db.support_tickets.find_one({"id": ticket_id, "tenant_id": user["tenant_id"]})

    if not ticket:
        raise HTTPException(404, "Ticket not found")

    # Non-admin users can only close/reopen their own tickets
    if not user.get("is_platform_admin") and not user.get("is_admin") and req.status == "in_progress":
        raise HTTPException(403, "Only admin can set in_progress status")

    await db.support_tickets.update_one(
        {"id": ticket_id},
        {"$set": {"status": req.status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    tid = ticket.get("tenant_id", user.get("tenant_id", ""))
    await log_audit(tid, user["id"], "ticket_status_updated", f"Ticket: {ticket_id}, Status: {req.status}", request.client.host if request.client else "", "support")
    return {"message": "Ticket status updated", "status": req.status}

@api_router.post("/support/tickets/{ticket_id}/notes")
async def add_ticket_note(ticket_id: str, req: TicketNoteCreate, request: Request):
    user = await get_current_user(request)

    if user.get("is_platform_admin") or user.get("is_admin"):
        ticket = await db.support_tickets.find_one({"id": ticket_id})
    else:
        ticket = await db.support_tickets.find_one({"id": ticket_id, "tenant_id": user["tenant_id"]})

    if not ticket:
        raise HTTPException(404, "Ticket not found")

    note = {
        "id": str(uuid.uuid4()),
        "ticket_id": ticket_id,
        "author_id": user["id"],
        "author_name": user.get("name", ""),
        "author_type": "admin" if (user.get("is_platform_admin") or user.get("is_admin")) else "owner",
        "message": req.message,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.ticket_notes.insert_one(note)

    # Update ticket's updated_at
    await db.support_tickets.update_one({"id": ticket_id}, {"$set": {"updated_at": datetime.now(timezone.utc).isoformat()}})

    note.pop("_id", None)
    return note

# ═══════════════════════════════════════════════════════════════════
#  FINANCIAL ACCESS REQUEST SYSTEM
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/platform/access-requests")
async def create_access_request(req: AccessRequestCreate, request: Request):
    """Platform admin or admin requests access to tenant financial data"""
    admin = await get_platform_or_admin(request)

    # Verify the target owner/tenant exists
    target_owner = await db.users.find_one({"_id": ObjectId(req.owner_id), "tenant_id": req.tenant_id})
    if not target_owner:
        raise HTTPException(404, "Target owner/tenant not found")

    if req.request_type not in ["revenue", "transactions", "full_financial"]:
        raise HTTPException(400, "request_type must be revenue, transactions, or full_financial")

    # Check for existing pending request
    existing = await db.access_requests.find_one({
        "admin_id": admin["id"], "tenant_id": req.tenant_id,
        "status": "pending"
    })
    if existing:
        raise HTTPException(400, "You already have a pending request for this tenant")

    expires_at = (datetime.now(timezone.utc) + timedelta(hours=req.duration_hours)).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "admin_id": admin["id"],
        "admin_name": admin.get("name", ""),
        "owner_id": req.owner_id,
        "tenant_id": req.tenant_id,
        "request_type": req.request_type,
        "reason": req.reason,
        "duration_hours": req.duration_hours,
        "status": "pending",
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "responded_at": None,
        "response_note": ""
    }
    await db.access_requests.insert_one(doc)
    await log_audit(req.tenant_id, admin["id"], "financial_access_requested",
                    f"Admin {admin['id']} requested {req.request_type} access. Reason: {req.reason}",
                    request.client.host if request.client else "", "data_access")
    doc.pop("_id", None)
    return doc

@api_router.get("/access-requests")
async def list_access_requests(request: Request, status: str = "", page: int = 1, limit: int = 20):
    """Owners see pending requests for their tenant, admins see their own requests"""
    user = await get_current_user(request)

    if user.get("is_platform_admin") or user.get("is_admin"):
        # Admins/platform admins see all requests (or their own)
        query = {"admin_id": user["id"]}
    elif user["role"] == "OWNER":
        query = {"tenant_id": user["tenant_id"]}
    else:
        raise HTTPException(403, "Only OWNER can view access requests")

    if status:
        query["status"] = status

    total = await db.access_requests.count_documents(query)
    skip = (page - 1) * limit
    requests_list = await db.access_requests.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"requests": requests_list, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.put("/access-requests/{request_id}/respond")
async def respond_to_access_request(request_id: str, req: AccessRequestRespond, request: Request):
    """Owner approves or rejects a financial access request"""
    user = await get_current_user(request)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can respond to access requests")

    if req.action not in ["approve", "reject"]:
        raise HTTPException(400, "Action must be approve or reject")

    access_req = await db.access_requests.find_one({"id": request_id, "tenant_id": user["tenant_id"], "status": "pending"})
    if not access_req:
        raise HTTPException(404, "Pending access request not found")

    new_status = "approved" if req.action == "approve" else "rejected"
    update_data = {
        "status": new_status,
        "responded_at": datetime.now(timezone.utc).isoformat(),
        "response_note": req.response_note
    }
    if req.action == "approve":
        # Set expiry from now based on requested duration
        update_data["expires_at"] = (datetime.now(timezone.utc) + timedelta(hours=access_req.get("duration_hours", 24))).isoformat()

    await db.access_requests.update_one({"id": request_id}, {"$set": update_data})
    await log_audit(user["tenant_id"], user["id"], f"access_request_{new_status}",
                    f"Request {request_id} {new_status}. Note: {req.response_note}",
                    request.client.host if request.client else "", "data_access")
    return {"message": f"Access request {new_status}", "status": new_status}

@api_router.get("/platform/tenant-financials/{tenant_id}")
async def get_tenant_financials(tenant_id: str, request: Request):
    """Platform admin or admin views tenant financial data (requires approved access)"""
    admin = await get_platform_or_admin(request)

    # Check for approved, non-expired access
    now = datetime.now(timezone.utc).isoformat()
    access = await db.access_requests.find_one({
        "admin_id": admin["id"], "tenant_id": tenant_id,
        "status": "approved", "expires_at": {"$gt": now}
    })
    if not access:
        raise HTTPException(403, "No approved access for this tenant. Submit an access request first.")

    request_type = access.get("request_type", "revenue")
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(404, "Tenant not found")

    result = {"tenant": tenant, "access_type": request_type, "access_expires_at": access["expires_at"]}

    if request_type in ["revenue", "full_financial"]:
        # Revenue summary
        invoices = await db.invoices.find({"tenant_id": tenant_id}, {"_id": 0, "grand_total": 1, "created_at": 1}).to_list(10000)
        total_revenue = sum(inv.get("grand_total", 0) for inv in invoices)
        result["total_revenue"] = round(total_revenue, 2)
        result["total_invoices"] = len(invoices)

    if request_type in ["transactions", "full_financial"]:
        # Recent transactions
        recent_invoices = await db.invoices.find({"tenant_id": tenant_id}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
        result["recent_transactions"] = recent_invoices

    await log_audit(tenant_id, admin["id"], "financial_data_accessed",
                    f"Admin {admin['id']} accessed {request_type} data for tenant {tenant_id}",
                    request.client.host if request.client else "", "data_access")
    return result

# ═══════════════════════════════════════════════════════════════════
#  ANALYTICS & MONITORING SYSTEM
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/analytics/owner/overview")
async def analytics_owner_overview(request: Request, period: str = "30d"):
    """Comprehensive owner analytics dashboard data."""
    user = await get_current_user(request)
    if user["role"] not in ("OWNER", "MANAGER"):
        raise HTTPException(403, "Analytics requires OWNER or MANAGER role")
    tid = user["tenant_id"]

    # Determine date range
    days = {"7d": 7, "30d": 30, "90d": 90, "365d": 365}.get(period, 30)
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    prev_start = (datetime.now(timezone.utc) - timedelta(days=days * 2)).isoformat()
    prev_end = start_date

    # --- Business KPIs ---
    total_invoices = await db.invoices.count_documents({"tenant_id": tid, "created_at": {"$gte": start_date}})
    prev_invoices = await db.invoices.count_documents({"tenant_id": tid, "created_at": {"$gte": prev_start, "$lt": prev_end}})

    invoices_cursor = db.invoices.find({"tenant_id": tid, "created_at": {"$gte": start_date}}, {"_id": 0, "grand_total": 1, "items": 1, "discount": 1, "created_at": 1})
    invoices_list = await invoices_cursor.to_list(10000)
    total_revenue = sum(inv.get("grand_total", 0) for inv in invoices_list)
    total_items_sold = sum(len(inv.get("items", [])) for inv in invoices_list)
    avg_order_value = round(total_revenue / max(total_invoices, 1), 2)
    total_discount = sum(inv.get("discount", 0) for inv in invoices_list)

    prev_invoices_list = await db.invoices.find({"tenant_id": tid, "created_at": {"$gte": prev_start, "$lt": prev_end}}, {"_id": 0, "grand_total": 1}).to_list(10000)
    prev_revenue = sum(inv.get("grand_total", 0) for inv in prev_invoices_list)
    prev_aov = round(prev_revenue / max(prev_invoices, 1), 2)

    # --- Product Analytics ---
    total_products = await db.products.count_documents({"tenant_id": tid})
    low_stock = await db.products.count_documents({"tenant_id": tid, "$expr": {"$lte": ["$stock", "$low_stock_threshold"]}})
    out_of_stock = await db.products.count_documents({"tenant_id": tid, "stock": {"$lte": 0}})

    # --- Customer Analytics ---
    total_customers = await db.customers.count_documents({"tenant_id": tid})
    new_customers = await db.customers.count_documents({"tenant_id": tid, "created_at": {"$gte": start_date}})

    # --- Active Users ---
    active_user_ids = await db.api_analytics.distinct("user_id", {"tenant_id": tid, "date": {"$gte": (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")}})
    active_users = len([u for u in active_user_ids if u])

    # --- API Usage ---
    api_calls = await db.api_analytics.count_documents({"tenant_id": tid, "timestamp": {"$gte": start_date}})
    prev_api_calls = await db.api_analytics.count_documents({"tenant_id": tid, "timestamp": {"$gte": prev_start, "$lt": prev_end}})

    # --- Avg Response Time ---
    perf_pipeline = [
        {"$match": {"tenant_id": tid, "timestamp": {"$gte": start_date}}},
        {"$group": {"_id": None, "avg_ms": {"$avg": "$duration_ms"}, "p95_ms": {"$percentile": {"input": "$duration_ms", "p": [0.95], "method": "approximate"}}}},
    ]
    try:
        perf_results = await db.api_analytics.aggregate(perf_pipeline).to_list(1)
        avg_response_ms = round(perf_results[0]["avg_ms"], 1) if perf_results else 0
    except Exception:
        avg_response_ms = 0

    return {
        "period": period, "days": days,
        "kpis": {
            "total_revenue": round(total_revenue, 2),
            "prev_revenue": round(prev_revenue, 2),
            "revenue_change": round(((total_revenue - prev_revenue) / max(prev_revenue, 1)) * 100, 1),
            "total_invoices": total_invoices,
            "prev_invoices": prev_invoices,
            "invoice_change": round(((total_invoices - prev_invoices) / max(prev_invoices, 1)) * 100, 1),
            "avg_order_value": avg_order_value,
            "prev_aov": prev_aov,
            "aov_change": round(((avg_order_value - prev_aov) / max(prev_aov, 1)) * 100, 1),
            "total_items_sold": total_items_sold,
            "total_discount": round(total_discount, 2),
            "total_products": total_products,
            "low_stock": low_stock,
            "out_of_stock": out_of_stock,
            "total_customers": total_customers,
            "new_customers": new_customers,
            "active_users": active_users,
            "api_calls": api_calls,
            "prev_api_calls": prev_api_calls,
            "api_change": round(((api_calls - prev_api_calls) / max(prev_api_calls, 1)) * 100, 1),
            "avg_response_ms": avg_response_ms,
        }
    }


@api_router.get("/analytics/owner/revenue-trend")
async def analytics_revenue_trend(request: Request, period: str = "30d", granularity: str = "daily"):
    """Revenue time-series data."""
    user = await get_current_user(request)
    if user["role"] not in ("OWNER", "MANAGER"):
        raise HTTPException(403, "Requires OWNER/MANAGER")
    tid = user["tenant_id"]
    days = {"7d": 7, "30d": 30, "90d": 90, "365d": 365}.get(period, 30)

    trend = []
    for i in range(days - 1, -1, -1):
        day = datetime.now(timezone.utc) - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        day_inv = await db.invoices.find(
            {"tenant_id": tid, "created_at": {"$gte": day_start.isoformat(), "$lt": day_end.isoformat()}},
            {"_id": 0, "grand_total": 1}
        ).to_list(5000)
        rev = sum(inv.get("grand_total", 0) for inv in day_inv)
        trend.append({"date": day_start.strftime("%Y-%m-%d"), "label": day_start.strftime("%b %d"), "revenue": round(rev, 2), "orders": len(day_inv)})

    return {"trend": trend, "period": period}


@api_router.get("/analytics/owner/top-products")
async def analytics_top_products(request: Request, period: str = "30d", limit: int = 10):
    """Top products by revenue and quantity sold."""
    user = await get_current_user(request)
    if user["role"] not in ("OWNER", "MANAGER"):
        raise HTTPException(403, "Requires OWNER/MANAGER")
    tid = user["tenant_id"]
    days = {"7d": 7, "30d": 30, "90d": 90, "365d": 365}.get(period, 30)
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    invoices = await db.invoices.find({"tenant_id": tid, "created_at": {"$gte": start_date}}, {"_id": 0, "items": 1}).to_list(10000)
    product_sales = {}
    for inv in invoices:
        for item in inv.get("items", []):
            pid = item.get("product_id", item.get("name", "Unknown"))
            name = item.get("name", "Unknown")
            if pid not in product_sales:
                product_sales[pid] = {"name": name, "qty": 0, "revenue": 0}
            product_sales[pid]["qty"] += item.get("quantity", 0)
            product_sales[pid]["revenue"] += item.get("quantity", 0) * item.get("price", 0)

    by_revenue = sorted(product_sales.values(), key=lambda x: x["revenue"], reverse=True)[:limit]
    by_qty = sorted(product_sales.values(), key=lambda x: x["qty"], reverse=True)[:limit]
    return {"by_revenue": by_revenue, "by_quantity": by_qty}


@api_router.get("/analytics/owner/customer-insights")
async def analytics_customer_insights(request: Request, period: str = "30d"):
    """Customer behavior analytics — platform admin and admin only."""
    user = await get_platform_or_admin(request)
    days = {"7d": 7, "30d": 30, "90d": 90, "365d": 365}.get(period, 30)
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    # Aggregate across all tenants for platform admin
    query = {"created_at": {"$gte": start_date}}

    # Top customers by revenue
    invoices = await db.invoices.find(query, {"_id": 0, "customer_id": 1, "customer_name": 1, "grand_total": 1}).to_list(10000)
    customer_totals = {}
    for inv in invoices:
        cid = inv.get("customer_id", "walk-in")
        name = inv.get("customer_name", "Walk-in Customer")
        if cid not in customer_totals:
            customer_totals[cid] = {"name": name, "orders": 0, "total": 0}
        customer_totals[cid]["orders"] += 1
        customer_totals[cid]["total"] += inv.get("grand_total", 0)

    top_customers = sorted(customer_totals.values(), key=lambda x: x["total"], reverse=True)[:10]
    for c in top_customers:
        c["total"] = round(c["total"], 2)
        c["avg_order"] = round(c["total"] / max(c["orders"], 1), 2)

    # Customer purchase frequency distribution
    freq_dist = {"1_order": 0, "2_5_orders": 0, "6_10_orders": 0, "10_plus": 0}
    for c in customer_totals.values():
        if c["orders"] == 1:
            freq_dist["1_order"] += 1
        elif c["orders"] <= 5:
            freq_dist["2_5_orders"] += 1
        elif c["orders"] <= 10:
            freq_dist["6_10_orders"] += 1
        else:
            freq_dist["10_plus"] += 1

    # Revenue by category
    all_invoices = await db.invoices.find(query, {"_id": 0, "items": 1}).to_list(10000)
    cat_revenue = {}
    for inv in all_invoices:
        for item in inv.get("items", []):
            cat = item.get("category", "Uncategorized")
            cat_revenue[cat] = cat_revenue.get(cat, 0) + item.get("quantity", 0) * item.get("price", 0)

    categories = [{"name": k, "revenue": round(v, 2)} for k, v in sorted(cat_revenue.items(), key=lambda x: x[1], reverse=True)[:8]]

    return {
        "top_customers": top_customers,
        "frequency_distribution": freq_dist,
        "total_unique_customers": len(customer_totals),
        "revenue_by_category": categories
    }


@api_router.get("/analytics/owner/usage-heatmap")
async def analytics_usage_heatmap(request: Request, period: str = "30d"):
    """API usage heatmap — platform admin and admin only."""
    user = await get_platform_or_admin(request)
    days = {"7d": 7, "30d": 30, "90d": 90}.get(period, 30)
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")

    # Aggregate across all tenants for platform admin
    pipeline = [
        {"$match": {"date": {"$gte": start_date}}},
        {"$group": {"_id": {"hour": "$hour", "feature": "$feature"}, "count": {"$sum": 1}, "avg_ms": {"$avg": "$duration_ms"}}},
        {"$sort": {"count": -1}}
    ]
    results = await db.api_analytics.aggregate(pipeline).to_list(200)

    # Hourly distribution
    hourly = [0] * 24
    for r in results:
        h = r["_id"].get("hour", 0)
        if 0 <= h < 24:
            hourly[h] += r["count"]

    # Feature breakdown
    feature_counts = {}
    feature_perf = {}
    for r in results:
        f = r["_id"].get("feature", "other")
        feature_counts[f] = feature_counts.get(f, 0) + r["count"]
        if f not in feature_perf:
            feature_perf[f] = []
        feature_perf[f].append(r.get("avg_ms", 0))

    features = [{"name": k, "calls": v, "avg_ms": round(sum(feature_perf.get(k, [0])) / max(len(feature_perf.get(k, [1])), 1), 1)} for k, v in sorted(feature_counts.items(), key=lambda x: x[1], reverse=True)[:12]]

    # Error rate — across all tenants for platform admin
    total_calls = await db.api_analytics.count_documents({"date": {"$gte": start_date}})
    error_calls = await db.api_analytics.count_documents({"date": {"$gte": start_date}, "status_code": {"$gte": 400}})
    error_rate = round((error_calls / max(total_calls, 1)) * 100, 2)

    return {
        "hourly_distribution": [{"hour": i, "calls": hourly[i]} for i in range(24)],
        "feature_breakdown": features,
        "total_calls": total_calls,
        "error_calls": error_calls,
        "error_rate": error_rate
    }


@api_router.get("/analytics/realtime")
async def analytics_realtime(request: Request):
    """Real-time metrics — platform admin and admin only."""
    user = await get_platform_or_admin(request)
    cutoff = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()

    # Aggregate across all tenants for platform admin
    recent = await db.api_analytics.find({"timestamp": {"$gte": cutoff}}, {"_id": 0, "endpoint": 1, "status_code": 1, "duration_ms": 1, "timestamp": 1, "method": 1, "feature": 1}).sort("timestamp", -1).to_list(100)

    total = len(recent)
    errors = len([r for r in recent if r.get("status_code", 200) >= 400])
    avg_ms = round(sum(r.get("duration_ms", 0) for r in recent) / max(total, 1), 1)
    rpm = round(total / 5, 1)  # Requests per minute

    # Active endpoints
    endpoints = {}
    for r in recent:
        ep = r.get("endpoint", "")
        endpoints[ep] = endpoints.get(ep, 0) + 1

    top_endpoints = sorted(endpoints.items(), key=lambda x: x[1], reverse=True)[:8]

    return {
        "window": "5min",
        "total_requests": total,
        "requests_per_minute": rpm,
        "error_count": errors,
        "error_rate": round((errors / max(total, 1)) * 100, 1),
        "avg_response_ms": avg_ms,
        "top_endpoints": [{"endpoint": ep, "count": c} for ep, c in top_endpoints],
        "recent_requests": recent[:20],
    }


@api_router.get("/analytics/platform/overview")
async def analytics_platform_overview(request: Request, period: str = "30d"):
    """Platform-wide analytics for platform admin and admin."""
    admin = await get_platform_or_admin(request)
    days = {"7d": 7, "30d": 30, "90d": 90, "365d": 365}.get(period, 30)
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    start_date_str = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")

    total_tenants = await db.tenants.count_documents({})
    active_tenants = len(await db.api_analytics.distinct("tenant_id", {"date": {"$gte": start_date_str}}))
    total_users = await db.users.count_documents({})
    total_products = await db.products.count_documents({})
    total_invoices = await db.invoices.count_documents({"created_at": {"$gte": start_date}})

    # System-wide revenue
    inv_cursor = db.invoices.find({"created_at": {"$gte": start_date}}, {"_id": 0, "grand_total": 1, "tenant_id": 1})
    all_inv = await inv_cursor.to_list(50000)
    total_revenue = sum(inv.get("grand_total", 0) for inv in all_inv)

    # Revenue per tenant
    tenant_revenue = {}
    for inv in all_inv:
        t = inv.get("tenant_id", "unknown")
        tenant_revenue[t] = tenant_revenue.get(t, 0) + inv.get("grand_total", 0)

    # API calls system-wide
    total_api_calls = await db.api_analytics.count_documents({"date": {"$gte": start_date_str}})
    error_calls = await db.api_analytics.count_documents({"date": {"$gte": start_date_str}, "status_code": {"$gte": 400}})

    # Top tenants by usage
    tenant_usage_pipeline = [
        {"$match": {"date": {"$gte": start_date_str}}},
        {"$group": {"_id": "$tenant_id", "calls": {"$sum": 1}, "avg_ms": {"$avg": "$duration_ms"}}},
        {"$sort": {"calls": -1}},
        {"$limit": 10}
    ]
    tenant_usage = await db.api_analytics.aggregate(tenant_usage_pipeline).to_list(10)

    # Enrich with tenant names
    for tu in tenant_usage:
        if tu["_id"]:
            tenant = await db.tenants.find_one({"id": tu["_id"]}, {"_id": 0, "name": 1})
            tu["tenant_name"] = tenant["name"] if tenant else "Unknown"
            tu["revenue"] = round(tenant_revenue.get(tu["_id"], 0), 2)
        tu["avg_ms"] = round(tu.get("avg_ms", 0), 1)

    # Daily API trends
    daily_trend = []
    for i in range(min(days, 30) - 1, -1, -1):
        day = (datetime.now(timezone.utc) - timedelta(days=i)).strftime("%Y-%m-%d")
        count = await db.api_analytics.count_documents({"date": day})
        daily_trend.append({"date": day, "calls": count})

    return {
        "period": period,
        "kpis": {
            "total_tenants": total_tenants,
            "active_tenants": active_tenants,
            "total_users": total_users,
            "total_products": total_products,
            "total_invoices": total_invoices,
            "total_revenue": round(total_revenue, 2),
            "total_api_calls": total_api_calls,
            "error_calls": error_calls,
            "error_rate": round((error_calls / max(total_api_calls, 1)) * 100, 2),
        },
        "top_tenants": tenant_usage,
        "daily_api_trend": daily_trend,
    }


@api_router.get("/analytics/export")
async def analytics_export(request: Request, type: str = "revenue", period: str = "30d", format: str = "csv"):
    """Export analytics data as CSV — platform admin and admin only."""
    user = await get_platform_or_admin(request)
    days = {"7d": 7, "30d": 30, "90d": 90, "365d": 365}.get(period, 30)
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    if type == "revenue":
        # Export across all tenants for platform admin
        invoices = await db.invoices.find({"created_at": {"$gte": start_date}}, {"_id": 0, "invoice_number": 1, "created_at": 1, "grand_total": 1, "subtotal": 1, "discount": 1, "customer_name": 1, "tenant_id": 1}).sort("created_at", -1).to_list(10000)
        import io, csv
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=["invoice_number", "created_at", "customer_name", "subtotal", "discount", "grand_total", "tenant_id"])
        writer.writeheader()
        for inv in invoices:
            writer.writerow({k: inv.get(k, "") for k in writer.fieldnames})
        from starlette.responses import StreamingResponse
        return StreamingResponse(io.BytesIO(output.getvalue().encode()), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=revenue_{period}.csv"})

    elif type == "api_usage":
        start_date_str = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
        # Export across all tenants for platform admin
        records = await db.api_analytics.find({"date": {"$gte": start_date_str}}, {"_id": 0, "timestamp": 1, "endpoint": 1, "method": 1, "status_code": 1, "duration_ms": 1, "user_id": 1, "tenant_id": 1}).sort("timestamp", -1).to_list(10000)
        import io, csv
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=["timestamp", "endpoint", "method", "status_code", "duration_ms", "user_id", "tenant_id"])
        writer.writeheader()
        for r in records:
            writer.writerow({k: r.get(k, "") for k in writer.fieldnames})
        from starlette.responses import StreamingResponse
        return StreamingResponse(io.BytesIO(output.getvalue().encode()), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=api_usage_{period}.csv"})

    raise HTTPException(400, "Invalid export type")


# ═══════════════════════════════════════════════════════════════════
#  PLATFORM ADMIN ROUTES
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/admin/cache-stats")
async def get_cache_stats(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "OWNER":
        raise HTTPException(403, "Only owners can view cache stats")
    return {
        "product_cache": product_cache.stats,
        "category_cache": category_cache.stats,
        "barcode_cache": barcode_cache.stats,
        "dashboard_cache": dashboard_cache.stats,
        "customer_cache": customer_cache.stats,
    }


@api_router.get("/platform/tenants")
async def list_tenants(request: Request, page: int = 1, limit: int = 20):
    """Platform admin or admin lists all tenants (no financial data)"""
    admin = await get_platform_or_admin(request)
    total = await db.tenants.count_documents({})
    skip = (page - 1) * limit
    tenants = await db.tenants.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)

    # Enrich with user count (no financial data)
    for t in tenants:
        t["user_count"] = await db.users.count_documents({"tenant_id": t["id"]})
        t["product_count"] = await db.products.count_documents({"tenant_id": t["id"]})

    return {"tenants": tenants, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.get("/platform/stats")
async def platform_stats(request: Request):
    """Platform overview stats (no per-tenant financial data)"""
    admin = await get_platform_or_admin(request)
    total_tenants = await db.tenants.count_documents({})
    total_users = await db.users.count_documents({})
    total_products = await db.products.count_documents({})
    active_tenants = await db.tenants.count_documents({"is_active": True})
    open_tickets = await db.support_tickets.count_documents({"status": {"$in": ["open", "in_progress"]}})
    pending_access = await db.access_requests.count_documents({"status": "pending"})

    return {
        "total_tenants": total_tenants,
        "active_tenants": active_tenants,
        "total_users": total_users,
        "total_products": total_products,
        "open_support_tickets": open_tickets,
        "pending_access_requests": pending_access
    }

# ═══════════════════════════════════════════════════════════════════
#  ENHANCED AUDIT LOGGING
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/audit-logs/export")
async def export_audit_logs(request: Request, format: str = "csv", start_date: str = "", end_date: str = ""):
    """Export audit logs as CSV/Excel - immutable, no delete endpoint"""
    user = await get_current_user(request)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")

    query = {"tenant_id": user["tenant_id"]}
    if start_date:
        query["timestamp"] = {"$gte": start_date}
    if end_date:
        query.setdefault("timestamp", {})["$lte"] = end_date

    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).to_list(50000)
    headers = ["Timestamp", "Action", "User ID", "Details", "IP Address", "Event Category"]
    keys = ["timestamp", "action", "user_id", "details", "ip_address", "event_category"]

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        ws.append(headers)
        for log_entry in logs:
            ws.append([log_entry.get(k, "") for k in keys])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               headers={"Content-Disposition": "attachment; filename=audit_logs.xlsx"})
    else:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        for log_entry in logs:
            writer.writerow([log_entry.get(k, "") for k in keys])
        buf.seek(0)
        return StreamingResponse(io.BytesIO(buf.getvalue().encode()), media_type="text/csv",
                               headers={"Content-Disposition": "attachment; filename=audit_logs.csv"})

# ═══════════════════════════════════════════════════════════════════
#  SMART IP WHITELISTING / TEMP ACCESS
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/admin/temp-access")
async def grant_temp_access(req: TempAccessCreate, request: Request):
    """Admin/Owner grants temporary IP access to a user"""
    admin = await get_current_user(request)
    if admin["role"] not in ["OWNER", "MANAGER"] and not admin.get("is_platform_admin") and not admin.get("is_admin"):
        raise HTTPException(403, "Insufficient permissions")

    # Verify target user exists in same tenant (or platform admin / admin)
    if admin.get("is_platform_admin") or admin.get("is_admin"):
        target = await db.users.find_one({"_id": ObjectId(req.user_id)})
    else:
        target = await db.users.find_one({"_id": ObjectId(req.user_id), "tenant_id": admin["tenant_id"]})
    if not target:
        raise HTTPException(404, "User not found")

    expires_at = (datetime.now(timezone.utc) + timedelta(hours=req.duration_hours)).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": req.user_id,
        "tenant_id": target["tenant_id"],
        "granted_by": admin["id"],
        "granted_by_name": admin.get("name", ""),
        "reason": req.reason,
        "allowed_ip": req.allowed_ip,
        "expires_at": expires_at,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.temp_access.insert_one(doc)
    await log_audit(target["tenant_id"], admin["id"], "temp_access_granted",
                    f"Temp IP access for user {req.user_id}, IP: {req.allowed_ip}, Duration: {req.duration_hours}h",
                    request.client.host if request.client else "", "security")
    doc.pop("_id", None)
    return doc

@api_router.get("/admin/temp-access")
async def list_temp_access(request: Request, user_id: str = ""):
    """List active temporary access grants"""
    admin = await get_current_user(request)
    if admin["role"] not in ["OWNER", "MANAGER"] and not admin.get("is_platform_admin") and not admin.get("is_admin"):
        raise HTTPException(403, "Insufficient permissions")

    now = datetime.now(timezone.utc).isoformat()
    if admin.get("is_platform_admin") or admin.get("is_admin"):
        query = {"is_active": True, "expires_at": {"$gt": now}}
    else:
        query = {"tenant_id": admin["tenant_id"], "is_active": True, "expires_at": {"$gt": now}}

    if user_id:
        query["user_id"] = user_id

    grants = await db.temp_access.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)

    # Enrich with user names
    for g in grants:
        try:
            u = await db.users.find_one({"_id": ObjectId(g["user_id"])}, {"name": 1, "email": 1})
            g["user_name"] = u.get("name", "") if u else "Unknown"
            g["user_email"] = u.get("email", "") if u else ""
        except Exception:
            g["user_name"] = "Unknown"
            g["user_email"] = ""

    return {"grants": grants, "total": len(grants)}

@api_router.delete("/admin/temp-access/{grant_id}")
async def revoke_temp_access(grant_id: str, request: Request):
    """Revoke a temporary access grant"""
    admin = await get_current_user(request)
    if admin["role"] not in ["OWNER", "MANAGER"] and not admin.get("is_platform_admin") and not admin.get("is_admin"):
        raise HTTPException(403, "Insufficient permissions")

    if admin.get("is_platform_admin") or admin.get("is_admin"):
        grant = await db.temp_access.find_one({"id": grant_id})
    else:
        grant = await db.temp_access.find_one({"id": grant_id, "tenant_id": admin["tenant_id"]})

    if not grant:
        raise HTTPException(404, "Temp access grant not found")

    await db.temp_access.update_one({"id": grant_id}, {"$set": {"is_active": False}})
    tid = grant.get("tenant_id", admin.get("tenant_id", ""))
    await log_audit(tid, admin["id"], "temp_access_revoked",
                    f"Revoked temp access {grant_id} for user {grant.get('user_id')}",
                    request.client.host if request.client else "", "security")
    return {"message": "Temporary access revoked"}

# ═══════════════════════════════════════════════════════════════════
#  FRAUD DETECTION / SECURITY ALERTS
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/security/alerts")
async def list_security_alerts(request: Request, is_read: str = "", severity: str = "", page: int = 1, limit: int = 50):
    """Get security alerts for the tenant"""
    user = await get_current_user(request)
    if user["role"] not in ["OWNER", "MANAGER"] and not user.get("is_platform_admin"):
        raise HTTPException(403, "Insufficient permissions")

    if user.get("is_platform_admin"):
        query = {}
    else:
        query = {"tenant_id": user["tenant_id"]}

    if is_read == "true":
        query["is_read"] = True
    elif is_read == "false":
        query["is_read"] = False

    if severity:
        query["severity"] = severity

    total = await db.security_alerts.count_documents(query)
    skip = (page - 1) * limit
    alerts = await db.security_alerts.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)

    # Enrich with user info
    for a in alerts:
        try:
            u = await db.users.find_one({"_id": ObjectId(a.get("user_id", ""))}, {"name": 1, "email": 1})
            a["user_name"] = u.get("name", "") if u else "Unknown"
            a["user_email"] = u.get("email", "") if u else ""
        except Exception:
            a["user_name"] = "Unknown"
            a["user_email"] = ""

    return {"alerts": alerts, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.put("/security/alerts/{alert_id}/read")
async def mark_alert_read(alert_id: str, request: Request):
    """Mark a security alert as read"""
    user = await get_current_user(request)
    if user["role"] not in ["OWNER", "MANAGER"] and not user.get("is_platform_admin"):
        raise HTTPException(403, "Insufficient permissions")

    if user.get("is_platform_admin"):
        result = await db.security_alerts.update_one({"id": alert_id}, {"$set": {"is_read": True}})
    else:
        result = await db.security_alerts.update_one(
            {"id": alert_id, "tenant_id": user["tenant_id"]},
            {"$set": {"is_read": True}}
        )

    if result.modified_count == 0:
        raise HTTPException(404, "Alert not found")
    return {"message": "Alert marked as read"}

@api_router.put("/security/alerts/read-all")
async def mark_all_alerts_read(request: Request):
    """Mark all security alerts as read"""
    user = await get_current_user(request)
    if user["role"] not in ["OWNER", "MANAGER"] and not user.get("is_platform_admin"):
        raise HTTPException(403, "Insufficient permissions")

    if user.get("is_platform_admin"):
        await db.security_alerts.update_many({"is_read": False}, {"$set": {"is_read": True}})
    else:
        await db.security_alerts.update_many(
            {"tenant_id": user["tenant_id"], "is_read": False},
            {"$set": {"is_read": True}}
        )
    return {"message": "All alerts marked as read"}

@api_router.get("/security/alerts/summary")
async def security_alerts_summary(request: Request):
    """Get summary counts of security alerts"""
    user = await get_current_user(request)
    if user["role"] not in ["OWNER", "MANAGER"] and not user.get("is_platform_admin"):
        raise HTTPException(403, "Insufficient permissions")

    if user.get("is_platform_admin"):
        base_query = {}
    else:
        base_query = {"tenant_id": user["tenant_id"]}

    total = await db.security_alerts.count_documents(base_query)
    unread = await db.security_alerts.count_documents({**base_query, "is_read": False})
    high_severity = await db.security_alerts.count_documents({**base_query, "severity": "high", "is_read": False})
    medium_severity = await db.security_alerts.count_documents({**base_query, "severity": "medium", "is_read": False})

    # Counts by type
    type_counts = {}
    for alert_type in ["failed_logins", "new_ip_login", "ip_blocked", "suspicious_activity"]:
        type_counts[alert_type] = await db.security_alerts.count_documents({**base_query, "alert_type": alert_type, "is_read": False})

    return {
        "total": total,
        "unread": unread,
        "high_severity_unread": high_severity,
        "medium_severity_unread": medium_severity,
        "by_type": type_counts
    }

# ═══════════════════════════════════════════════════════════════════
#  PREMIUM: PROMO CODE SYSTEM
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/promo-codes")
async def create_promo_code(req: PromoCodeCreate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Only OWNER/MANAGER can manage promo codes")

    existing = await db.promo_codes.find_one({"tenant_id": user["tenant_id"], "code": req.code.upper()})
    if existing:
        raise HTTPException(400, "Promo code already exists")

    doc = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        "code": req.code.upper(), "discount_type": req.discount_type,
        "value": req.value, "min_order_amount": req.min_order_amount,
        "max_discount": req.max_discount, "valid_from": req.valid_from or datetime.now(timezone.utc).isoformat(),
        "valid_to": req.valid_to, "max_uses": req.max_uses, "current_uses": 0,
        "is_active": True, "description": req.description,
        "created_by": user["id"], "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.promo_codes.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "promo_code_created", f"Code: {req.code.upper()}", request.client.host if request.client else "", "general")
    doc.pop("_id", None)
    return doc

@api_router.get("/promo-codes")
async def list_promo_codes(request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    codes = await db.promo_codes.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"promo_codes": codes}

@api_router.put("/promo-codes/{code_id}")
async def update_promo_code(code_id: str, req: PromoCodeUpdate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")
    updates = {k: v for k, v in req.dict().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No updates provided")
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.promo_codes.update_one({"id": code_id, "tenant_id": user["tenant_id"]}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(404, "Promo code not found")
    return {"message": "Promo code updated"}

@api_router.delete("/promo-codes/{code_id}")
async def delete_promo_code(code_id: str, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")
    await db.promo_codes.update_one({"id": code_id, "tenant_id": user["tenant_id"]}, {"$set": {"is_active": False}})
    return {"message": "Promo code deactivated"}

@api_router.post("/promo-codes/validate")
async def validate_promo_code(req: PromoCodeValidate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    code = await db.promo_codes.find_one({"tenant_id": user["tenant_id"], "code": req.code.upper(), "is_active": True}, {"_id": 0})
    if not code:
        raise HTTPException(404, "Invalid or inactive promo code")
    now = datetime.now(timezone.utc).isoformat()
    if code.get("valid_to") and code["valid_to"] < now:
        raise HTTPException(400, "Promo code has expired")
    if code.get("valid_from") and code["valid_from"] > now:
        raise HTTPException(400, "Promo code is not yet valid")
    if code.get("max_uses") and code.get("max_uses") > 0 and code.get("current_uses", 0) >= code["max_uses"]:
        raise HTTPException(400, "Promo code usage limit reached")
    if req.order_amount < code.get("min_order_amount", 0):
        raise HTTPException(400, f"Minimum order amount is ₹{code['min_order_amount']}")

    if code["discount_type"] == "percentage":
        discount = round(req.order_amount * code["value"] / 100, 2)
        if code.get("max_discount") and code["max_discount"] > 0:
            discount = min(discount, code["max_discount"])
    else:
        discount = code["value"]

    return {"valid": True, "code": code["code"], "discount_type": code["discount_type"],
            "discount_value": code["value"], "discount_amount": discount,
            "final_amount": round(req.order_amount - discount, 2)}

# ═══════════════════════════════════════════════════════════════════
#  PREMIUM: AUTO REORDER SYSTEM
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/reorder/rules")
async def create_reorder_rule(req: ReorderRuleCreate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")
    product = await db.products.find_one({"id": req.product_id, "tenant_id": user["tenant_id"]})
    if not product:
        raise HTTPException(404, "Product not found")
    existing = await db.reorder_rules.find_one({"tenant_id": user["tenant_id"], "product_id": req.product_id})
    if existing:
        raise HTTPException(400, "Reorder rule already exists for this product. Update instead.")
    doc = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        **req.dict(), "product_name": product.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.reorder_rules.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/reorder/rules")
async def list_reorder_rules(request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    rules = await db.reorder_rules.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(500)
    for r in rules:
        p = await db.products.find_one({"id": r.get("product_id"), "tenant_id": user["tenant_id"]}, {"name": 1, "stock": 1, "sku": 1})
        if p:
            r["product_name"] = p.get("name", "")
            r["current_stock"] = p.get("stock", 0)
            r["sku"] = p.get("sku", "")
    return {"rules": rules}

@api_router.put("/reorder/rules/{rule_id}")
async def update_reorder_rule(rule_id: str, req: ReorderRuleUpdate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")
    updates = {k: v for k, v in req.dict().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No updates provided")
    result = await db.reorder_rules.update_one({"id": rule_id, "tenant_id": user["tenant_id"]}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(404, "Reorder rule not found")
    return {"message": "Reorder rule updated"}

@api_router.delete("/reorder/rules/{rule_id}")
async def delete_reorder_rule(rule_id: str, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    result = await db.reorder_rules.delete_one({"id": rule_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Reorder rule not found")
    return {"message": "Reorder rule deleted"}

@api_router.post("/reorder/check")
async def manual_reorder_check(request: Request):
    """Manually check all products for reorder needs"""
    user = await get_current_user(request)
    await require_premium(user)
    rules = await db.reorder_rules.find({"tenant_id": user["tenant_id"], "is_active": True}).to_list(1000)
    triggered = 0
    for rule in rules:
        product = await db.products.find_one({"id": rule["product_id"], "tenant_id": user["tenant_id"]}, {"_id": 0})
        if product and product.get("stock", 0) <= rule.get("threshold", 0):
            await trigger_reorder_notifications(user["tenant_id"], product, rule, user)
            triggered += 1
    return {"message": f"Reorder check complete. {triggered} product(s) below threshold."}

# ─── Notification Templates ──────────────────────────────────────

@api_router.post("/notification-templates")
async def create_notification_template(req: NotificationTemplateCreate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")
    if req.channel not in ["whatsapp", "email", "voice"]:
        raise HTTPException(400, "Channel must be whatsapp, email, or voice")
    doc = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        "channel": req.channel, "name": req.name, "subject": req.subject,
        "template_text": req.template_text, "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notification_templates.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/notification-templates")
async def list_notification_templates(request: Request, channel: str = ""):
    user = await get_current_user(request)
    await require_premium(user)
    query = {"tenant_id": user["tenant_id"]}
    if channel:
        query["channel"] = channel
    templates = await db.notification_templates.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"templates": templates}

@api_router.put("/notification-templates/{template_id}")
async def update_notification_template(template_id: str, req: NotificationTemplateUpdate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    updates = {k: v for k, v in req.dict().items() if v is not None}
    result = await db.notification_templates.update_one({"id": template_id, "tenant_id": user["tenant_id"]}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(404, "Template not found")
    return {"message": "Template updated"}

@api_router.get("/notification-logs")
async def list_notification_logs(request: Request, channel: str = "", page: int = 1, limit: int = 50):
    user = await get_current_user(request)
    await require_premium(user)
    query = {"tenant_id": user["tenant_id"]}
    if channel:
        query["channel"] = channel
    total = await db.notification_logs.count_documents(query)
    skip = (page - 1) * limit
    logs = await db.notification_logs.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "page": page, "pages": (total + limit - 1) // limit}

# ─── SMTP Settings ───────────────────────────────────────────────

@api_router.put("/settings/smtp")
async def update_smtp_settings(req: SMTPSettingsUpdate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can configure SMTP")
    await db.tenants.update_one({"id": user["tenant_id"]}, {"$set": {"smtp_settings": req.dict()}})
    await log_audit(user["tenant_id"], user["id"], "smtp_settings_updated", "SMTP email settings configured", request.client.host if request.client else "", "settings")
    return {"message": "SMTP settings saved"}

@api_router.get("/settings/smtp")
async def get_smtp_settings(request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER")
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"smtp_settings": 1})
    smtp = tenant.get("smtp_settings", {}) if tenant else {}
    smtp.pop("smtp_password", None)  # Don't expose password
    return smtp

# ═══════════════════════════════════════════════════════════════════
#  PREMIUM: ADVANCE PAYMENT HANDLING
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/advance-orders")
async def create_advance_order(req: AdvanceOrderCreate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if req.advance_amount <= 0:
        raise HTTPException(400, "Advance amount must be positive")
    doc = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        "customer_name": req.customer_name, "customer_phone": req.customer_phone,
        "products": req.products, "advance_amount": req.advance_amount,
        "total_estimated": req.total_estimated, "balance_due": round(req.total_estimated - req.advance_amount, 2),
        "notes": req.notes, "status": "pending",
        "created_by": user["id"], "created_by_name": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "fulfilled_at": None, "invoice_id": None
    }
    await db.advance_orders.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "advance_order_created", f"Customer: {req.customer_name}, Advance: ₹{req.advance_amount}", request.client.host if request.client else "", "billing")
    doc.pop("_id", None)
    return doc

@api_router.get("/advance-orders")
async def list_advance_orders(request: Request, status: str = "", page: int = 1, limit: int = 50):
    user = await get_current_user(request)
    await require_premium(user)
    query = {"tenant_id": user["tenant_id"]}
    if status:
        query["status"] = status
    total = await db.advance_orders.count_documents(query)
    skip = (page - 1) * limit
    orders = await db.advance_orders.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"orders": orders, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.get("/advance-orders/{order_id}")
async def get_advance_order(order_id: str, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    order = await db.advance_orders.find_one({"id": order_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Advance order not found")
    return order

@api_router.put("/advance-orders/{order_id}/fulfill")
async def fulfill_advance_order(order_id: str, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    order = await db.advance_orders.find_one({"id": order_id, "tenant_id": user["tenant_id"]})
    if not order:
        raise HTTPException(404, "Advance order not found")
    if order["status"] != "pending":
        raise HTTPException(400, f"Order is already {order['status']}")
    await db.advance_orders.update_one(
        {"id": order_id},
        {"$set": {"status": "fulfilled", "fulfilled_at": datetime.now(timezone.utc).isoformat()}}
    )
    await log_audit(user["tenant_id"], user["id"], "advance_order_fulfilled", f"Order: {order_id}", request.client.host if request.client else "", "billing")
    return {"message": "Order fulfilled"}

@api_router.put("/advance-orders/{order_id}/cancel")
async def cancel_advance_order(order_id: str, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    order = await db.advance_orders.find_one({"id": order_id, "tenant_id": user["tenant_id"]})
    if not order:
        raise HTTPException(404, "Advance order not found")
    if order["status"] != "pending":
        raise HTTPException(400, f"Order is already {order['status']}")
    await db.advance_orders.update_one({"id": order_id}, {"$set": {"status": "cancelled"}})
    return {"message": "Order cancelled"}

# ═══════════════════════════════════════════════════════════════════
#  PREMIUM: SMART PRODUCT RECOMMENDATIONS
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/inventory/recommendations/{product_id}")
async def get_product_recommendations(product_id: str, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    product = await db.products.find_one({"id": product_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not product:
        raise HTTPException(404, "Product not found")
    # Find alternatives in same category with stock > 0
    alternatives = await db.products.find({
        "tenant_id": user["tenant_id"], "category": product.get("category", ""),
        "id": {"$ne": product_id}, "stock": {"$gt": 0}
    }, {"_id": 0}).limit(10).to_list(10)
    return {"product": product, "alternatives": alternatives}

@api_router.get("/inventory/search-alternatives")
async def search_alternatives(request: Request, category: str = "", exclude_id: str = "", search: str = ""):
    user = await get_current_user(request)
    await require_premium(user)
    query = {"tenant_id": user["tenant_id"], "stock": {"$gt": 0}}
    if category:
        query["category"] = category
    if exclude_id:
        query["id"] = {"$ne": exclude_id}
    if search:
        query["$or"] = [{"name": {"$regex": search, "$options": "i"}}, {"sku": {"$regex": search, "$options": "i"}}]
    products = await db.products.find(query, {"_id": 0}).limit(15).to_list(15)
    return {"alternatives": products}

# ═══════════════════════════════════════════════════════════════════
#  PREMIUM: ADMIN-CONTROLLED OWNER ACCOUNT MANAGEMENT
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/platform/create-owner")
async def platform_create_owner(req: CreateOwnerAccount, request: Request):
    admin = await get_platform_or_admin(request)
    existing = await db.users.find_one({"email": req.email})
    if existing:
        raise HTTPException(400, "Email already registered")

    tenant_id = str(uuid.uuid4())
    valid_until = (datetime.now(timezone.utc) + timedelta(days=req.valid_days)).isoformat()
    await db.tenants.insert_one({
        "id": tenant_id, "shop_name": req.shop_name, "business_type": req.business_type,
        "plan": req.plan, "max_users": 10 if req.plan == "basic" else 50 if req.plan == "standard" else 999,
        "is_active": True, "is_revoked": False, "valid_until": valid_until,
        "created_by_admin": admin["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    await db.users.insert_one({
        "email": req.email, "password_hash": hash_password(req.password),
        "name": req.name, "role": "OWNER", "tenant_id": tenant_id,
        "is_active": True, "mfa_enabled": False, "mfa_secret": None,
        "is_platform_admin": False, "known_login_ips": [],
        "permissions": {"can_view_revenue": True, "can_manage_inventory": True},
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    await log_audit("system", admin["id"], "owner_account_created",
                    f"Created owner {req.email} for shop {req.shop_name}, plan: {req.plan}, valid: {req.valid_days} days",
                    request.client.host if request.client else "", "admin")
    return {"message": "Owner account created", "tenant_id": tenant_id, "email": req.email, "valid_until": valid_until}

@api_router.put("/platform/tenants/{tenant_id}/status")
async def update_tenant_status(tenant_id: str, request: Request):
    admin = await get_platform_or_admin(request)
    body = await request.json()
    action = body.get("action")  # activate, deactivate, revoke
    if action not in ["activate", "deactivate", "revoke"]:
        raise HTTPException(400, "Action must be activate, deactivate, or revoke")
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(404, "Tenant not found")
    if action == "activate":
        await db.tenants.update_one({"id": tenant_id}, {"$set": {"is_active": True, "is_revoked": False}})
        await db.users.update_many({"tenant_id": tenant_id}, {"$set": {"is_active": True}})
    elif action == "deactivate":
        await db.tenants.update_one({"id": tenant_id}, {"$set": {"is_active": False}})
        await db.users.update_many({"tenant_id": tenant_id}, {"$set": {"is_active": False}})
    elif action == "revoke":
        await db.tenants.update_one({"id": tenant_id}, {"$set": {"is_revoked": True}})
    await log_audit("system", admin["id"], f"tenant_{action}d", f"Tenant {tenant_id} ({tenant.get('shop_name', '')}) {action}d",
                    request.client.host if request.client else "", "admin")
    return {"message": f"Tenant {action}d successfully"}

@api_router.put("/platform/tenants/{tenant_id}/extend")
async def extend_tenant_validity(tenant_id: str, request: Request):
    admin = await get_platform_or_admin(request)
    body = await request.json()
    days = body.get("days", 30)
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(404, "Tenant not found")
    current_valid = tenant.get("valid_until", datetime.now(timezone.utc).isoformat())
    try:
        base_date = datetime.fromisoformat(current_valid.replace('Z', '+00:00'))
    except Exception:
        base_date = datetime.now(timezone.utc)
    if base_date < datetime.now(timezone.utc):
        base_date = datetime.now(timezone.utc)
    new_valid = (base_date + timedelta(days=days)).isoformat()
    await db.tenants.update_one({"id": tenant_id}, {"$set": {"valid_until": new_valid, "is_revoked": False, "is_active": True}})
    await db.users.update_many({"tenant_id": tenant_id}, {"$set": {"is_active": True}})
    await log_audit("system", admin["id"], "tenant_extended", f"Tenant {tenant_id} extended by {days} days until {new_valid}",
                    request.client.host if request.client else "", "admin")
    return {"message": f"Validity extended by {days} days", "valid_until": new_valid}

@api_router.put("/platform/tenants/{tenant_id}/plan")
async def update_tenant_plan(tenant_id: str, request: Request):
    admin = await get_platform_or_admin(request)
    body = await request.json()
    plan = body.get("plan")
    if plan not in ["basic", "standard", "premium"]:
        raise HTTPException(400, "Plan must be basic, standard, or premium")
    await db.tenants.update_one({"id": tenant_id}, {"$set": {"plan": plan}})
    return {"message": f"Plan updated to {plan}"}

# ═══════════════════════════════════════════════════════════════════
#  PLATFORM ADMIN: ADMIN ACCOUNT MANAGEMENT (Platform Admin only)
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/platform/admins")
async def list_admins(request: Request):
    """List all admin accounts — platform admin only."""
    admin = await get_platform_admin(request)
    admins = await db.users.find({"is_admin": True}).to_list(100)
    result = []
    for a in admins:
        doc = serialize_doc(a)
        doc.pop("password_hash", None)
        doc.pop("mfa_secret", None)
        result.append(doc)
    return {"admins": result}

@api_router.post("/platform/create-admin")
async def platform_create_admin(req: CreateAdminAccount, request: Request):
    """Create a new admin account — platform admin only."""
    admin = await get_platform_admin(request)
    existing = await db.users.find_one({"email": req.email.lower().strip()})
    if existing:
        raise HTTPException(400, "Email already registered")

    admin_user = {
        "email": req.email.lower().strip(),
        "password_hash": hash_password(req.password),
        "name": req.name,
        "role": "ADMIN",
        "tenant_id": "system",
        "is_active": True,
        "is_admin": True,
        "is_platform_admin": False,
        "mfa_enabled": False,
        "mfa_secret": None,
        "known_login_ips": [],
        "permissions": {"can_view_revenue": True, "can_manage_inventory": True},
        "created_by": admin["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.users.insert_one(admin_user)
    await log_audit("system", admin["id"], "admin_account_created",
                    f"Created admin {req.email}", request.client.host if request.client else "", "admin")
    return {"message": "Admin account created", "id": str(result.inserted_id), "email": req.email}

@api_router.put("/platform/admins/{admin_id}/status")
async def update_admin_status(admin_id: str, request: Request):
    """Activate/deactivate admin account — platform admin only."""
    admin = await get_platform_admin(request)
    body = await request.json()
    is_active = body.get("is_active", True)
    target = await db.users.find_one({"_id": ObjectId(admin_id), "is_admin": True})
    if not target:
        raise HTTPException(404, "Admin account not found")
    await db.users.update_one({"_id": ObjectId(admin_id)}, {"$set": {"is_active": is_active}})
    action = "activated" if is_active else "deactivated"
    await log_audit("system", admin["id"], f"admin_account_{action}",
                    f"Admin {target.get('email', '')} {action}", request.client.host if request.client else "", "admin")
    return {"message": f"Admin account {action}"}

@api_router.delete("/platform/admins/{admin_id}")
async def delete_admin(admin_id: str, request: Request):
    """Delete admin account — platform admin only."""
    admin = await get_platform_admin(request)
    target = await db.users.find_one({"_id": ObjectId(admin_id), "is_admin": True})
    if not target:
        raise HTTPException(404, "Admin account not found")
    await db.users.delete_one({"_id": ObjectId(admin_id)})
    await log_audit("system", admin["id"], "admin_account_deleted",
                    f"Deleted admin {target.get('email', '')}", request.client.host if request.client else "", "admin")
    return {"message": "Admin account deleted"}

# ═══════════════════════════════════════════════════════════════════
#  TENANT / SETTINGS
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/tenant")
async def get_tenant(request: Request):
    user = await get_current_user(request)
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    if not tenant:
        raise HTTPException(404, "Tenant not found")
    return tenant

@api_router.put("/tenant")
async def update_tenant(request: Request):
    user = await get_current_user(request)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can update settings")
    body = await request.json()
    allowed = ["shop_name", "business_type", "address", "phone", "gst_number"]
    updates = {k: v for k, v in body.items() if k in allowed}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.tenants.update_one({"id": user["tenant_id"]}, {"$set": updates})
    return {"message": "Settings updated"}

# ═══════════════════════════════════════════════════════════════════
#  FEATURE: SMART DIGITAL RECEIPTS WITH VIRAL BRANDING
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/invoices/{invoice_id}/digital-receipt")
async def get_digital_receipt(invoice_id: str, request: Request):
    """Get digital receipt data with branding and QR code."""
    user = await get_current_user(request)
    tid = user["tenant_id"]
    invoice = await db.invoices.find_one({"id": invoice_id, "tenant_id": tid}, {"_id": 0})
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    tenant = await db.tenants.find_one({"id": tid}, {"_id": 0})
    shop_name = tenant.get("shop_name", "RetailPro Store") if tenant else "RetailPro Store"

    # Generate or retrieve share token
    share_token = invoice.get("share_token")
    if not share_token:
        share_token = str(uuid.uuid4())[:12]
        await db.invoices.update_one({"id": invoice_id}, {"$set": {"share_token": share_token}})

    frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
    share_url = f"{frontend_url}/receipt/{share_token}"

    # Calculate loyalty points (1 point per ₹100 spent)
    loyalty_points = int(invoice.get("grand_total", 0) / 100)

    return {
        "invoice": invoice,
        "shop_name": shop_name,
        "share_token": share_token,
        "share_url": share_url,
        "loyalty_points": loyalty_points,
        "branding": {
            "footer": "Billing powered by RetailPro — Smarter retail starts here",
            "logo_text": "RetailPro",
            "tagline": "Smart Billing for Smart Retailers"
        },
        "whatsapp_share_url": f"https://wa.me/?text={share_url}%0A%0AHere%27s%20your%20digital%20receipt%20from%20{shop_name.replace(' ', '%20')}%0APowered%20by%20RetailPro"
    }

@api_router.get("/receipt/{share_token}")
async def get_public_receipt(share_token: str):
    """Public endpoint — view shared receipt (no auth required)."""
    invoice = await db.invoices.find_one({"share_token": share_token}, {"_id": 0})
    if not invoice:
        raise HTTPException(404, "Receipt not found")

    tenant = await db.tenants.find_one({"id": invoice["tenant_id"]}, {"_id": 0})
    shop_name = tenant.get("shop_name", "RetailPro Store") if tenant else "RetailPro Store"

    loyalty_points = int(invoice.get("grand_total", 0) / 100)

    return {
        "invoice_number": invoice.get("invoice_number"),
        "date": invoice.get("created_at"),
        "shop_name": shop_name,
        "customer_name": invoice.get("customer_name", "Walk-in Customer"),
        "items": invoice.get("items", []),
        "subtotal": invoice.get("subtotal", 0),
        "discount": invoice.get("discount", 0),
        "tax": invoice.get("tax", 0),
        "grand_total": invoice.get("grand_total", 0),
        "payment_method": invoice.get("payment_method", "cash"),
        "loyalty_points": loyalty_points,
        "branding": {
            "footer": "Billing powered by RetailPro — Smarter retail starts here",
            "logo_text": "RetailPro",
            "tagline": "Smart Billing for Smart Retailers"
        }
    }

@api_router.post("/invoices/{invoice_id}/send-receipt")
async def send_digital_receipt(invoice_id: str, request: Request):
    """Send digital receipt via WhatsApp/SMS (generates shareable link)."""
    user = await get_current_user(request)
    tid = user["tenant_id"]
    body = await request.json()
    phone = body.get("phone", "")
    channel = body.get("channel", "whatsapp")  # whatsapp | sms

    invoice = await db.invoices.find_one({"id": invoice_id, "tenant_id": tid}, {"_id": 0})
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    share_token = invoice.get("share_token")
    if not share_token:
        share_token = str(uuid.uuid4())[:12]
        await db.invoices.update_one({"id": invoice_id}, {"$set": {"share_token": share_token}})

    tenant = await db.tenants.find_one({"id": tid}, {"_id": 0})
    shop_name = tenant.get("shop_name", "RetailPro Store") if tenant else "RetailPro Store"
    frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
    share_url = f"{frontend_url}/receipt/{share_token}"

    message = f"🧾 Digital Receipt from {shop_name}\n\nInvoice: {invoice.get('invoice_number', '')}\nTotal: ₹{invoice.get('grand_total', 0):,.2f}\n\nView receipt: {share_url}\n\n✨ Billing powered by RetailPro"

    # Log the receipt send attempt
    await db.receipt_sends.insert_one({
        "id": str(uuid.uuid4()),
        "invoice_id": invoice_id,
        "tenant_id": tid,
        "phone": phone,
        "channel": channel,
        "message": message,
        "status": "generated",  # In production: "sent" via Twilio/Meta API
        "created_at": datetime.now(timezone.utc).isoformat()
    })

    wa_link = f"https://wa.me/{phone.replace('+', '')}?text={message.replace(chr(10), '%0A').replace(' ', '%20')}" if phone else ""

    await log_audit(tid, user["id"], "receipt_sent", f"Digital receipt sent for invoice {invoice_id} via {channel}", request.client.host if request.client else "")

    return {
        "message": "Receipt link generated",
        "share_url": share_url,
        "whatsapp_link": wa_link,
        "sms_message": message,
        "status": "generated"
    }

# ═══════════════════════════════════════════════════════════════════
#  FEATURE: SMART SUBSTITUTION ENGINE
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/products/{product_id}/substitutes")
async def get_product_substitutes(product_id: str, request: Request):
    """Find substitute products (same category, in stock)."""
    user = await get_current_user(request)
    tid = user["tenant_id"]

    product = await db.products.find_one({"id": product_id, "tenant_id": tid}, {"_id": 0})
    if not product:
        raise HTTPException(404, "Product not found")

    category = product.get("category", "")
    product_name = product.get("name", "")
    price = product.get("price", 0)

    # Find alternatives: same category, in stock, different product
    query = {
        "tenant_id": tid,
        "id": {"$ne": product_id},
        "stock": {"$gt": 0}
    }
    if category:
        query["category"] = category

    candidates = await db.products.find(query, {"_id": 0}).to_list(20)

    # Sort by price similarity
    for c in candidates:
        c["price_diff"] = round(c.get("price", 0) - price, 2)
        c["margin_diff"] = round(c.get("price", 0) - c.get("cost_price", 0), 2) - round(price - product.get("cost_price", 0), 2)
        c["margin_diff"] = round(c["margin_diff"], 2)

    candidates.sort(key=lambda x: abs(x.get("price_diff", 0)))

    return {
        "original": {
            "id": product_id,
            "name": product_name,
            "price": price,
            "stock": product.get("stock", 0),
            "category": category
        },
        "substitutes": candidates[:8]
    }

@api_router.post("/products/ai-substitute")
async def ai_smart_substitute(request: Request):
    """AI-powered substitution with reasoning using LLM."""
    user = await get_current_user(request)
    tid = user["tenant_id"]
    body = await request.json()
    product_name = body.get("product_name", "")
    product_id = body.get("product_id", "")

    if not product_name and not product_id:
        raise HTTPException(400, "Provide product_name or product_id")

    # Get original product info
    original = None
    if product_id:
        original = await db.products.find_one({"id": product_id, "tenant_id": tid}, {"_id": 0})
        product_name = original.get("name", product_name) if original else product_name

    # Get all in-stock products for context
    in_stock = await db.products.find(
        {"tenant_id": tid, "stock": {"$gt": 0}},
        {"_id": 0, "name": 1, "price": 1, "stock": 1, "category": 1, "cost_price": 1, "id": 1}
    ).to_list(200)

    if not in_stock:
        return {"suggestions": [], "reasoning": "No products in stock to suggest alternatives."}

    inventory_text = "\n".join([
        f"- {p['name']} | Price: ₹{p.get('price', 0)} | Cost: ₹{p.get('cost_price', 0)} | Stock: {p.get('stock', 0)} | Category: {p.get('category', 'N/A')} | ID: {p['id']}"
        for p in in_stock[:50]
    ])

    from emergentintegrations.llm.chat import LlmChat, UserMessage
    llm_key = os.environ.get("EMERGENT_LLM_KEY", "")
    chat = LlmChat(
        api_key=llm_key,
        session_id=f"substitute-{tid}-{uuid.uuid4()}",
        system_message="You are a retail product substitution expert. You help shopkeepers find alternative products when something is out of stock. For pharmacies, suggest generic alternatives. Always consider price similarity, purpose, and margin. Respond ONLY in valid JSON."
    ).with_model("openai", "gpt-5.2")

    prompt = f"""Customer wants: "{product_name}" which is out of stock or unavailable.

Available inventory:
{inventory_text}

Suggest up to 3 best substitutes. For each, explain why it's a good alternative.
Reply in this exact JSON format:
{{"suggestions": [{{"product_id": "...", "name": "...", "reason": "Brief reason why this is a good substitute", "confidence": 0.0-1.0}}], "customer_message": "A friendly message to tell the customer about the alternative"}}"""

    try:
        response = await chat.send_message(UserMessage(text=prompt))
        # Parse JSON from response
        import json as json_lib
        # Try to extract JSON from response
        resp_text = response.strip()
        if resp_text.startswith("```"):
            resp_text = resp_text.split("```")[1]
            if resp_text.startswith("json"):
                resp_text = resp_text[4:]
        parsed = json_lib.loads(resp_text)
        return parsed
    except Exception as e:
        logger.error(f"AI substitute error: {e}")
        # Fallback to simple category match
        category = original.get("category", "") if original else ""
        fallback = [p for p in in_stock if p.get("category") == category][:3]
        return {
            "suggestions": [{"product_id": p["id"], "name": p["name"], "reason": f"Same category ({category}), in stock", "confidence": 0.5} for p in fallback],
            "customer_message": f"We're sorry, {product_name} is currently unavailable. Here are some alternatives we have in stock."
        }

# ═══════════════════════════════════════════════════════════════════
#  FEATURE: AI BUSINESS PULSE — DAILY BRIEFING
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/pulse/today")
async def get_daily_pulse(request: Request):
    """Get today's AI-generated business briefing for the owner."""
    user = await get_current_user(request)
    if user["role"] not in ("OWNER", "MANAGER") and not user.get("is_platform_admin") and not user.get("is_admin"):
        raise HTTPException(403, "Requires OWNER/MANAGER or admin access")

    tid = user["tenant_id"]
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Check if pulse already generated today
    existing = await db.business_pulses.find_one({"tenant_id": tid, "date": today}, {"_id": 0})
    if existing:
        return existing

    # Generate new pulse
    return await _generate_pulse(tid, today)

@api_router.post("/pulse/generate")
async def regenerate_pulse(request: Request):
    """Force regenerate today's business pulse."""
    user = await get_current_user(request)
    if user["role"] not in ("OWNER", "MANAGER") and not user.get("is_platform_admin") and not user.get("is_admin"):
        raise HTTPException(403, "Requires OWNER/MANAGER or admin access")

    tid = user["tenant_id"]
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Delete existing and regenerate
    await db.business_pulses.delete_many({"tenant_id": tid, "date": today})
    return await _generate_pulse(tid, today)

async def _generate_pulse(tid: str, today: str):
    """Internal: Generate AI business pulse from data."""
    now = datetime.now(timezone.utc)
    yesterday_start = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0).isoformat()
    yesterday_end = (now - timedelta(days=1)).replace(hour=23, minute=59, second=59).isoformat()
    last_week_same_day = (now - timedelta(days=8)).replace(hour=0, minute=0, second=0).isoformat()
    last_week_same_day_end = (now - timedelta(days=8)).replace(hour=23, minute=59, second=59).isoformat()

    # Yesterday's data
    yesterday_invoices = await db.invoices.find(
        {"tenant_id": tid, "created_at": {"$gte": yesterday_start, "$lte": yesterday_end}},
        {"_id": 0, "grand_total": 1, "items": 1, "customer_name": 1, "customer_id": 1}
    ).to_list(10000)

    yesterday_revenue = sum(inv.get("grand_total", 0) for inv in yesterday_invoices)
    yesterday_orders = len(yesterday_invoices)

    # Last week same day comparison
    last_week_invoices = await db.invoices.find(
        {"tenant_id": tid, "created_at": {"$gte": last_week_same_day, "$lte": last_week_same_day_end}},
        {"_id": 0, "grand_total": 1}
    ).to_list(10000)
    last_week_revenue = sum(inv.get("grand_total", 0) for inv in last_week_invoices)

    # Revenue change
    revenue_change = 0
    if last_week_revenue > 0:
        revenue_change = round(((yesterday_revenue - last_week_revenue) / last_week_revenue) * 100, 1)

    # Top sellers yesterday
    product_sales = {}
    for inv in yesterday_invoices:
        for item in inv.get("items", []):
            name = item.get("name", "Unknown")
            qty = item.get("quantity", 0)
            product_sales[name] = product_sales.get(name, 0) + qty
    top_sellers = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:5]

    # Low stock items
    low_stock = await db.products.find(
        {"tenant_id": tid, "stock": {"$lte": 10, "$gt": 0}},
        {"_id": 0, "name": 1, "stock": 1}
    ).to_list(20)

    # Slow-moving items (no sales in 14 days)
    two_weeks_ago = (now - timedelta(days=14)).isoformat()
    recent_invoices = await db.invoices.find(
        {"tenant_id": tid, "created_at": {"$gte": two_weeks_ago}},
        {"_id": 0, "items": 1}
    ).to_list(10000)
    sold_products = set()
    for inv in recent_invoices:
        for item in inv.get("items", []):
            sold_products.add(item.get("name", ""))

    all_products = await db.products.find(
        {"tenant_id": tid, "stock": {"$gt": 0}},
        {"_id": 0, "name": 1, "stock": 1, "price": 1}
    ).to_list(500)
    slow_moving = [p for p in all_products if p["name"] not in sold_products][:5]

    # Build pulse data
    pulse_data = {
        "yesterday_revenue": round(yesterday_revenue, 2),
        "yesterday_orders": yesterday_orders,
        "last_week_revenue": round(last_week_revenue, 2),
        "revenue_change": revenue_change,
        "top_sellers": [{"name": name, "qty": qty} for name, qty in top_sellers],
        "low_stock_count": len(low_stock),
        "low_stock_items": low_stock[:5],
        "slow_moving_count": len(slow_moving),
        "slow_moving_items": slow_moving[:5],
    }

    # Generate AI insight using LLM
    ai_message = ""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        llm_key = os.environ.get("EMERGENT_LLM_KEY", "")
        if llm_key:
            tenant = await db.tenants.find_one({"id": tid}, {"_id": 0})
            shop_name = tenant.get("shop_name", "Your Store") if tenant else "Your Store"
            day_name = (now - timedelta(days=1)).strftime("%A")

            chat = LlmChat(
                api_key=llm_key,
                session_id=f"pulse-{tid}-{today}",
                system_message="You are a friendly retail business advisor. Generate a brief, encouraging daily briefing for a shop owner. Be concise (3-5 sentences), include actionable suggestions. Use emojis sparingly. Include specific numbers from the data."
            ).with_model("openai", "gpt-5.2")

            data_summary = f"""Shop: {shop_name}
Yesterday ({day_name}):
- Revenue: ₹{yesterday_revenue:,.2f} ({yesterday_orders} orders)
- vs last {day_name}: ₹{last_week_revenue:,.2f} ({'+' if revenue_change > 0 else ''}{revenue_change}%)
- Top sellers: {', '.join([f'{n} ({q} units)' for n, q in top_sellers[:3]]) or 'No sales yesterday'}
- Low stock items: {len(low_stock)} items below 10 units
- Slow-moving (no sales in 14 days): {len(slow_moving)} items{f' — including {slow_moving[0]["name"]} (₹{slow_moving[0].get("price", 0)})' if slow_moving else ''}"""

            response = await chat.send_message(UserMessage(text=f"Generate today's business briefing based on this data:\n\n{data_summary}"))
            ai_message = response.strip()
    except Exception as e:
        logger.error(f"AI pulse generation error: {e}")
        # Fallback to template-based message
        trend = "↑" if revenue_change > 0 else "↓" if revenue_change < 0 else "→"
        top_seller_text = f"Top seller: {top_sellers[0][0]} ({top_sellers[0][1]} units)" if top_sellers else "No sales recorded yesterday"
        low_stock_text = f"{len(low_stock)} items need restocking." if low_stock else "Stock levels look good!"
        slow_text = ""
        if slow_moving:
            sm_name = slow_moving[0].get("name", "an item")
            slow_text = f" Consider a discount on {sm_name} — no sales in 14 days."
        ai_message = f"Good morning! Yesterday's revenue was ₹{yesterday_revenue:,.2f} {trend}{abs(revenue_change)}% vs last week. {top_seller_text}. {low_stock_text}{slow_text}"

    pulse = {
        "id": str(uuid.uuid4()),
        "tenant_id": tid,
        "date": today,
        "data": pulse_data,
        "ai_message": ai_message,
        "generated_at": now.isoformat()
    }
    await db.business_pulses.insert_one({**pulse, "_id": None})
    # Remove _id for response
    pulse.pop("_id", None)
    return pulse

# ═══════════════════════════════════════════════════════════════════
#  FEATURE: PREDICTIVE CUSTOMER ENGAGEMENT (Medicine Refill Reminders)
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/refill-predictions")
async def get_refill_predictions(request: Request):
    """Detect customers who buy recurring items (e.g. monthly medicines) and predict refill dates."""
    user = await get_current_user(request)
    if user["role"] not in ("OWNER", "MANAGER") and not user.get("is_platform_admin") and not user.get("is_admin"):
        raise HTTPException(403, "Requires OWNER/MANAGER or admin access")

    tid = user["tenant_id"]
    now = datetime.now(timezone.utc)

    # Get all invoices for last 6 months to detect patterns
    six_months_ago = (now - timedelta(days=180)).isoformat()
    invoices = await db.invoices.find(
        {"tenant_id": tid, "created_at": {"$gte": six_months_ago}, "customer_id": {"$exists": True, "$ne": None}},
        {"_id": 0, "customer_id": 1, "customer_name": 1, "items": 1, "created_at": 1}
    ).sort("created_at", 1).to_list(50000)

    # Build customer purchase patterns
    customer_products = {}  # {customer_id: {product_name: [purchase_dates]}}
    for inv in invoices:
        cid = inv.get("customer_id")
        cname = inv.get("customer_name", "Unknown")
        if not cid:
            continue
        if cid not in customer_products:
            customer_products[cid] = {"name": cname, "products": {}}
        for item in inv.get("items", []):
            pname = item.get("name", "")
            if pname not in customer_products[cid]["products"]:
                customer_products[cid]["products"][pname] = []
            customer_products[cid]["products"][pname].append(inv.get("created_at", ""))

    predictions = []
    for cid, cdata in customer_products.items():
        for pname, dates in cdata["products"].items():
            if len(dates) < 2:
                continue  # Need at least 2 purchases to detect pattern

            # Calculate average interval between purchases
            sorted_dates = sorted(dates)
            intervals = []
            for i in range(1, len(sorted_dates)):
                try:
                    d1 = datetime.fromisoformat(sorted_dates[i-1].replace("Z", "+00:00"))
                    d2 = datetime.fromisoformat(sorted_dates[i].replace("Z", "+00:00"))
                    interval = (d2 - d1).days
                    if 7 <= interval <= 90:  # Only consider reasonable intervals (weekly to quarterly)
                        intervals.append(interval)
                except:
                    continue

            if not intervals:
                continue

            avg_interval = sum(intervals) / len(intervals)
            # Only flag as recurring if interval is reasonably consistent (std dev < 30% of mean)
            if len(intervals) >= 2:
                variance = sum((i - avg_interval) ** 2 for i in intervals) / len(intervals)
                std_dev = variance ** 0.5
                if std_dev > avg_interval * 0.3:
                    continue  # Too irregular, likely not a recurring purchase

            # Predict next purchase date
            try:
                last_purchase = datetime.fromisoformat(sorted_dates[-1].replace("Z", "+00:00"))
            except:
                continue
            predicted_next = last_purchase + timedelta(days=int(avg_interval))
            days_until = (predicted_next - now).days
            days_overdue = -days_until if days_until < 0 else 0

            # Only include if due within 7 days or overdue by up to 14 days
            if -14 <= days_until <= 7:
                predictions.append({
                    "customer_id": cid,
                    "customer_name": cdata["name"],
                    "product_name": pname,
                    "avg_interval_days": round(avg_interval),
                    "purchase_count": len(dates),
                    "last_purchase": sorted_dates[-1],
                    "predicted_next": predicted_next.isoformat(),
                    "days_until": days_until,
                    "days_overdue": days_overdue,
                    "is_overdue": days_until < 0,
                    "urgency": "high" if days_overdue > 3 else "medium" if days_until <= 2 else "low",
                    "suggested_message": f"Hi {cdata['name']}, time for your {pname} refill? Your usual order is ready! 💊"
                })

    # Sort by urgency (overdue first, then upcoming)
    predictions.sort(key=lambda x: x["days_until"])

    return {
        "predictions": predictions,
        "total": len(predictions),
        "summary": {
            "overdue": len([p for p in predictions if p["is_overdue"]]),
            "due_soon": len([p for p in predictions if not p["is_overdue"]])
        }
    }

@api_router.post("/refill-reminders/send")
async def send_refill_reminder(request: Request):
    """Send a refill reminder to a customer via WhatsApp."""
    user = await get_current_user(request)
    tid = user["tenant_id"]
    body = await request.json()
    customer_id = body.get("customer_id")
    product_name = body.get("product_name")
    message = body.get("message", "")

    customer = await db.customers.find_one({"id": customer_id, "tenant_id": tid}, {"_id": 0})
    if not customer:
        raise HTTPException(404, "Customer not found")

    phone = customer.get("phone", "")
    cname = customer.get("name", "Customer")

    if not message:
        tenant = await db.tenants.find_one({"id": tid}, {"_id": 0})
        shop_name = tenant.get("shop_name", "Your Store") if tenant else "Your Store"
        message = f"Hi {cname}! 👋\n\nTime for your {product_name} refill? Your usual order is ready at {shop_name}!\n\nVisit us today or reply to confirm your order.\n\n— {shop_name}\nPowered by RetailPro"

    # Log reminder
    await db.refill_reminders.insert_one({
        "id": str(uuid.uuid4()),
        "tenant_id": tid,
        "customer_id": customer_id,
        "customer_name": cname,
        "product_name": product_name,
        "phone": phone,
        "message": message,
        "status": "generated",
        "sent_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    })

    wa_link = f"https://wa.me/{phone.replace('+', '').replace(' ', '')}?text={message.replace(chr(10), '%0A').replace(' ', '%20')}" if phone else ""

    await log_audit(tid, user["id"], "refill_reminder_sent", f"Refill reminder for {cname} - {product_name}", request.client.host if request.client else "")

    return {
        "message": "Reminder generated",
        "whatsapp_link": wa_link,
        "sms_message": message,
        "status": "generated"
    }

# ═══════════════════════════════════════════════════════════════════
#  MULTI-BRANCH MANAGEMENT (Premium)
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/branches")
async def create_branch(req: BranchCreate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can manage branches")
    if req.code:
        existing = await db.branches.find_one({"tenant_id": user["tenant_id"], "code": req.code})
        if existing:
            raise HTTPException(400, "Branch code already exists")
    doc = req.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["tenant_id"] = user["tenant_id"]
    doc["is_active"] = True
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.branches.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "branch_created", f"Branch: {req.name}", request.client.host if request.client else "")
    doc.pop("_id", None)
    return doc

@api_router.get("/branches")
async def list_branches(request: Request):
    user = await get_current_user(request)
    branches = await db.branches.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("name", 1).to_list(100)
    for b in branches:
        b["product_count"] = await db.products.count_documents({"tenant_id": user["tenant_id"], "branch_id": b["id"]})
        b["user_count"] = await db.users.count_documents({"tenant_id": user["tenant_id"], "branch_id": b["id"]})
    return {"branches": branches}

@api_router.put("/branches/{branch_id}")
async def update_branch(branch_id: str, req: BranchUpdate, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can manage branches")
    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.branches.update_one({"id": branch_id, "tenant_id": user["tenant_id"]}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(404, "Branch not found")
    await log_audit(user["tenant_id"], user["id"], "branch_updated", f"Branch: {branch_id}", request.client.host if request.client else "")
    return {"message": "Branch updated"}

@api_router.delete("/branches/{branch_id}")
async def delete_branch(branch_id: str, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can manage branches")
    product_count = await db.products.count_documents({"tenant_id": user["tenant_id"], "branch_id": branch_id})
    if product_count > 0:
        raise HTTPException(400, f"Cannot delete branch with {product_count} products. Transfer them first.")
    result = await db.branches.delete_one({"id": branch_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Branch not found")
    await log_audit(user["tenant_id"], user["id"], "branch_deleted", f"Branch: {branch_id}", request.client.host if request.client else "")
    return {"message": "Branch deleted"}

@api_router.get("/branches/{branch_id}/stats")
async def branch_stats(branch_id: str, request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    branch = await db.branches.find_one({"id": branch_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not branch:
        raise HTTPException(404, "Branch not found")
    tid = user["tenant_id"]
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0).isoformat()
    total_products = await db.products.count_documents({"tenant_id": tid, "branch_id": branch_id})
    total_invoices = await db.invoices.count_documents({"tenant_id": tid, "branch_id": branch_id})
    today_invoices = await db.invoices.find({"tenant_id": tid, "branch_id": branch_id, "created_at": {"$gte": today}}, {"_id": 0, "grand_total": 1}).to_list(5000)
    today_revenue = sum(inv.get("grand_total", 0) for inv in today_invoices)
    low_stock = await db.products.count_documents({"tenant_id": tid, "branch_id": branch_id, "$expr": {"$lte": ["$stock", "$low_stock_threshold"]}})
    return {
        "branch": branch, "total_products": total_products, "total_invoices": total_invoices,
        "today_revenue": round(today_revenue, 2), "today_orders": len(today_invoices), "low_stock_count": low_stock
    }

@api_router.post("/branches/transfer-products")
async def transfer_products(request: Request):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can transfer products")
    body = await request.json()
    product_ids = body.get("product_ids", [])
    target_branch_id = body.get("target_branch_id", "")
    if not product_ids or not target_branch_id:
        raise HTTPException(400, "product_ids and target_branch_id required")
    target = await db.branches.find_one({"id": target_branch_id, "tenant_id": user["tenant_id"]})
    if not target:
        raise HTTPException(404, "Target branch not found")
    result = await db.products.update_many(
        {"id": {"$in": product_ids}, "tenant_id": user["tenant_id"]},
        {"$set": {"branch_id": target_branch_id, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    product_cache.invalidate(user["tenant_id"])
    await log_audit(user["tenant_id"], user["id"], "products_transferred", f"{result.modified_count} products → branch {target_branch_id}", request.client.host if request.client else "")
    return {"message": f"{result.modified_count} products transferred"}

# ═══════════════════════════════════════════════════════════════════
#  CATEGORY HIERARCHY MANAGEMENT
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/categories")
async def create_category(req: CategoryCreate, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot manage categories")
    if req.parent_id:
        parent = await db.categories.find_one({"id": req.parent_id, "tenant_id": user["tenant_id"]})
        if not parent:
            raise HTTPException(404, "Parent category not found")
        level = parent.get("level", 0) + 1
        path = parent.get("path", []) + [parent["id"]]
    else:
        level = 0
        path = []
    doc = req.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["tenant_id"] = user["tenant_id"]
    doc["level"] = level
    doc["path"] = path
    doc["is_active"] = True
    doc["product_count"] = 0
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.categories.insert_one(doc)
    category_cache.invalidate(user["tenant_id"])
    await log_audit(user["tenant_id"], user["id"], "category_created", f"Category: {req.name}", request.client.host if request.client else "")
    doc.pop("_id", None)
    return doc

@api_router.get("/categories")
async def list_categories_hierarchy(request: Request, flat: bool = False):
    user = await get_current_user(request)
    cats = await db.categories.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("sort_order", 1).to_list(1000)
    # Update product counts
    for c in cats:
        c["product_count"] = await db.products.count_documents({"tenant_id": user["tenant_id"], "category_id": c["id"]})
    if flat:
        return {"categories": cats}
    # Build tree
    cat_map = {c["id"]: {**c, "children": []} for c in cats}
    roots = []
    for c in cats:
        node = cat_map[c["id"]]
        if c.get("parent_id") and c["parent_id"] in cat_map:
            cat_map[c["parent_id"]]["children"].append(node)
        else:
            roots.append(node)
    return {"categories": roots, "total": len(cats)}

@api_router.put("/categories/{category_id}")
async def update_category(category_id: str, req: CategoryUpdate, request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot manage categories")
    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.categories.update_one({"id": category_id, "tenant_id": user["tenant_id"]}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(404, "Category not found")
    category_cache.invalidate(user["tenant_id"])
    return {"message": "Category updated"}

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["OWNER", "MANAGER"]:
        raise HTTPException(403, "Insufficient permissions")
    children = await db.categories.count_documents({"tenant_id": user["tenant_id"], "parent_id": category_id})
    if children > 0:
        raise HTTPException(400, f"Cannot delete category with {children} subcategories. Delete them first.")
    products = await db.products.count_documents({"tenant_id": user["tenant_id"], "category_id": category_id})
    if products > 0:
        raise HTTPException(400, f"Cannot delete category with {products} products. Reassign them first.")
    result = await db.categories.delete_one({"id": category_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Category not found")
    category_cache.invalidate(user["tenant_id"])
    return {"message": "Category deleted"}

# ═══════════════════════════════════════════════════════════════════
#  BULK UPLOAD (CSV / Excel / JSON)
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/inventory/bulk-upload")
async def bulk_upload_products(request: Request):
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot upload products")

    content_type = request.headers.get("content-type", "")
    body = await request.body()

    products = []
    errors = []

    if "application/json" in content_type:
        try:
            data = json.loads(body)
            if isinstance(data, dict):
                products = data.get("products", [])
            elif isinstance(data, list):
                products = data
        except Exception as e:
            raise HTTPException(400, f"Invalid JSON: {str(e)}")

    elif "text/csv" in content_type or "csv" in content_type:
        try:
            text = body.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                products.append(row)
        except Exception as e:
            raise HTTPException(400, f"Invalid CSV: {str(e)}")

    elif "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type:
        try:
            wb = load_workbook(filename=io.BytesIO(body), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    products.append(dict(zip(headers, row)))
        except Exception as e:
            raise HTTPException(400, f"Invalid Excel file: {str(e)}")
    else:
        raise HTTPException(400, "Unsupported format. Use JSON, CSV, or Excel (.xlsx)")

    if not products:
        raise HTTPException(400, "No products found in upload")
    if len(products) > 10000:
        raise HTTPException(400, "Maximum 10,000 products per upload")

    created = 0
    updated = 0
    skipped = 0
    branch_id = ""

    for i, p in enumerate(products):
        try:
            name = str(p.get("name", "")).strip()
            if not name:
                errors.append({"row": i + 2, "error": "Missing product name"})
                skipped += 1
                continue
            sku = str(p.get("sku", "")).strip()
            barcode_val = str(p.get("barcode", "")).strip()
            branch_id = str(p.get("branch_id", "")).strip()

            # Check for existing by SKU or barcode
            existing = None
            if sku:
                existing = await db.products.find_one({"tenant_id": user["tenant_id"], "sku": sku})
            if not existing and barcode_val:
                existing = await db.products.find_one({"tenant_id": user["tenant_id"], "barcode": barcode_val})

            doc = {
                "name": name,
                "sku": sku,
                "barcode": barcode_val,
                "category": str(p.get("category", "")).strip(),
                "category_id": str(p.get("category_id", "")).strip(),
                "price": float(p.get("price", 0) or 0),
                "cost_price": float(p.get("cost_price", 0) or 0),
                "stock": int(float(p.get("stock", 0) or 0)),
                "low_stock_threshold": int(float(p.get("low_stock_threshold", 10) or 10)),
                "unit": str(p.get("unit", "pcs")).strip() or "pcs",
                "batch_number": str(p.get("batch_number", "")).strip(),
                "expiry_date": str(p.get("expiry_date", "")).strip() if p.get("expiry_date") else "",
                "description": str(p.get("description", "")).strip(),
                "hsn_code": str(p.get("hsn_code", "")).strip(),
                "gst_rate": float(p.get("gst_rate", 0) or 0),
                "branch_id": branch_id,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }

            if existing:
                await db.products.update_one({"_id": existing["_id"]}, {"$set": doc})
                updated += 1
            else:
                doc["id"] = str(uuid.uuid4())
                doc["tenant_id"] = user["tenant_id"]
                doc["created_at"] = datetime.now(timezone.utc).isoformat()
                doc["created_by"] = user["id"]
                await db.products.insert_one(doc)
                created += 1
        except Exception as e:
            errors.append({"row": i + 2, "error": str(e)})
            skipped += 1

    product_cache.invalidate(user["tenant_id"])
    category_cache.invalidate(user["tenant_id"])
    await log_audit(user["tenant_id"], user["id"], "bulk_upload",
                    f"Created: {created}, Updated: {updated}, Skipped: {skipped}",
                    request.client.host if request.client else "")

    return {
        "message": f"Bulk upload complete",
        "created": created, "updated": updated, "skipped": skipped,
        "total_processed": len(products), "errors": errors[:50]
    }

@api_router.get("/inventory/bulk-template")
async def get_bulk_template(request: Request, format: str = "csv"):
    user = await get_current_user(request)
    headers = ["name", "sku", "barcode", "category", "price", "cost_price", "stock",
               "low_stock_threshold", "unit", "batch_number", "expiry_date", "description",
               "hsn_code", "gst_rate", "branch_id"]
    sample = ["Sample Product", "SKU001", "8901234567890", "General", "100.00", "80.00",
              "50", "10", "pcs", "BATCH001", "2027-12-31", "Sample description",
              "3004", "12", ""]

    if format == "json":
        template = {"products": [dict(zip(headers, sample))]}
        return JSONResponse(content=template, headers={"Content-Disposition": "attachment; filename=bulk_template.json"})
    elif format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(headers)
        ws.append(sample)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               headers={"Content-Disposition": "attachment; filename=bulk_template.xlsx"})
    else:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerow(sample)
        buf.seek(0)
        return StreamingResponse(io.BytesIO(buf.getvalue().encode()), media_type="text/csv",
                               headers={"Content-Disposition": "attachment; filename=bulk_template.csv"})

# ═══════════════════════════════════════════════════════════════════
#  BARCODE LABEL PRINTING
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/inventory/barcode-labels")
async def generate_barcode_labels(request: Request):
    user = await get_current_user(request)
    body = await request.json()
    product_ids = body.get("product_ids", [])
    label_size = body.get("label_size", "medium")  # small, medium, large
    copies = min(int(body.get("copies", 1)), 10)

    if not product_ids:
        raise HTTPException(400, "No product IDs provided")
    if len(product_ids) > 100:
        raise HTTPException(400, "Maximum 100 products per batch")

    products = []
    for pid in product_ids:
        p = await db.products.find_one({"id": pid, "tenant_id": user["tenant_id"]}, {"_id": 0})
        if p:
            products.append(p)

    if not products:
        raise HTTPException(404, "No products found")

    # Generate PDF with barcode labels
    sizes = {
        "small": {"w": 38*mm, "h": 25*mm, "cols": 5, "font": 6},
        "medium": {"w": 50*mm, "h": 30*mm, "cols": 4, "font": 7},
        "large": {"w": 70*mm, "h": 40*mm, "cols": 3, "font": 8},
    }
    sz = sizes.get(label_size, sizes["medium"])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=10*mm, bottomMargin=10*mm, leftMargin=5*mm, rightMargin=5*mm)
    styles = getSampleStyleSheet()
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontSize=sz["font"], alignment=TA_CENTER, leading=sz["font"] + 2)
    price_style = ParagraphStyle('Price', parent=styles['Normal'], fontSize=sz["font"] + 2, alignment=TA_CENTER, leading=sz["font"] + 4, fontName='Helvetica-Bold')

    all_labels = []
    for p in products:
        for _ in range(copies):
            all_labels.append(p)

    # Build table of labels
    table_data = []
    row = []
    for p in all_labels:
        barcode_val = p.get("barcode") or p.get("sku") or p.get("id", "")[:12]
        # Generate barcode image
        try:
            code128 = python_barcode.get("code128", barcode_val, writer=ImageWriter())
            bc_buffer = io.BytesIO()
            code128.write(bc_buffer, options={"module_width": 0.25, "module_height": 8, "font_size": 6, "text_distance": 1, "quiet_zone": 1})
            bc_buffer.seek(0)
            from reportlab.platypus import Image as RLImage
            bc_img = RLImage(bc_buffer, width=sz["w"] - 4*mm, height=12*mm)
        except Exception:
            bc_img = Paragraph(barcode_val, label_style)

        cell_content = [
            Paragraph(p.get("name", "")[:30], label_style),
            bc_img,
            Paragraph(f"₹{p.get('price', 0):.2f}", price_style),
        ]
        cell = Table([[c] for c in cell_content], colWidths=[sz["w"] - 2*mm])
        cell.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#94A3B8')),
            ('TOPPADDING', (0, 0), (-1, -1), 1*mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1*mm),
        ]))
        row.append(cell)
        if len(row) >= sz["cols"]:
            table_data.append(row)
            row = []

    if row:
        while len(row) < sz["cols"]:
            row.append("")
        table_data.append(row)

    if table_data:
        main_table = Table(table_data, colWidths=[sz["w"]] * sz["cols"])
        main_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 1*mm),
            ('RIGHTPADDING', (0, 0), (-1, -1), 1*mm),
            ('TOPPADDING', (0, 0), (-1, -1), 1*mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1*mm),
        ]))
        doc.build([main_table])
    else:
        doc.build([Paragraph("No labels to generate", styles['Normal'])])

    buffer.seek(0)
    await log_audit(user["tenant_id"], user["id"], "barcode_labels_generated", f"{len(all_labels)} labels", request.client.host if request.client else "")
    return StreamingResponse(buffer, media_type="application/pdf",
                           headers={"Content-Disposition": "attachment; filename=barcode_labels.pdf"})

# ═══════════════════════════════════════════════════════════════════
#  INVOICE EMAIL SENDING
# ═══════════════════════════════════════════════════════════════════

@api_router.post("/pos/invoices/{invoice_id}/email")
async def email_invoice(invoice_id: str, request: Request):
    user = await get_current_user(request)
    body = await request.json()
    recipient_email = body.get("email", "")
    if not recipient_email:
        raise HTTPException(400, "Recipient email required")

    invoice = await db.invoices.find_one({"id": invoice_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    shop_name = tenant.get("shop_name", "RetailPro") if tenant else "RetailPro"

    # Build HTML email
    items_html = ""
    for item in invoice.get("items", []):
        items_html += f"<tr><td style='padding:6px;border-bottom:1px solid #eee'>{item['name']}</td><td style='padding:6px;border-bottom:1px solid #eee;text-align:center'>{item['quantity']}</td><td style='padding:6px;border-bottom:1px solid #eee;text-align:right'>₹{item['price']:.2f}</td><td style='padding:6px;border-bottom:1px solid #eee;text-align:right'>₹{item.get('total', item['price']*item['quantity']):.2f}</td></tr>"

    html_body = f"""
    <div style='max-width:600px;margin:auto;font-family:sans-serif;color:#333'>
      <div style='background:#0F172A;color:white;padding:20px;text-align:center'>
        <h1 style='margin:0;font-size:24px'>{shop_name}</h1>
        <p style='margin:5px 0 0;opacity:0.8'>Invoice #{invoice.get('invoice_number','')}</p>
      </div>
      <div style='padding:20px'>
        <p><strong>Date:</strong> {invoice.get('created_at','')[:19].replace('T',' ')}</p>
        <p><strong>Customer:</strong> {invoice.get('customer_name','Walk-in')}</p>
        <p><strong>Payment:</strong> {invoice.get('payment_method','cash').upper()}</p>
        <table style='width:100%;border-collapse:collapse;margin:15px 0'>
          <tr style='background:#F1F5F9'><th style='padding:8px;text-align:left'>Item</th><th style='padding:8px;text-align:center'>Qty</th><th style='padding:8px;text-align:right'>Price</th><th style='padding:8px;text-align:right'>Total</th></tr>
          {items_html}
        </table>
        <div style='text-align:right;margin-top:10px'>
          <p>Subtotal: ₹{invoice.get('subtotal',0):.2f}</p>
          <p>Tax (GST): ₹{invoice.get('tax_total',0):.2f}</p>
          {"<p>Discount: -₹" + f"{invoice.get('discount',0):.2f}</p>" if invoice.get('discount',0) > 0 else ""}
          <p style='font-size:18px;font-weight:bold;color:#0F172A'>Grand Total: ₹{invoice.get('grand_total',0):.2f}</p>
        </div>
      </div>
      <div style='background:#F8FAFC;padding:15px;text-align:center;color:#64748B;font-size:12px'>
        Thank you for your business! | Powered by RetailPro
      </div>
    </div>"""

    subject = f"Invoice {invoice.get('invoice_number','')} from {shop_name}"
    sent = await send_email_notification(user["tenant_id"], recipient_email, subject, html_body)

    await log_audit(user["tenant_id"], user["id"], "invoice_emailed",
                    f"Invoice {invoice.get('invoice_number','')} emailed to {recipient_email}",
                    request.client.host if request.client else "")

    return {"message": "Invoice email sent" if sent else "Email sending failed. Check SMTP settings.", "sent": sent}

# ═══════════════════════════════════════════════════════════════════
#  ADVANCED PROFIT MARGIN DASHBOARD
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/reports/profit-dashboard")
async def advanced_profit_dashboard(request: Request, period: str = "30d", branch_id: str = ""):
    user = await get_current_user(request)
    if not get_user_permission(user, "can_view_revenue"):
        raise HTTPException(403, "Permission denied")

    tid = user["tenant_id"]
    days = {"7d": 7, "30d": 30, "90d": 90, "365d": 365}.get(period, 30)
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    prev_start = (datetime.now(timezone.utc) - timedelta(days=days * 2)).isoformat()

    query = {"tenant_id": tid, "created_at": {"$gte": start_date}}
    prev_query = {"tenant_id": tid, "created_at": {"$gte": prev_start, "$lt": start_date}}
    if branch_id:
        query["branch_id"] = branch_id
        prev_query["branch_id"] = branch_id

    invoices = await db.invoices.find(query, {"_id": 0}).to_list(50000)
    prev_invoices = await db.invoices.find(prev_query, {"_id": 0}).to_list(50000)

    # Product-level profit analysis
    prod_map = {}
    products_list = await db.products.find({"tenant_id": tid}, {"_id": 0, "id": 1, "cost_price": 1, "category": 1, "name": 1}).to_list(50000)
    cost_map = {p["id"]: p for p in products_list}

    product_data = {}
    daily_profit = {}
    category_profit = {}

    for inv in invoices:
        inv_date = inv.get("created_at", "")[:10]
        for item in inv.get("items", []):
            pid = item.get("product_id", "")
            revenue = item.get("total", item["price"] * item["quantity"])
            cost = cost_map.get(pid, {}).get("cost_price", 0) * item["quantity"]
            profit = revenue - cost
            cat = cost_map.get(pid, {}).get("category", "Uncategorized") or "Uncategorized"

            if pid not in product_data:
                product_data[pid] = {"name": item["name"], "revenue": 0, "cost": 0, "qty": 0, "category": cat}
            product_data[pid]["revenue"] += revenue
            product_data[pid]["cost"] += cost
            product_data[pid]["qty"] += item["quantity"]

            daily_profit[inv_date] = daily_profit.get(inv_date, {"revenue": 0, "cost": 0})
            daily_profit[inv_date]["revenue"] += revenue
            daily_profit[inv_date]["cost"] += cost

            if cat not in category_profit:
                category_profit[cat] = {"revenue": 0, "cost": 0, "qty": 0}
            category_profit[cat]["revenue"] += revenue
            category_profit[cat]["cost"] += cost
            category_profit[cat]["qty"] += item["quantity"]

    # Previous period totals
    prev_revenue = sum(inv.get("grand_total", 0) for inv in prev_invoices)
    prev_cost = 0
    for inv in prev_invoices:
        for item in inv.get("items", []):
            prev_cost += cost_map.get(item.get("product_id", ""), {}).get("cost_price", 0) * item["quantity"]

    total_revenue = sum(d["revenue"] for d in product_data.values())
    total_cost = sum(d["cost"] for d in product_data.values())
    total_profit = total_revenue - total_cost
    prev_profit = prev_revenue - prev_cost

    # Top/bottom products by margin
    products_margins = []
    for pid, data in product_data.items():
        profit = data["revenue"] - data["cost"]
        margin = (profit / data["revenue"] * 100) if data["revenue"] > 0 else 0
        products_margins.append({
            "product_id": pid, "name": data["name"], "category": data["category"],
            "revenue": round(data["revenue"], 2), "cost": round(data["cost"], 2),
            "profit": round(profit, 2), "margin_pct": round(margin, 1), "qty": data["qty"]
        })
    products_margins.sort(key=lambda x: x["profit"], reverse=True)

    # Daily trend
    daily_trend = []
    for i in range(days - 1, -1, -1):
        d = (datetime.now(timezone.utc) - timedelta(days=i)).strftime("%Y-%m-%d")
        dp = daily_profit.get(d, {"revenue": 0, "cost": 0})
        daily_trend.append({"date": d, "revenue": round(dp["revenue"], 2), "cost": round(dp["cost"], 2), "profit": round(dp["revenue"] - dp["cost"], 2)})

    # Category breakdown
    categories = []
    for cat, data in category_profit.items():
        profit = data["revenue"] - data["cost"]
        margin = (profit / data["revenue"] * 100) if data["revenue"] > 0 else 0
        share = (data["revenue"] / total_revenue * 100) if total_revenue > 0 else 0
        categories.append({
            "category": cat, "revenue": round(data["revenue"], 2), "cost": round(data["cost"], 2),
            "profit": round(profit, 2), "margin_pct": round(margin, 1), "qty": data["qty"],
            "revenue_share": round(share, 1)
        })
    categories.sort(key=lambda x: x["revenue"], reverse=True)

    return {
        "summary": {
            "total_revenue": round(total_revenue, 2), "total_cost": round(total_cost, 2),
            "total_profit": round(total_profit, 2),
            "overall_margin": round((total_profit / total_revenue * 100) if total_revenue > 0 else 0, 1),
            "prev_revenue": round(prev_revenue, 2), "prev_profit": round(prev_profit, 2),
            "revenue_change": round(((total_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0, 1),
            "profit_change": round(((total_profit - prev_profit) / prev_profit * 100) if prev_profit > 0 else 0, 1),
        },
        "top_products": products_margins[:15],
        "bottom_products": list(reversed(products_margins[-10:])) if len(products_margins) > 10 else [],
        "daily_trend": daily_trend[-30:],
        "categories": categories,
        "period": period
    }

# ═══════════════════════════════════════════════════════════════════
#  BATCH EXPIRY ALERT SYSTEM
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/inventory/expiry-dashboard")
async def expiry_dashboard(request: Request, branch_id: str = ""):
    user = await get_current_user(request)
    tid = user["tenant_id"]
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    cutoff_30 = (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%Y-%m-%d")
    cutoff_90 = (datetime.now(timezone.utc) + timedelta(days=90)).strftime("%Y-%m-%d")
    cutoff_180 = (datetime.now(timezone.utc) + timedelta(days=180)).strftime("%Y-%m-%d")

    query = {"tenant_id": tid, "expiry_date": {"$nin": [None, ""]}}
    if branch_id:
        query["branch_id"] = branch_id

    products = await db.products.find(query, {"_id": 0}).sort("expiry_date", 1).to_list(5000)

    expired = []
    critical = []  # 0-30 days
    warning = []   # 30-90 days
    notice = []    # 90-180 days
    ok = []        # 180+ days

    total_expired_value = 0
    total_at_risk_value = 0

    for p in products:
        exp = p.get("expiry_date", "")
        if not exp:
            continue
        exp_date = exp[:10]
        stock_value = p.get("stock", 0) * p.get("cost_price", 0)

        if exp_date <= today:
            p["status"] = "expired"
            p["days_until_expiry"] = 0
            expired.append(p)
            total_expired_value += stock_value
        elif exp_date <= cutoff_30:
            days_left = (datetime.fromisoformat(exp_date) - datetime.fromisoformat(today)).days
            p["status"] = "critical"
            p["days_until_expiry"] = days_left
            critical.append(p)
            total_at_risk_value += stock_value
        elif exp_date <= cutoff_90:
            days_left = (datetime.fromisoformat(exp_date) - datetime.fromisoformat(today)).days
            p["status"] = "warning"
            p["days_until_expiry"] = days_left
            warning.append(p)
        elif exp_date <= cutoff_180:
            days_left = (datetime.fromisoformat(exp_date) - datetime.fromisoformat(today)).days
            p["status"] = "notice"
            p["days_until_expiry"] = days_left
            notice.append(p)
        else:
            days_left = (datetime.fromisoformat(exp_date) - datetime.fromisoformat(today)).days
            p["status"] = "ok"
            p["days_until_expiry"] = days_left
            ok.append(p)

    # Category breakdown of expiring products
    expiring_by_category = {}
    for p in expired + critical + warning:
        cat = p.get("category", "Uncategorized") or "Uncategorized"
        if cat not in expiring_by_category:
            expiring_by_category[cat] = {"count": 0, "value": 0}
        expiring_by_category[cat]["count"] += 1
        expiring_by_category[cat]["value"] += p.get("stock", 0) * p.get("cost_price", 0)

    return {
        "summary": {
            "total_tracked": len(products), "total_expired": len(expired),
            "total_critical": len(critical), "total_warning": len(warning),
            "total_notice": len(notice), "total_ok": len(ok),
            "expired_value": round(total_expired_value, 2),
            "at_risk_value": round(total_at_risk_value, 2),
        },
        "expired": expired[:50], "critical": critical[:50],
        "warning": warning[:50], "notice": notice[:50],
        "by_category": [{"category": k, **v} for k, v in sorted(expiring_by_category.items(), key=lambda x: x[1]["count"], reverse=True)],
    }

# ═══════════════════════════════════════════════════════════════════
#  SUPERMARKET ADVANCED ANALYTICS
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/analytics/sales-trends")
async def sales_trends(request: Request, period: str = "30d", branch_id: str = ""):
    user = await get_current_user(request)
    if user["role"] not in ("OWNER", "MANAGER"):
        raise HTTPException(403, "Requires OWNER or MANAGER")
    tid = user["tenant_id"]
    days = {"7d": 7, "30d": 30, "90d": 90}.get(period, 30)

    query = {"tenant_id": tid}
    if branch_id:
        query["branch_id"] = branch_id

    # Hourly distribution (all time last N days)
    start = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    query["created_at"] = {"$gte": start}
    invoices = await db.invoices.find(query, {"_id": 0, "created_at": 1, "grand_total": 1, "items": 1}).to_list(50000)

    hourly = [{"hour": h, "orders": 0, "revenue": 0} for h in range(24)]
    daily = {}
    weekday = [{"day": d, "orders": 0, "revenue": 0} for d in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]]

    for inv in invoices:
        created = inv.get("created_at", "")
        try:
            dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
            h = dt.hour
            hourly[h]["orders"] += 1
            hourly[h]["revenue"] += inv.get("grand_total", 0)
            wd = dt.weekday()
            weekday[wd]["orders"] += 1
            weekday[wd]["revenue"] += inv.get("grand_total", 0)
            d = created[:10]
            if d not in daily:
                daily[d] = {"date": d, "orders": 0, "revenue": 0, "items_sold": 0}
            daily[d]["orders"] += 1
            daily[d]["revenue"] += inv.get("grand_total", 0)
            daily[d]["items_sold"] += sum(item.get("quantity", 0) for item in inv.get("items", []))
        except Exception:
            pass

    for h in hourly:
        h["revenue"] = round(h["revenue"], 2)
    for w in weekday:
        w["revenue"] = round(w["revenue"], 2)

    daily_sorted = sorted(daily.values(), key=lambda x: x["date"])
    for d in daily_sorted:
        d["revenue"] = round(d["revenue"], 2)

    # Peak hours
    peak_hour = max(hourly, key=lambda x: x["revenue"])
    peak_day = max(weekday, key=lambda x: x["revenue"])
    avg_daily_revenue = round(sum(d["revenue"] for d in daily_sorted) / max(len(daily_sorted), 1), 2)

    return {
        "hourly": hourly, "daily": daily_sorted[-30:], "weekday": weekday,
        "insights": {
            "peak_hour": peak_hour["hour"], "peak_hour_revenue": peak_hour["revenue"],
            "peak_day": peak_day["day"], "peak_day_revenue": peak_day["revenue"],
            "avg_daily_revenue": avg_daily_revenue,
            "total_orders": len(invoices),
            "total_revenue": round(sum(inv.get("grand_total", 0) for inv in invoices), 2),
        }
    }

@api_router.get("/analytics/branch-comparison")
async def branch_comparison(request: Request, period: str = "30d"):
    user = await get_current_user(request)
    await require_premium(user)
    if user["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER")
    tid = user["tenant_id"]
    days = {"7d": 7, "30d": 30, "90d": 90}.get(period, 30)
    start = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    branches = await db.branches.find({"tenant_id": tid, "is_active": True}, {"_id": 0}).to_list(100)
    if not branches:
        return {"branches": [], "message": "No branches found"}

    result = []
    for b in branches:
        bid = b["id"]
        invoices = await db.invoices.find({"tenant_id": tid, "branch_id": bid, "created_at": {"$gte": start}}, {"_id": 0, "grand_total": 1}).to_list(50000)
        revenue = sum(inv.get("grand_total", 0) for inv in invoices)
        products = await db.products.count_documents({"tenant_id": tid, "branch_id": bid})
        low_stock = await db.products.count_documents({"tenant_id": tid, "branch_id": bid, "$expr": {"$lte": ["$stock", "$low_stock_threshold"]}})
        result.append({
            "branch_id": bid, "name": b.get("name", ""),
            "revenue": round(revenue, 2), "orders": len(invoices),
            "avg_order_value": round(revenue / max(len(invoices), 1), 2),
            "products": products, "low_stock": low_stock,
        })

    result.sort(key=lambda x: x["revenue"], reverse=True)
    return {"branches": result, "period": period}

@api_router.get("/analytics/customer-rfm")
async def customer_rfm_analysis(request: Request, branch_id: str = ""):
    """RFM (Recency, Frequency, Monetary) analysis for customer segmentation"""
    user = await get_current_user(request)
    if user["role"] not in ("OWNER", "MANAGER"):
        raise HTTPException(403, "Requires OWNER or MANAGER")
    tid = user["tenant_id"]

    query = {"tenant_id": tid}
    if branch_id:
        query["branch_id"] = branch_id

    invoices = await db.invoices.find(query, {"_id": 0, "customer_id": 1, "customer_name": 1, "grand_total": 1, "created_at": 1}).to_list(100000)

    customer_data = {}
    now = datetime.now(timezone.utc)
    for inv in invoices:
        cid = inv.get("customer_id", "") or inv.get("customer_name", "Walk-in")
        if cid == "Walk-in" or not cid:
            continue
        if cid not in customer_data:
            customer_data[cid] = {"name": inv.get("customer_name", ""), "last_purchase": "", "frequency": 0, "monetary": 0}
        customer_data[cid]["frequency"] += 1
        customer_data[cid]["monetary"] += inv.get("grand_total", 0)
        created = inv.get("created_at", "")
        if created > customer_data[cid]["last_purchase"]:
            customer_data[cid]["last_purchase"] = created

    segments = {"champions": [], "loyal": [], "at_risk": [], "lost": [], "new": []}
    for cid, data in customer_data.items():
        try:
            last = datetime.fromisoformat(data["last_purchase"].replace("Z", "+00:00"))
            recency_days = (now - last).days
        except Exception:
            recency_days = 999

        # Simple RFM scoring
        r_score = 5 if recency_days <= 7 else 4 if recency_days <= 30 else 3 if recency_days <= 60 else 2 if recency_days <= 90 else 1
        f_score = 5 if data["frequency"] >= 20 else 4 if data["frequency"] >= 10 else 3 if data["frequency"] >= 5 else 2 if data["frequency"] >= 2 else 1
        m_score = 5 if data["monetary"] >= 50000 else 4 if data["monetary"] >= 20000 else 3 if data["monetary"] >= 5000 else 2 if data["monetary"] >= 1000 else 1

        total_score = r_score + f_score + m_score
        if total_score >= 13:
            segment = "champions"
        elif total_score >= 10:
            segment = "loyal"
        elif r_score <= 2 and f_score >= 3:
            segment = "at_risk"
        elif r_score <= 2:
            segment = "lost"
        else:
            segment = "new"

        entry = {
            "customer_id": cid, "name": data["name"],
            "recency_days": recency_days, "frequency": data["frequency"],
            "monetary": round(data["monetary"], 2),
            "r_score": r_score, "f_score": f_score, "m_score": m_score,
            "segment": segment
        }
        segments[segment].append(entry)

    for seg in segments.values():
        seg.sort(key=lambda x: x["monetary"], reverse=True)

    return {
        "segments": {k: v[:20] for k, v in segments.items()},
        "summary": {k: len(v) for k, v in segments.items()},
        "total_customers": len(customer_data)
    }

@api_router.get("/analytics/product-performance")
async def product_performance(request: Request, period: str = "30d", branch_id: str = ""):
    user = await get_current_user(request)
    if user["role"] not in ("OWNER", "MANAGER"):
        raise HTTPException(403, "Requires OWNER or MANAGER")
    tid = user["tenant_id"]
    days = {"7d": 7, "30d": 30, "90d": 90}.get(period, 30)
    start = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    query = {"tenant_id": tid, "created_at": {"$gte": start}}
    if branch_id:
        query["branch_id"] = branch_id

    invoices = await db.invoices.find(query, {"_id": 0}).to_list(50000)

    product_perf = {}
    for inv in invoices:
        for item in inv.get("items", []):
            pid = item.get("product_id", "")
            if pid not in product_perf:
                product_perf[pid] = {"name": item["name"], "qty": 0, "revenue": 0, "order_count": 0}
            product_perf[pid]["qty"] += item["quantity"]
            product_perf[pid]["revenue"] += item.get("total", item["price"] * item["quantity"])
            product_perf[pid]["order_count"] += 1

    # Get cost prices
    for pid in product_perf:
        prod = await db.products.find_one({"id": pid, "tenant_id": tid}, {"_id": 0, "cost_price": 1, "stock": 1, "category": 1})
        if prod:
            product_perf[pid]["cost_price"] = prod.get("cost_price", 0)
            product_perf[pid]["stock"] = prod.get("stock", 0)
            product_perf[pid]["category"] = prod.get("category", "")

    scored = []
    for pid, data in product_perf.items():
        profit = data["revenue"] - data.get("cost_price", 0) * data["qty"]
        velocity = data["qty"] / days
        margin = (profit / data["revenue"] * 100) if data["revenue"] > 0 else 0
        # Performance score (0-100)
        score = min(100, int(velocity * 20 + margin * 0.5 + data["order_count"] * 0.5))
        scored.append({
            "product_id": pid, "name": data["name"], "category": data.get("category", ""),
            "qty_sold": data["qty"], "revenue": round(data["revenue"], 2),
            "profit": round(profit, 2), "margin_pct": round(margin, 1),
            "velocity_per_day": round(velocity, 2), "order_count": data["order_count"],
            "stock": data.get("stock", 0), "performance_score": score,
        })

    scored.sort(key=lambda x: x["performance_score"], reverse=True)

    # Slow movers
    all_products = await db.products.find({"tenant_id": tid, "stock": {"$gt": 0}}, {"_id": 0, "id": 1, "name": 1, "stock": 1, "price": 1, "category": 1}).to_list(10000)
    sold_ids = set(product_perf.keys())
    slow_movers = [{"product_id": p["id"], "name": p["name"], "stock": p["stock"], "value": round(p["stock"] * p.get("price", 0), 2), "category": p.get("category", "")}
                   for p in all_products if p["id"] not in sold_ids]
    slow_movers.sort(key=lambda x: x["value"], reverse=True)

    return {
        "top_performers": scored[:20],
        "bottom_performers": list(reversed(scored[-10:])) if len(scored) > 10 else [],
        "slow_movers": slow_movers[:20],
        "total_products_sold": len(scored),
        "total_slow_movers": len(slow_movers),
    }

# ═══════════════════════════════════════════════════════════════════
#  SUPERMARKET-SCALE: CROSS-BRANCH AVAILABILITY & TRANSFER REQUESTS
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/inventory/cross-branch/{product_id}")
async def cross_branch_availability(product_id: str, request: Request):
    """Check product availability across all branches"""
    user = await get_current_user(request)
    tid = user["tenant_id"]

    # Find the product in any branch
    products = await db.products.find(
        {"tenant_id": tid, "$or": [{"id": product_id}, {"sku": product_id}, {"barcode": product_id}]},
        {"_id": 0}
    ).to_list(100)

    if not products:
        # Search by name (partial match for the same product across branches)
        sample = await db.products.find_one({"id": product_id, "tenant_id": tid}, {"_id": 0, "name": 1, "sku": 1})
        if sample:
            products = await db.products.find(
                {"tenant_id": tid, "name": sample["name"]}, {"_id": 0}
            ).to_list(100)

    # Group by branch
    branches = await db.branches.find({"tenant_id": tid}, {"_id": 0}).to_list(100)
    branch_map = {b["id"]: b for b in branches}
    branch_map[""] = {"name": "Unassigned", "id": "", "address": ""}

    availability = []
    for p in products:
        bid = p.get("branch_id", "")
        branch = branch_map.get(bid, {"name": "Unknown", "id": bid})
        availability.append({
            "product_id": p["id"], "name": p.get("name", ""),
            "branch_id": bid, "branch_name": branch.get("name", ""),
            "branch_address": branch.get("address", ""),
            "stock": p.get("stock", 0), "price": p.get("price", 0),
            "sku": p.get("sku", ""), "barcode": p.get("barcode", ""),
        })

    availability.sort(key=lambda x: x["stock"], reverse=True)
    total_stock = sum(a["stock"] for a in availability)

    return {
        "product_name": products[0]["name"] if products else "",
        "total_stock_all_branches": total_stock,
        "branches": availability,
        "branch_count": len(availability),
    }

@api_router.post("/transfer-requests")
async def create_transfer_request(request: Request):
    """Staff/Manager creates a transfer request from another branch"""
    user = await get_current_user(request)
    body = await request.json()
    product_id = body.get("product_id", "")
    source_branch_id = body.get("source_branch_id", "")
    target_branch_id = body.get("target_branch_id", "")
    quantity = int(body.get("quantity", 1))
    reason = body.get("reason", "")

    if not product_id or not source_branch_id or not target_branch_id:
        raise HTTPException(400, "product_id, source_branch_id, and target_branch_id required")
    if source_branch_id == target_branch_id:
        raise HTTPException(400, "Source and target branch cannot be the same")
    if quantity < 1:
        raise HTTPException(400, "Quantity must be at least 1")

    # Verify product exists in source branch
    product = await db.products.find_one(
        {"id": product_id, "tenant_id": user["tenant_id"], "branch_id": source_branch_id}, {"_id": 0}
    )
    if not product:
        raise HTTPException(404, "Product not found in source branch")
    if product.get("stock", 0) < quantity:
        raise HTTPException(400, f"Insufficient stock. Available: {product.get('stock', 0)}")

    doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": user["tenant_id"],
        "product_id": product_id,
        "product_name": product.get("name", ""),
        "source_branch_id": source_branch_id,
        "target_branch_id": target_branch_id,
        "quantity": quantity,
        "reason": reason,
        "status": "pending",
        "requested_by": user["id"],
        "requested_by_name": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.transfer_requests.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "transfer_request_created",
                    f"Transfer {quantity}x {product.get('name','')} from branch {source_branch_id} to {target_branch_id}",
                    request.client.host if request.client else "")
    doc.pop("_id", None)
    return doc

@api_router.get("/transfer-requests")
async def list_transfer_requests(request: Request, status: str = "", page: int = 1, limit: int = 20):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    user_branch = user.get("branch_id", "")
    # Staff see only requests involving their branch
    if user["role"] == "STAFF" and user_branch:
        query["$or"] = [{"source_branch_id": user_branch}, {"target_branch_id": user_branch}]
    if status:
        query["status"] = status

    total = await db.transfer_requests.count_documents(query)
    skip = (page - 1) * limit
    requests_list = await db.transfer_requests.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)

    # Enrich with branch names
    branches = await db.branches.find({"tenant_id": user["tenant_id"]}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    bmap = {b["id"]: b["name"] for b in branches}
    for r in requests_list:
        r["source_branch_name"] = bmap.get(r.get("source_branch_id", ""), "Unknown")
        r["target_branch_name"] = bmap.get(r.get("target_branch_id", ""), "Unknown")

    return {"requests": requests_list, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.put("/transfer-requests/{request_id}")
async def handle_transfer_request(request_id: str, request: Request):
    """OWNER/MANAGER approves or rejects a transfer request"""
    user = await get_current_user(request)
    if user["role"] == "STAFF":
        raise HTTPException(403, "Staff cannot approve transfer requests")
    body = await request.json()
    action = body.get("action", "")  # approve / reject

    tr = await db.transfer_requests.find_one({"id": request_id, "tenant_id": user["tenant_id"]})
    if not tr:
        raise HTTPException(404, "Transfer request not found")
    if tr["status"] != "pending":
        raise HTTPException(400, f"Request is already {tr['status']}")

    if action == "approve":
        # Deduct from source, add to target
        source_product = await db.products.find_one(
            {"id": tr["product_id"], "tenant_id": user["tenant_id"], "branch_id": tr["source_branch_id"]}
        )
        if not source_product or source_product.get("stock", 0) < tr["quantity"]:
            raise HTTPException(400, "Insufficient stock in source branch")

        # Deduct from source
        await db.products.update_one(
            {"id": tr["product_id"], "tenant_id": user["tenant_id"], "branch_id": tr["source_branch_id"]},
            {"$inc": {"stock": -tr["quantity"]}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
        )

        # Find or create in target branch
        target_product = await db.products.find_one(
            {"tenant_id": user["tenant_id"], "branch_id": tr["target_branch_id"],
             "$or": [{"id": tr["product_id"]}, {"sku": source_product.get("sku", "")}, {"name": source_product.get("name", "")}]}
        )

        if target_product:
            await db.products.update_one(
                {"_id": target_product["_id"]},
                {"$inc": {"stock": tr["quantity"]}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
            )
        else:
            # Clone product to target branch
            new_product = {k: v for k, v in source_product.items() if k != "_id"}
            new_product["id"] = str(uuid.uuid4())
            new_product["branch_id"] = tr["target_branch_id"]
            new_product["stock"] = tr["quantity"]
            new_product["created_at"] = datetime.now(timezone.utc).isoformat()
            await db.products.insert_one(new_product)

        await db.transfer_requests.update_one(
            {"id": request_id},
            {"$set": {"status": "approved", "approved_by": user["id"], "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        product_cache.invalidate(user["tenant_id"])
        await log_audit(user["tenant_id"], user["id"], "transfer_approved",
                        f"Approved transfer of {tr['quantity']}x {tr.get('product_name','')}",
                        request.client.host if request.client else "")
        return {"message": "Transfer approved and stock updated"}

    elif action == "reject":
        await db.transfer_requests.update_one(
            {"id": request_id},
            {"$set": {"status": "rejected", "rejected_by": user["id"], "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        await log_audit(user["tenant_id"], user["id"], "transfer_rejected",
                        f"Rejected transfer of {tr['quantity']}x {tr.get('product_name','')}",
                        request.client.host if request.client else "")
        return {"message": "Transfer rejected"}
    else:
        raise HTTPException(400, "Action must be 'approve' or 'reject'")

# ═══════════════════════════════════════════════════════════════════
#  SUPERMARKET-SCALE: STAFF BRANCH ASSIGNMENT
# ═══════════════════════════════════════════════════════════════════

@api_router.put("/users/{user_id}/assign-branch")
async def assign_user_branch(user_id: str, request: Request):
    """OWNER assigns a staff/manager to a specific branch"""
    admin = await get_current_user(request)
    if admin["role"] != "OWNER":
        raise HTTPException(403, "Only OWNER can assign branches")

    body = await request.json()
    branch_id = body.get("branch_id", "")

    if branch_id:
        branch = await db.branches.find_one({"id": branch_id, "tenant_id": admin["tenant_id"]})
        if not branch:
            raise HTTPException(404, "Branch not found")

    target = await db.users.find_one({"_id": ObjectId(user_id), "tenant_id": admin["tenant_id"]})
    if not target:
        raise HTTPException(404, "User not found")

    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"branch_id": branch_id, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    branch_name = branch.get("name", "None") if branch_id else "All Branches"
    await log_audit(admin["tenant_id"], admin["id"], "user_branch_assigned",
                    f"User {target.get('email','')} assigned to branch: {branch_name}",
                    request.client.host if request.client else "", "security")
    return {"message": f"User assigned to {branch_name}", "branch_id": branch_id}

@api_router.get("/users/{user_id}/branch")
async def get_user_branch(user_id: str, request: Request):
    admin = await get_current_user(request)
    target = await db.users.find_one({"_id": ObjectId(user_id), "tenant_id": admin["tenant_id"]}, {"_id": 0, "branch_id": 1, "name": 1, "email": 1})
    if not target:
        raise HTTPException(404, "User not found")
    branch = None
    if target.get("branch_id"):
        branch = await db.branches.find_one({"id": target["branch_id"], "tenant_id": admin["tenant_id"]}, {"_id": 0, "name": 1, "code": 1})
    return {"user_id": user_id, "branch_id": target.get("branch_id", ""), "branch": branch}

# ═══════════════════════════════════════════════════════════════════
#  SUPERMARKET-SCALE: SMART SEARCH WITH CATEGORY SUGGESTIONS
# ═══════════════════════════════════════════════════════════════════

@api_router.get("/search/products")
async def smart_product_search(request: Request, q: str = "", category_id: str = "", branch_id: str = "",
                                page: int = 1, limit: int = 50, sort_by: str = "name", sort_dir: str = "asc"):
    """High-performance product search with category filtering, branch filtering, and pagination"""
    user = await get_current_user(request)
    tid = user["tenant_id"]

    query = {"tenant_id": tid}
    # Branch filtering: Staff see only their branch, others can filter
    user_branch = user.get("branch_id", "")
    if user["role"] == "STAFF" and user_branch:
        query["branch_id"] = user_branch
    elif branch_id:
        query["branch_id"] = branch_id

    if category_id:
        # Include subcategories
        subcats = await db.categories.find({"tenant_id": tid, "path": category_id}, {"_id": 0, "id": 1}).to_list(500)
        cat_ids = [category_id] + [c["id"] for c in subcats]
        query["category_id"] = {"$in": cat_ids}

    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"sku": {"$regex": q, "$options": "i"}},
            {"barcode": {"$regex": q, "$options": "i"}},
            {"description": {"$regex": q, "$options": "i"}},
            {"category": {"$regex": q, "$options": "i"}},
        ]

    sort_field = sort_by if sort_by in ("name", "price", "stock", "created_at", "category") else "name"
    sort_order = 1 if sort_dir == "asc" else -1

    total = await db.products.count_documents(query)
    skip = (page - 1) * limit
    products = await db.products.find(query, {"_id": 0}).sort(sort_field, sort_order).skip(skip).limit(limit).to_list(limit)

    # Enrich with branch names
    if products:
        branch_ids = list(set(p.get("branch_id", "") for p in products if p.get("branch_id")))
        if branch_ids:
            branches = await db.branches.find({"tenant_id": tid, "id": {"$in": branch_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
            bmap = {b["id"]: b["name"] for b in branches}
            for p in products:
                p["branch_name"] = bmap.get(p.get("branch_id", ""), "")

    return {
        "products": products, "total": total, "page": page,
        "pages": (total + limit - 1) // limit, "query": q
    }

@api_router.get("/search/suggestions")
async def search_suggestions(request: Request, q: str = ""):
    """Auto-suggest categories and products as user types"""
    user = await get_current_user(request)
    if not q or len(q) < 2:
        return {"categories": [], "products": []}

    tid = user["tenant_id"]

    # Category suggestions
    cats = await db.categories.find(
        {"tenant_id": tid, "name": {"$regex": q, "$options": "i"}},
        {"_id": 0, "id": 1, "name": 1, "level": 1, "parent_id": 1}
    ).limit(5).to_list(5)

    # Also match the flat category field
    flat_cats = await db.products.distinct("category", {"tenant_id": tid, "category": {"$regex": q, "$options": "i"}})
    for fc in flat_cats[:5]:
        if fc and not any(c["name"] == fc for c in cats):
            cats.append({"id": "", "name": fc, "level": 0, "type": "flat"})

    # Product suggestions
    products = await db.products.find(
        {"tenant_id": tid, "$or": [
            {"name": {"$regex": q, "$options": "i"}},
            {"sku": {"$regex": q, "$options": "i"}},
            {"barcode": q}
        ]},
        {"_id": 0, "id": 1, "name": 1, "sku": 1, "price": 1, "stock": 1, "category": 1, "branch_id": 1}
    ).limit(8).to_list(8)

    return {"categories": cats[:5], "products": products[:8]}

@api_router.get("/categories/breadcrumb/{category_id}")
async def category_breadcrumb(category_id: str, request: Request):
    """Get breadcrumb trail for a category"""
    user = await get_current_user(request)
    cat = await db.categories.find_one({"id": category_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not cat:
        raise HTTPException(404, "Category not found")

    breadcrumb = []
    path_ids = cat.get("path", []) + [category_id]
    if path_ids:
        ancestors = await db.categories.find(
            {"tenant_id": user["tenant_id"], "id": {"$in": path_ids}},
            {"_id": 0, "id": 1, "name": 1, "level": 1}
        ).to_list(20)
        # Order by level
        ancestors.sort(key=lambda x: x.get("level", 0))
        breadcrumb = ancestors

    return {"breadcrumb": breadcrumb, "current": {"id": cat["id"], "name": cat["name"]}}

# ═══════════════════════════════════════════════════════════════════
#  APP SETUP
# ═══════════════════════════════════════════════════════════════════

app.include_router(api_router)

# CORS: When credentials are used, wildcard "*" is not allowed by browsers.
# Must specify explicit origins.
frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
cors_origins_env = os.environ.get("CORS_ORIGINS", "")
origins = [frontend_url]
if cors_origins_env and cors_origins_env != "*":
    origins.extend([o.strip() for o in cors_origins_env.split(",") if o.strip()])
# Add common local origins for development
origins.extend(["http://localhost:3000", "http://127.0.0.1:3000"])
# Deduplicate
origins = list(set(origins))

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Security Headers Middleware ──────────────────────────────────
@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    # Prevent MIME type sniffing
    response.headers["X-Content-Type-Options"] = "nosniff"
    # Prevent clickjacking
    response.headers["X-Frame-Options"] = "DENY"
    # XSS protection (legacy browsers)
    response.headers["X-XSS-Protection"] = "1; mode=block"
    # Control referrer info leakage
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    # Restrict browser features
    response.headers["Permissions-Policy"] = "camera=(self), microphone=(), geolocation=(), payment=(self)"
    # Content Security Policy — tight but allows our needs
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval'; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: https: blob:; "
        "font-src 'self' data:; "
        "connect-src 'self' https://api.upcitemdb.com https://world.openfoodfacts.org wss: ws:; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'"
    )
    # Force HTTPS
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains; preload"
    # Prevent caching of sensitive responses
    if "/api/auth/" in str(request.url) or "/api/admin/" in str(request.url):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, private"
        response.headers["Pragma"] = "no-cache"
    return response

# ─── Request Fingerprint & CSRF Protection ────────────────────────
CSRF_SAFE_METHODS = {"GET", "HEAD", "OPTIONS"}

@app.middleware("http")
async def csrf_and_fingerprint_middleware(request: Request, call_next):
    # Skip CSRF check for safe methods and mobile scanner endpoints (no auth)
    path = request.url.path
    if request.method in CSRF_SAFE_METHODS or "/scan/mobile/" in path or path == "/api/auth/login" or path == "/api/auth/register":
        return await call_next(request)

    # For state-changing requests, verify the request comes from our frontend
    # by checking Origin/Referer headers against allowed origins
    origin = request.headers.get("origin", "")
    referer = request.headers.get("referer", "")
    # Allow requests from known origins or same-origin (no origin header = same-origin)
    if origin and origin not in origins and not any(origin.startswith(o) for o in origins):
        # Check referer as fallback
        if referer and not any(referer.startswith(o) for o in origins):
            logger.warning(f"CSRF blocked: origin={origin}, referer={referer}, path={path}")
            return JSONResponse(status_code=403, content={"detail": "Request blocked by CSRF protection"})

    return await call_next(request)

# ─── Analytics Tracking Middleware ─────────────────────────────────
import time as _time_mod

ANALYTICS_SKIP_PATHS = {"/api/analytics/realtime", "/api/auth/heartbeat", "/health", "/favicon.ico"}

@app.middleware("http")
async def analytics_tracking_middleware(request: Request, call_next):
    """Track every API call for analytics. Non-blocking — writes in background."""
    path = request.url.path
    # Skip non-API and analytics paths to avoid recursion
    if not path.startswith("/api/") or path in ANALYTICS_SKIP_PATHS:
        return await call_next(request)

    start_time = _time_mod.time()
    response = await call_next(request)
    duration_ms = round((_time_mod.time() - start_time) * 1000, 2)

    # Extract user info from JWT cookie (non-blocking, best-effort)
    user_id = None
    tenant_id = None
    try:
        token = request.cookies.get("access_token")
        if token:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user_id = payload.get("sub")
            tenant_id = payload.get("tenant_id")
    except Exception:
        pass

    # Categorize the endpoint
    endpoint_parts = path.replace("/api/", "").split("/")
    feature = endpoint_parts[0] if endpoint_parts else "unknown"

    # Write analytics asynchronously (don't block response)
    try:
        doc = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "hour": datetime.now(timezone.utc).hour,
            "endpoint": path,
            "method": request.method,
            "feature": feature,
            "status_code": response.status_code,
            "duration_ms": duration_ms,
            "user_id": user_id,
            "tenant_id": tenant_id,
            "ip": request.client.host if request.client else None,
            "user_agent": request.headers.get("user-agent", "")[:200],
        }
        # Fire and forget — don't await
        import asyncio
        asyncio.create_task(db.api_analytics.insert_one(doc))
    except Exception:
        pass  # Never block the response

    return response

@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.users.create_index("tenant_id")
    await db.products.create_index([("tenant_id", 1), ("name", 1)])
    await db.products.create_index([("tenant_id", 1), ("sku", 1)])
    await db.products.create_index([("tenant_id", 1), ("expiry_date", 1)])
    await db.invoices.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.invoices.create_index([("tenant_id", 1), ("customer_id", 1)])
    await db.audit_logs.create_index([("tenant_id", 1), ("timestamp", -1)])
    await db.audit_logs.create_index([("tenant_id", 1), ("user_id", 1)])
    await db.audit_logs.create_index([("tenant_id", 1), ("event_category", 1)])
    await db.login_attempts.create_index("identifier")
    await db.payment_transactions.create_index("session_id")
    await db.mfa_backup_codes.create_index([("user_id", 1), ("used", 1)])
    await db.purchases.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.suppliers.create_index([("tenant_id", 1), ("name", 1)])
    await db.scan_sessions.create_index([("tenant_id", 1), ("user_id", 1), ("type", 1)])
    await db.scan_sessions.create_index("id")
    await db.scanned_barcodes.create_index([("session_id", 1), ("processed", 1)])
    await db.products.create_index([("tenant_id", 1), ("barcode", 1)])
    await db.customers.create_index([("tenant_id", 1), ("name", 1)])
    await db.customers.create_index([("tenant_id", 1), ("phone", 1)])
    await db.customer_transactions.create_index([("customer_id", 1), ("created_at", -1)])
    await db.api_keys.create_index("key_hash")
    await db.api_keys.create_index([("tenant_id", 1)])
    # New indexes for features
    await db.support_tickets.create_index([("tenant_id", 1), ("status", 1)])
    await db.support_tickets.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.ticket_notes.create_index([("ticket_id", 1), ("created_at", 1)])
    await db.access_requests.create_index([("tenant_id", 1), ("status", 1)])
    await db.access_requests.create_index([("admin_id", 1), ("status", 1)])
    await db.temp_access.create_index([("user_id", 1), ("is_active", 1), ("expires_at", 1)])
    await db.temp_access.create_index([("tenant_id", 1)])
    await db.security_alerts.create_index([("tenant_id", 1), ("is_read", 1), ("created_at", -1)])
    await db.security_alerts.create_index([("tenant_id", 1), ("alert_type", 1)])
    # Premium feature indexes
    await db.promo_codes.create_index([("tenant_id", 1), ("code", 1)], unique=True)
    await db.reorder_rules.create_index([("tenant_id", 1), ("product_id", 1)], unique=True)
    await db.notification_templates.create_index([("tenant_id", 1), ("channel", 1)])
    await db.notification_logs.create_index([("tenant_id", 1), ("created_at", -1)])
    # Analytics indexes
    await db.api_analytics.create_index([("tenant_id", 1), ("date", 1)])
    await db.api_analytics.create_index([("tenant_id", 1), ("timestamp", -1)])
    await db.api_analytics.create_index([("date", 1), ("feature", 1)])
    await db.api_analytics.create_index("timestamp", expireAfterSeconds=90 * 86400)  # Auto-delete after 90 days
    await db.advance_orders.create_index([("tenant_id", 1), ("status", 1)])
    await db.advance_orders.create_index([("tenant_id", 1), ("created_at", -1)])

    # Branch & Category indexes
    await db.branches.create_index([("tenant_id", 1), ("code", 1)], unique=True, sparse=True)
    await db.branches.create_index([("tenant_id", 1), ("is_active", 1)])
    await db.categories.create_index([("tenant_id", 1), ("parent_id", 1)])
    await db.categories.create_index([("tenant_id", 1), ("level", 1)])
    await db.products.create_index([("tenant_id", 1), ("branch_id", 1)])
    await db.products.create_index([("tenant_id", 1), ("category_id", 1)])
    await db.products.create_index([("tenant_id", 1), ("expiry_date", 1)])
    await db.invoices.create_index([("tenant_id", 1), ("branch_id", 1), ("created_at", -1)])

    # Transfer requests indexes
    await db.transfer_requests.create_index([("tenant_id", 1), ("status", 1)])
    await db.transfer_requests.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.transfer_requests.create_index([("tenant_id", 1), ("source_branch_id", 1)])
    await db.transfer_requests.create_index([("tenant_id", 1), ("target_branch_id", 1)])

    # Performance: compound indexes for large-scale search
    await db.products.create_index([("tenant_id", 1), ("name", 1)])
    await db.products.create_index([("tenant_id", 1), ("branch_id", 1), ("stock", 1)])
    await db.products.create_index([("tenant_id", 1), ("branch_id", 1), ("category_id", 1)])

    # User branch assignment
    await db.users.create_index([("tenant_id", 1), ("branch_id", 1)])

    # Seed admin user
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@retailsaas.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin@123")
    existing = await db.users.find_one({"email": admin_email})

    if existing is None:
        tenant_id = str(uuid.uuid4())
        await db.tenants.insert_one({
            "id": tenant_id, "shop_name": "Demo Retail Store", "business_type": "general",
            "plan": "premium", "max_users": 999, "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        await db.users.insert_one({
            "email": admin_email, "password_hash": hash_password(admin_password),
            "name": "Admin", "role": "OWNER", "tenant_id": tenant_id,
            "is_active": True, "mfa_enabled": False, "mfa_secret": None,
            "is_platform_admin": False, "known_login_ips": [],
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin seeded: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})

    # Ensure is_platform_admin and known_login_ips fields exist on all users
    await db.users.update_many(
        {"is_platform_admin": {"$exists": False}},
        {"$set": {"is_platform_admin": False}}
    )
    await db.users.update_many(
        {"known_login_ips": {"$exists": False}},
        {"$set": {"known_login_ips": []}}
    )

    # Seed platform admin
    platform_email = os.environ.get("PLATFORM_ADMIN_EMAIL", "platform@retailpro.com")
    platform_password = os.environ.get("PLATFORM_ADMIN_PASSWORD", "Platform@123")
    platform_user = await db.users.find_one({"email": platform_email})

    if platform_user is None:
        # Platform admin doesn't belong to a tenant - use a special system tenant
        sys_tenant = await db.tenants.find_one({"id": "system"})
        if not sys_tenant:
            await db.tenants.insert_one({
                "id": "system", "shop_name": "RetailPro Platform", "business_type": "platform",
                "plan": "premium", "max_users": 999, "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        await db.users.insert_one({
            "email": platform_email, "password_hash": hash_password(platform_password),
            "name": "Platform Admin", "role": "OWNER", "tenant_id": "system",
            "is_active": True, "mfa_enabled": False, "mfa_secret": None,
            "is_platform_admin": True, "known_login_ips": [],
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Platform admin seeded: {platform_email}")
    elif not verify_password(platform_password, platform_user["password_hash"]):
        await db.users.update_one({"email": platform_email}, {"$set": {"password_hash": hash_password(platform_password)}})

    # ── Seed default Admin account (product-side admin) ──
    default_admin_email = "admin@retailpro.com"
    default_admin_password = "AdminRP@123"
    admin_acct = await db.users.find_one({"email": default_admin_email})
    if admin_acct is None:
        await db.users.insert_one({
            "email": default_admin_email, "password_hash": hash_password(default_admin_password),
            "name": "RetailPro Admin", "role": "ADMIN", "tenant_id": "system",
            "is_active": True, "is_admin": True, "is_platform_admin": False,
            "mfa_enabled": False, "mfa_secret": None, "known_login_ips": [],
            "permissions": {"can_view_revenue": True, "can_manage_inventory": True},
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin account seeded: {default_admin_email}")
    elif not verify_password(default_admin_password, admin_acct["password_hash"]):
        await db.users.update_one({"email": default_admin_email}, {"$set": {"password_hash": hash_password(default_admin_password)}})

    creds_path = Path("/app/memory/test_credentials.md")
    creds_path.parent.mkdir(parents=True, exist_ok=True)
    with open(creds_path, "w") as f:
        f.write(f"# Test Credentials\n\n## Admin (Tenant Owner)\n- Email: {admin_email}\n- Password: {admin_password}\n- Role: OWNER\n\n## Platform Admin\n- Email: {platform_email}\n- Password: {platform_password}\n- Role: OWNER + is_platform_admin: true\n\n## Product Admin\n- Email: {default_admin_email}\n- Password: {default_admin_password}\n- Role: ADMIN + is_admin: true\n\n## Auth Endpoints\n- POST /api/auth/login\n- POST /api/auth/register\n- GET /api/auth/me\n- POST /api/auth/logout\n")

    logger.info("App started, indexes created, admin seeded")

@app.on_event("shutdown")
async def shutdown():
    client.close()
